#!/usr/bin/env python3
"""
LLM-Generated Task Prompt Detection Pipeline v0.65
===================================================
Multi-layer stylometric detection pipeline for identifying LLM-generated
or LLM-assisted task prompts in human data collection workflows.
"""

__version__ = '0.65.0'

# ============================================================================
# STANDARD LIBRARY IMPORTS
# ============================================================================

import re
import os
import sys
import json
import math
import zlib
import hashlib
import statistics
import unicodedata
import threading
import argparse
from collections import Counter, defaultdict
from datetime import datetime
from dataclasses import dataclass, field
from typing import List, Tuple, Dict, Optional, FrozenSet

# ============================================================================
# THIRD-PARTY IMPORTS
# ============================================================================

import pandas as pd

"""
Feature detection and optional dependency management.

Centralizes all try/except ImportError blocks so other modules can check
availability flags without repeating import logic.
"""


# ── tkinter ──────────────────────────────────────────────────────────────────
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    HAS_TK = True
except ImportError:
    HAS_TK = False

# ── spaCy: lightweight sentencizer ──────────────────────────────────────────
try:
    import spacy
    from spacy.lang.en import English
    _nlp = English()
    _nlp.add_pipe("sentencizer")
    HAS_SPACY = True
except ImportError:
    HAS_SPACY = False
    print("INFO: spacy not installed. Sentence segmentation will use regex fallback.")
except Exception as e:
    HAS_SPACY = False
    print(f"INFO: spacy sentencizer setup failed ({e}). Using regex fallback.")

# ── ftfy: robust text encoding repair ───────────────────────────────────────
try:
    import ftfy
    HAS_FTFY = True
except ImportError:
    HAS_FTFY = False

# ── sentence-transformers: semantic vector analysis ─────────────────────────
HAS_SEMANTIC = False
_EMBEDDER = None
_AI_CENTROIDS = None
_HUMAN_CENTROIDS = None
_SEMANTIC_INIT_LOCK = threading.Lock()
_SEMANTIC_INIT_DONE = False

_AI_ARCHETYPES = [
    "As an AI language model, I cannot provide personal opinions.",
    "Here is a comprehensive breakdown of the key factors to consider.",
    "To address this challenge, we must consider multiple perspectives.",
    "This thorough analysis demonstrates the critical importance of the topic.",
    "Furthermore, it is essential to note that this approach ensures alignment.",
    "In conclusion, by leveraging these strategies we can achieve optimal results.",
]
_HUMAN_ARCHETYPES = [
    "honestly idk maybe try restarting it lol",
    "so I went ahead and just hacked together a quick script",
    "tbh the whole thing is kinda janky but it works",
    "yeah no that's totally wrong, here's what actually happened",
    "I messed around with it for a bit and got something working",
]

try:
    from sentence_transformers import SentenceTransformer
    from sklearn.metrics.pairwise import cosine_similarity as _cosine_similarity
    import numpy as np
    HAS_SEMANTIC = True
except ImportError:
    pass
except Exception as e:
    print(f"INFO: sentence-transformers setup failed ({e}). Semantic layer disabled.")


def _ensure_semantic():
    """Lazy-initialize sentence-transformers model on first use."""
    global _EMBEDDER, _AI_CENTROIDS, _HUMAN_CENTROIDS, _SEMANTIC_INIT_DONE, HAS_SEMANTIC
    if _SEMANTIC_INIT_DONE or not HAS_SEMANTIC:
        return
    with _SEMANTIC_INIT_LOCK:
        if _SEMANTIC_INIT_DONE:
            return
        try:
            _EMBEDDER = SentenceTransformer('all-MiniLM-L6-v2')
            _AI_CENTROIDS = _EMBEDDER.encode(_AI_ARCHETYPES)
            _HUMAN_CENTROIDS = _EMBEDDER.encode(_HUMAN_ARCHETYPES)
        except Exception as e:
            HAS_SEMANTIC = False
            print(f"INFO: sentence-transformers model loading failed ({e}). Semantic layer disabled.")
        _SEMANTIC_INIT_DONE = True

# ── transformers: local perplexity scoring ──────────────────────────────────
HAS_PERPLEXITY = False
_PPL_MODEL = None
_PPL_TOKENIZER = None
_PPL_INIT_LOCK = threading.Lock()
_PPL_INIT_DONE = False

try:
    from transformers import GPT2LMHeadModel, GPT2TokenizerFast
    import torch as _torch
    HAS_PERPLEXITY = True
except ImportError:
    pass
except Exception as e:
    print(f"INFO: transformers/torch import failed ({e}). Perplexity scoring disabled.")


def _ensure_perplexity():
    """Lazy-initialize distilgpt2 model on first use."""
    global _PPL_MODEL, _PPL_TOKENIZER, _PPL_INIT_DONE, HAS_PERPLEXITY
    if _PPL_INIT_DONE or not HAS_PERPLEXITY:
        return
    with _PPL_INIT_LOCK:
        if _PPL_INIT_DONE:
            return
        try:
            _PPL_MODEL_ID = 'distilgpt2'
            _PPL_MODEL = GPT2LMHeadModel.from_pretrained(_PPL_MODEL_ID)
            _PPL_TOKENIZER = GPT2TokenizerFast.from_pretrained(_PPL_MODEL_ID)
            _PPL_MODEL.eval()
        except Exception as e:
            HAS_PERPLEXITY = False
            print(f"INFO: distilgpt2 model loading failed ({e}). Perplexity scoring disabled.")
        _PPL_INIT_DONE = True

# ── pypdf: PDF text extraction ──────────────────────────────────────────────
try:
    from pypdf import PdfReader
    HAS_PYPDF = True
except ImportError:
    HAS_PYPDF = False


# ==============================================================================
# TEXT UTILITIES
# ==============================================================================
"""Shared text utilities used across multiple modules."""


# Top-50 English function words (closed class, highly stable across registers)
ENGLISH_FUNCTION_WORDS = frozenset([
    'the', 'a', 'an', 'is', 'are', 'was', 'were', 'am', 'be', 'been',
    'being', 'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would',
    'shall', 'should', 'may', 'might', 'can', 'could', 'must',
    'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by', 'from', 'as',
    'into', 'about', 'between', 'through', 'after', 'before',
    'and', 'or', 'but', 'not', 'if', 'that', 'this', 'it', 'he', 'she',
    'they', 'we', 'i', 'you', 'my', 'your', 'his', 'her', 'its', 'our',
    'their', 'who', 'which', 'what', 'there',
])


def get_sentences(text):
    """Segment text into sentences using spacy sentencizer or regex fallback."""
    if HAS_SPACY:
        doc = _nlp(text)
        return [s.text for s in doc.sents]
    else:
        sents = re.split(r'(?<=[.!?])\s+', text)
        return [s for s in sents if s.strip()]


# ==============================================================================
# TEXT NORMALIZATION
# ==============================================================================
"""
Text normalization pre-pass.

Neutralizes cheap evasion attacks before all detection layers.
Ref: RAID benchmark (Dugan et al. 2024) — formatting perturbations,
     homoglyphs, and spacing attacks degrade metric-based detectors.
Ref: MGTBench (He et al. 2023) — paraphrasing sensitivity in rule-based
     detectors.
"""


# Common homoglyph mappings: visually similar Unicode -> ASCII
_HOMOGLYPH_MAP = str.maketrans({
    '\u0410': 'A', '\u0412': 'B', '\u0421': 'C', '\u0415': 'E',
    '\u041d': 'H', '\u041a': 'K', '\u041c': 'M', '\u041e': 'O',
    '\u0420': 'P', '\u0422': 'T', '\u0425': 'X',
    '\u0430': 'a', '\u0435': 'e', '\u043e': 'o', '\u0440': 'p',
    '\u0441': 'c', '\u0443': 'y', '\u0445': 'x',
    '\u2018': "'", '\u2019': "'", '\u201a': "'",
    '\u201c': '"', '\u201d': '"', '\u201e': '"',
    '\u2032': "'", '\u2033': '"',
    '\u2014': '--', '\u2013': '-', '\u2012': '-',
    '\u2026': '...', '\u22ef': '...',
    '\uff01': '!', '\uff1f': '?', '\uff0c': ',', '\uff0e': '.',
    '\uff1a': ':', '\uff1b': ';',
})

# Zero-width and invisible characters to strip
_INVISIBLE_RE = re.compile(
    '[\u200b\u200c\u200d\u200e\u200f'
    '\u2060\u2061\u2062\u2063\u2064'
    '\ufeff'
    '\u00ad'
    '\u034f'
    '\u180e'
    '\u2028\u2029'
    ']'
)

# Inter-character spacing: "l i k e  t h i s"
_INTERSPACED_RE = re.compile(r'(?<!\w)(\w) (\w) (\w) (\w)(?!\w)')


def normalize_text(text):
    """Normalize text to neutralize common evasion attacks.

    Returns (normalized_text, delta_report).
    """
    original = text
    original_len = max(len(text), 1)
    changes = 0
    ftfy_applied = False

    # 0. ftfy encoding repair
    if HAS_FTFY:
        pre_ftfy = text
        text = ftfy.fix_text(text)
        ftfy_changes = sum(1 for a, b in zip(pre_ftfy, text) if a != b)
        ftfy_changes += abs(len(pre_ftfy) - len(text))
        changes += ftfy_changes
        ftfy_applied = ftfy_changes > 0

    # 1. Strip invisible/zero-width characters
    invisible_count = len(_INVISIBLE_RE.findall(text))
    text = _INVISIBLE_RE.sub('', text)
    changes += invisible_count

    # 2. NFKC normalization
    pre_nfkc = text
    text = unicodedata.normalize('NFKC', text)
    nfkc_changes = sum(1 for a, b in zip(pre_nfkc, text) if a != b)
    changes += nfkc_changes

    # 3. Homoglyph folding
    pre_homoglyph = text
    text = text.translate(_HOMOGLYPH_MAP)
    homoglyph_count = sum(1 for a, b in zip(pre_homoglyph, text) if a != b)
    changes += homoglyph_count

    # 4. Inter-character spacing collapse
    interspacing_spans = len(_INTERSPACED_RE.findall(text))
    if interspacing_spans > 0:
        prev = None
        while prev != text:
            prev = text
            text = re.sub(r'(?<!\w)(\w) (?=\w(?:\s\w)*(?!\w))', r'\1', text)
        spacing_changes = len(original) - len(text)
        changes += max(spacing_changes, 0)

    # 5. Whitespace collapse
    pre_ws = text
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    text = text.strip()
    ws_collapsed = (pre_ws != text)
    if ws_collapsed:
        changes += abs(len(pre_ws) - len(text))

    # 6. Control character stripping
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)

    obfuscation_delta = changes / original_len

    return text, {
        'obfuscation_delta': round(obfuscation_delta, 4),
        'invisible_chars': invisible_count,
        'homoglyphs': homoglyph_count,
        'interspacing_spans': interspacing_spans,
        'whitespace_collapsed': ws_collapsed,
        'ftfy_applied': ftfy_applied,
    }


# ==============================================================================
# LANGUAGE SUPPORT GATE
# ==============================================================================
"""
Fairness / Language support gate.

Caps severity if text is outside the validated English-prose envelope.
Ref: Liang et al. (2023) "GPT Detectors Are Biased Against Non-Native
     English Writers"
Ref: Wang et al. (2023) "M4 -- multilingual detection remains harder."
"""



def check_language_support(text, word_count=None):
    """Assess whether text is within the validated English-prose envelope.

    Returns dict with support_level, function_word_coverage, non_latin_ratio, reason.
    """
    words = text.lower().split()
    if word_count is None:
        word_count = len(words)

    if word_count < 30:
        return {
            'support_level': 'REVIEW',
            'function_word_coverage': 0.0,
            'non_latin_ratio': 0.0,
            'reason': 'Text too short for reliable English detection',
        }

    fw_count = sum(1 for w in words if w in ENGLISH_FUNCTION_WORDS)
    fw_coverage = fw_count / max(word_count, 1)

    alpha_chars = [c for c in text if c.isalpha()]
    n_alpha = max(len(alpha_chars), 1)
    non_latin = sum(1 for c in alpha_chars
                    if unicodedata.category(c).startswith('L')
                    and not ('\u0041' <= c <= '\u007a' or '\u00c0' <= c <= '\u024f'))
    non_latin_ratio = non_latin / n_alpha

    if non_latin_ratio > 0.30:
        level = 'UNSUPPORTED'
        reason = f'High non-Latin script content ({non_latin_ratio:.0%})'
    elif fw_coverage < 0.08:
        level = 'UNSUPPORTED'
        reason = f'Very low English function-word coverage ({fw_coverage:.0%})'
    elif fw_coverage < 0.12:
        level = 'REVIEW'
        reason = f'Low English function-word coverage ({fw_coverage:.0%}) -- possible non-native or non-English text'
    elif non_latin_ratio > 0.10:
        level = 'REVIEW'
        reason = f'Mixed-script content ({non_latin_ratio:.0%} non-Latin)'
    else:
        level = 'SUPPORTED'
        reason = 'Text within validated English-prose envelope'

    return {
        'support_level': level,
        'function_word_coverage': round(fw_coverage, 4),
        'non_latin_ratio': round(non_latin_ratio, 4),
        'reason': reason,
    }


# ==============================================================================
# FILE I/O
# ==============================================================================
"""File loaders for xlsx, csv, and pdf input."""



def load_xlsx(filepath, sheet=None, prompt_col='prompt', id_col='task_id',
              occ_col='occupation', attempter_col='attempter_name', stage_col='pipeline_stage_name'):
    """Load tasks from an xlsx file. Returns list of dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)

    if sheet:
        ws = wb[sheet]
    else:
        for name in ['FullTaskX', 'Full Task Connected', 'Claim Sheet', 'Sample List']:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        else:
            ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip().lower() if h else '' for h in rows[0]]

    def find_col(candidates):
        for c in candidates:
            cl = c.lower()
            for i, h in enumerate(headers):
                if cl == h:
                    return i
        for c in candidates:
            cl = c.lower()
            if cl == 'id':
                continue
            for i, h in enumerate(headers):
                if cl in h:
                    return i
        return None

    prompt_idx = find_col([prompt_col, 'prompt', 'text', 'content'])
    id_idx = find_col([id_col, 'task_id', 'id'])
    occ_idx = find_col([occ_col, 'occupation', 'occ'])
    att_idx = find_col([attempter_col, 'attempter', 'claimed_by', 'claimed by'])
    stage_idx = find_col([stage_col, 'stage', 'pipeline_stage'])

    if prompt_idx is None:
        print(f"ERROR: Could not find prompt column. Headers: {headers}")
        return []

    tasks = []
    for row in rows[1:]:
        if not row or len(row) <= prompt_idx:
            continue
        prompt = str(row[prompt_idx]).strip() if row[prompt_idx] else ''
        if len(prompt) < 50:
            continue

        tasks.append({
            'prompt': prompt,
            'task_id': str(row[id_idx])[:20] if id_idx is not None and row[id_idx] else '',
            'occupation': str(row[occ_idx]) if occ_idx is not None and row[occ_idx] else '',
            'attempter': str(row[att_idx]) if att_idx is not None and row[att_idx] else '',
            'stage': str(row[stage_idx]) if stage_idx is not None and row[stage_idx] else '',
        })

    return tasks


def load_csv(filepath, prompt_col='prompt'):
    """Load tasks from CSV."""
    df = pd.read_csv(filepath)
    df = df.fillna('')

    col_map = {c.lower().strip(): c for c in df.columns}

    def resolve_col(*candidates):
        for c in candidates:
            key = c.lower().strip()
            if key in col_map:
                return col_map[key]
        for c in candidates:
            key = c.lower().strip()
            if key == 'id':
                continue
            for mapped_key, actual in col_map.items():
                if key in mapped_key:
                    return actual
        return None

    prompt_actual = resolve_col(prompt_col, 'prompt', 'text', 'content')
    id_actual = resolve_col('task_id', 'id')
    occ_actual = resolve_col('occupation', 'occ')
    att_actual = resolve_col('attempter_name', 'attempter', 'claimed_by')
    stage_actual = resolve_col('pipeline_stage_name', 'stage')

    if prompt_actual is None:
        print(f"ERROR: Could not find prompt column. Columns: {list(df.columns)}")
        return []

    tasks = []
    for _, row in df.iterrows():
        prompt = str(row.get(prompt_actual, ''))
        if len(prompt) < 50:
            continue
        tasks.append({
            'prompt': prompt,
            'task_id': str(row.get(id_actual, ''))[:20] if id_actual else '',
            'occupation': str(row.get(occ_actual, '')) if occ_actual else '',
            'attempter': str(row.get(att_actual, '')) if att_actual else '',
            'stage': str(row.get(stage_actual, '')) if stage_actual else '',
        })
    return tasks


def load_pdf(filepath):
    """Load text from PDF file. Each page becomes a separate task."""
    if not HAS_PYPDF:
        print("ERROR: pypdf not installed. Run: pip install pypdf")
        return []

    reader = PdfReader(filepath)
    tasks = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and len(text.strip()) >= 50:
            tasks.append({
                'prompt': text.strip(),
                'task_id': f"page_{i+1}",
                'occupation': '',
                'attempter': '',
                'stage': '',
            })

    if not tasks:
        full_text = '\n'.join(
            page.extract_text() for page in reader.pages
            if page.extract_text()
        ).strip()
        if len(full_text) >= 50:
            tasks.append({
                'prompt': full_text,
                'task_id': 'full_document',
                'occupation': '',
                'attempter': '',
                'stage': '',
            })

    return tasks


# ==============================================================================
# CONFORMAL CALIBRATION
# Conformal calibration from labeled baseline data.  Ref: Vovk et al. (2005), 
# ==============================================================================
_CALIBRATION_ALPHAS = [0.01, 0.05, 0.10]


def calibrate_from_baselines(jsonl_path):
    """Build calibration tables from labeled baseline data."""
    records = []
    with open(jsonl_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    rec = json.loads(line)
                    if rec.get('ground_truth') == 'human':
                        records.append(rec)
                except json.JSONDecodeError:
                    continue

    if len(records) < 20:
        return None

    nc_scores = [1.0 - float(r.get('confidence', 0)) for r in records]
    nc_scores.sort()

    global_cal = {}
    for alpha in _CALIBRATION_ALPHAS:
        idx = int(math.ceil((1 - alpha) * (len(nc_scores) + 1))) - 1
        idx = max(0, min(idx, len(nc_scores) - 1))
        global_cal[alpha] = nc_scores[idx]

    strata = defaultdict(list)
    for r in records:
        domain = r.get('domain', 'unknown') or 'unknown'
        length_bin = r.get('length_bin', 'unknown') or 'unknown'
        conf = float(r.get('confidence', 0))
        strata[(domain, length_bin)].append(1.0 - conf)

    strata_cal = {}
    strata_counts = {}
    for key, scores in strata.items():
        scores.sort()
        strata_counts[key] = len(scores)
        if len(scores) >= 10:
            strata_cal[key] = {}
            for alpha in _CALIBRATION_ALPHAS:
                idx = int(math.ceil((1 - alpha) * (len(scores) + 1))) - 1
                idx = max(0, min(idx, len(scores) - 1))
                strata_cal[key][alpha] = scores[idx]

    return {
        'global': global_cal,
        'strata': strata_cal,
        'n_calibration': len(records),
        'strata_counts': {f"{k[0]}_{k[1]}": v for k, v in strata_counts.items()},
    }


def apply_calibration(confidence, cal_table, domain=None, length_bin=None):
    """Apply conformal calibration to a raw confidence score."""
    if cal_table is None:
        return {
            'raw_confidence': confidence,
            'calibrated_confidence': confidence,
            'confidence_quantile': None,
            'stratum_used': 'uncalibrated',
        }

    nc_score = 1.0 - confidence

    stratum_key = (domain or 'unknown', length_bin or 'unknown')
    if stratum_key in cal_table.get('strata', {}):
        cal = cal_table['strata'][stratum_key]
        stratum_label = f"{stratum_key[0]}_{stratum_key[1]}"
    else:
        cal = cal_table.get('global', {})
        stratum_label = 'global'

    if nc_score <= cal.get(0.01, 0):
        confidence_quantile = 1.0
    elif nc_score <= cal.get(0.05, 0):
        confidence_quantile = 0.10
    elif nc_score <= cal.get(0.10, 0):
        confidence_quantile = 0.05
    else:
        confidence_quantile = 0.01

    alpha_05 = cal.get(0.05, 0.5)
    if nc_score > alpha_05:
        calibrated = min(confidence * 1.15, 0.99)
    elif nc_score < cal.get(0.10, 0.5):
        calibrated = confidence * 0.75
    else:
        calibrated = confidence

    return {
        'raw_confidence': round(confidence, 4),
        'calibrated_confidence': round(calibrated, 4),
        'confidence_quantile': round(confidence_quantile, 4) if confidence_quantile is not None else None,
        'stratum_used': stratum_label,
    }


def save_calibration(cal_table, path):
    """Save calibration table to JSON."""
    serializable = {
        'global': cal_table['global'],
        'strata': {f"{k[0]}|{k[1]}": v for k, v in cal_table.get('strata', {}).items()}
                  if isinstance(list(cal_table.get('strata', {}).keys() or [('',)])[0], tuple)
                  else cal_table.get('strata', {}),
        'n_calibration': cal_table['n_calibration'],
        'strata_counts': cal_table.get('strata_counts', {}),
    }
    with open(path, 'w') as f:
        json.dump(serializable, f, indent=2)
    print(f"  Calibration table saved to {path} ({cal_table['n_calibration']} records)")


def load_calibration(path):
    """Load calibration table from JSON."""
    with open(path, 'r') as f:
        raw = json.load(f)

    strata = {}
    for k, v in raw.get('strata', {}).items():
        parts = k.split('|')
        if len(parts) == 2:
            strata[(parts[0], parts[1])] = {float(ak): av for ak, av in v.items()}
        else:
            strata[k] = v

    return {
        'global': {float(k): v for k, v in raw.get('global', {}).items()},
        'strata': strata,
        'n_calibration': raw.get('n_calibration', 0),
        'strata_counts': raw.get('strata_counts', {}),
    }

# ==============================================================================
# BASELINE COLLECTION
# Labeled data collection and baseline analysis.
# ==============================================================================
_BASELINE_FIELDS = [
    'task_id', 'occupation', 'attempter', 'word_count', 'determination',
    'confidence', 'preamble_score', 'prompt_signature_composite', 'prompt_signature_cfd',
    'prompt_signature_mfsr', 'prompt_signature_framing', 'prompt_signature_must_rate',
    'prompt_signature_distinct_frames',
    'instruction_density_idi', 'instruction_density_imperatives', 'instruction_density_conditionals',
    'voice_dissonance_voice_score', 'voice_dissonance_spec_score', 'voice_dissonance_vsd',
    'voice_dissonance_voice_gated', 'voice_dissonance_hedges', 'voice_dissonance_casual_markers',
    'voice_dissonance_misspellings', 'ssi_triggered',
    'self_similarity_nssi_score', 'self_similarity_nssi_signals', 'self_similarity_determination',
    'continuation_bscore', 'continuation_determination',
    'self_similarity_sent_length_cv', 'self_similarity_comp_ratio', 'self_similarity_hapax_ratio',
    'norm_obfuscation_delta', 'norm_invisible_chars', 'norm_homoglyphs',
    'lang_support_level', 'lang_fw_coverage', 'lang_non_latin_ratio',
    'ground_truth', 'language', 'domain', 'mode',
    'window_max_score', 'window_mean_score', 'window_variance',
    'window_hot_span', 'window_mixed_signal',
    'stylo_fw_ratio', 'stylo_sent_dispersion', 'stylo_ttr',
    'perplexity_surprisal_variance', 'perplexity_volatility_decay', 'perplexity_n_tokens',
    'calibrated_confidence', 'confidence_quantile', 'calibration_stratum',
    'pack_constraint_score', 'pack_exec_spec_score', 'pack_schema_score',
    'pack_active_families', 'pack_prompt_boost', 'pack_idi_boost',
    # v0.65 additions
    'self_similarity_shuffled_comp_ratio', 'self_similarity_structural_compression_delta',
    'continuation_composite_stability', 'continuation_composite_variance',
    'continuation_improvement_rate', 'continuation_ncd_matrix_mean', 'continuation_ncd_matrix_variance',
    'perplexity_comp_ratio', 'perplexity_zlib_normalized_ppl', 'perplexity_comp_ppl_ratio',
    'window_fw_trajectory_cv', 'window_comp_trajectory_mean', 'window_comp_trajectory_cv',
    'tocsin_cohesiveness', 'tocsin_determination', 'tocsin_confidence',
    'surprisal_trajectory_cv', 'surprisal_stationarity',
    # v0.65 addendum: similarity feedback
    'similarity_upgraded',
]


def collect_baselines(results, output_path):
    """Append scored results to JSONL file for baseline accumulation."""
    timestamp = datetime.now().isoformat()
    n_written = 0

    with open(output_path, 'a') as f:
        for r in results:
            record = {k: r.get(k) for k in _BASELINE_FIELDS}
            record['_timestamp'] = timestamp
            record['_version'] = 'v0.65'
            wc = r.get('word_count', 0)
            if wc < 100:
                record['length_bin'] = 'short'
            elif wc < 300:
                record['length_bin'] = 'medium'
            elif wc < 800:
                record['length_bin'] = 'long'
            else:
                record['length_bin'] = 'very_long'
            f.write(json.dumps(record) + '\n')
            n_written += 1

    print(f"\n  Baseline data: {n_written} records appended to {output_path}")
    return n_written


def analyze_baselines(jsonl_path, output_csv=None):
    """Read accumulated baseline JSONL and compute per-occupation percentile tables."""
    records = []
    with open(jsonl_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    continue

    if not records:
        print(f"No records found in {jsonl_path}")
        return

    df = pd.DataFrame(records)
    print(f"\n{'='*90}")
    print(f"  BASELINE ANALYSIS -- {len(df)} records from {jsonl_path}")
    print(f"{'='*90}")

    det_counts = df['determination'].value_counts()
    print(f"\n  Overall distribution:")
    for det in ['RED', 'AMBER', 'YELLOW', 'GREEN']:
        ct = det_counts.get(det, 0)
        pct = ct / len(df) * 100
        print(f"    {det:>8}: {ct:>5} ({pct:.1f}%)")

    metrics = ['prompt_signature_composite', 'prompt_signature_cfd', 'prompt_signature_mfsr',
               'prompt_signature_must_rate', 'instruction_density_idi',
               'voice_dissonance_spec_score', 'voice_dissonance_voice_score', 'voice_dissonance_vsd',
               'self_similarity_nssi_score', 'self_similarity_comp_ratio', 'self_similarity_hapax_ratio',
               'self_similarity_sent_length_cv',
               'norm_obfuscation_delta', 'lang_fw_coverage', 'word_count']
    percentiles = [0.25, 0.50, 0.75, 0.90, 0.95, 0.99]

    occupations = sorted(df['occupation'].dropna().unique())
    if not occupations:
        occupations = ['(all)']
        df['occupation'] = '(all)'

    all_rows = []

    for occ in occupations:
        occ_df = df[df['occupation'] == occ]
        if len(occ_df) < 5:
            continue

        print(f"\n  -- {occ} (n={len(occ_df)}) --")
        det_pcts = occ_df['determination'].value_counts()
        flags = det_pcts.get('RED', 0) + det_pcts.get('AMBER', 0)
        flag_rate = flags / len(occ_df) * 100
        print(f"     Flag rate: {flag_rate:.1f}% ({flags}/{len(occ_df)})")

        for m in metrics:
            if m not in occ_df.columns:
                continue
            vals = pd.to_numeric(occ_df[m], errors='coerce').dropna()
            if len(vals) < 3:
                continue

            pct_vals = vals.quantile(percentiles).to_dict()
            row = {'occupation': occ, 'metric': m, 'n': len(vals),
                   'mean': vals.mean(), 'std': vals.std()}
            row.update({f'p{int(k*100)}': v for k, v in pct_vals.items()})
            all_rows.append(row)

            p50 = pct_vals.get(0.50, 0)
            p90 = pct_vals.get(0.90, 0)
            p99 = pct_vals.get(0.99, 0)
            print(f"     {m:40s}  p50={p50:7.2f}  p90={p90:7.2f}  p99={p99:7.2f}  mean={vals.mean():7.2f}")

    if output_csv and all_rows:
        baseline_df = pd.DataFrame(all_rows)
        baseline_df.to_csv(output_csv, index=False)
        print(f"\n  Baseline percentiles written to {output_csv}")

    if 'ground_truth' in df.columns:
        labeled = df[df['ground_truth'].isin(['human', 'ai'])].copy()
        if len(labeled) >= 20:
            n_human = (labeled['ground_truth'] == 'human').sum()
            n_ai = (labeled['ground_truth'] == 'ai').sum()
            print(f"\n  -- TPR @ FPR (n={len(labeled)}: {n_human} human, {n_ai} AI) --")

            if n_human >= 5 and n_ai >= 5:
                scores = pd.to_numeric(labeled['confidence'], errors='coerce').fillna(0)
                labels = (labeled['ground_truth'] == 'ai').astype(int)

                thresholds = sorted(scores.unique(), reverse=True)
                for target_fpr, label in [(0.01, '1%'), (0.05, '5%'), (0.10, '10%')]:
                    best_tpr = 0.0
                    best_thresh = 1.0
                    for t in thresholds:
                        predicted_pos = (scores >= t)
                        fp = ((predicted_pos) & (labels == 0)).sum()
                        tp = ((predicted_pos) & (labels == 1)).sum()
                        fpr = fp / max(n_human, 1)
                        tpr = tp / max(n_ai, 1)
                        if fpr <= target_fpr and tpr > best_tpr:
                            best_tpr = tpr
                            best_thresh = t
                    print(f"     TPR @ {label:>3} FPR: {best_tpr:.1%}  (threshold={best_thresh:.3f})")

                for gt_label in ['human', 'ai']:
                    subset = labeled[labeled['ground_truth'] == gt_label]
                    flagged = subset['determination'].isin(['RED', 'AMBER']).sum()
                    rate = flagged / max(len(subset), 1) * 100
                    print(f"     Flag rate ({gt_label:>5}): {rate:.1f}% ({flagged}/{len(subset)})")

    if 'domain' in df.columns and 'length_bin' in df.columns:
        df['_stratum'] = df['domain'].fillna('unknown').astype(str) + 'x' + df['length_bin'].fillna('unknown').astype(str)
        strata = df['_stratum'].unique()
        if len(strata) > 1:
            print(f"\n  -- STRATIFIED FLAG RATES (domain x length_bin) --")
            stratum_rates = {}
            for s in sorted(strata):
                s_df = df[df['_stratum'] == s]
                if len(s_df) < 3:
                    continue
                flagged = s_df['determination'].isin(['RED', 'AMBER']).sum()
                rate = flagged / len(s_df) * 100
                stratum_rates[s] = rate
                print(f"     {s:30s}  n={len(s_df):>4}  flag_rate={rate:5.1f}%")

            if stratum_rates:
                rates = list(stratum_rates.values())
                max_rate = max(rates)
                min_rate = min(rates)
                if max_rate - min_rate > 20:
                    print(f"\n  WARNING: Flag rate disparity across strata")
                    print(f"     Range: {min_rate:.1f}% -- {max_rate:.1f}% (delta={max_rate - min_rate:.1f}pp)")

    return all_rows

# ==============================================================================
# CROSS-SUBMISSION SIMILARITY
# Cross-submission similarity analysis.
# ==============================================================================
def _word_shingles(text, k=3):
    words = re.findall(r'\w+', text.lower())
    if len(words) < k:
        return {tuple(words)} if words else set()
    return set(tuple(words[i:i+k]) for i in range(len(words) - k + 1))


def _jaccard(a, b):
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


_STRUCT_FEATURES = [
    # Original v0.53 features
    'prompt_signature_composite', 'prompt_signature_cfd', 'prompt_signature_mfsr',
    'prompt_signature_must_rate', 'instruction_density_idi',
    'voice_dissonance_spec_score', 'voice_dissonance_voice_score',
    # v0.58+ additions
    'self_similarity_nssi_score', 'self_similarity_comp_ratio',
    'self_similarity_hapax_ratio', 'self_similarity_sent_length_cv',
    'window_max_score', 'window_mean_score',
    'stylo_fw_ratio', 'stylo_ttr', 'stylo_sent_dispersion',
]


def _structural_similarity(r1, r2):
    diff_sq = sum((r1.get(f, 0) - r2.get(f, 0)) ** 2 for f in _STRUCT_FEATURES)
    return 1.0 / (1.0 + math.sqrt(diff_sq))


def _adaptive_thresholds(group, shingle_cache, jaccard_threshold, struct_threshold):
    """Compute adaptive similarity thresholds from occupation group distribution.

    Uses median + 2*std of all pairwise scores within the group. Falls back
    to flat thresholds when fewer than 4 results are available.
    """
    if len(group) < 4:
        return jaccard_threshold, struct_threshold

    jac_vals = []
    struct_vals = []
    for i in range(len(group)):
        for j in range(i + 1, len(group)):
            tid_a = group[i].get('task_id', '')
            tid_b = group[j].get('task_id', '')
            jac_vals.append(_jaccard(
                shingle_cache.get(tid_a, set()),
                shingle_cache.get(tid_b, set()),
            ))
            struct_vals.append(_structural_similarity(group[i], group[j]))

    if len(jac_vals) < 3:
        return jaccard_threshold, struct_threshold

    med_jac = statistics.median(jac_vals)
    std_jac = statistics.stdev(jac_vals) if len(jac_vals) >= 2 else 0.0
    med_struct = statistics.median(struct_vals)
    std_struct = statistics.stdev(struct_vals) if len(struct_vals) >= 2 else 0.0

    adj_jac = max(med_jac + 2 * std_jac, 0.15)
    adj_struct = max(med_struct + 2 * std_struct, 0.50)

    # Never exceed user-specified thresholds
    adj_jac = min(adj_jac, jaccard_threshold)
    adj_struct = min(adj_struct, struct_threshold)

    return adj_jac, adj_struct


def _minhash_signature(shingles, n_hashes=128):
    """Compute MinHash signature for a set of shingles (pure Python, zero deps)."""
    sig = [float('inf')] * n_hashes
    for shingle in shingles:
        s = ' '.join(shingle) if isinstance(shingle, tuple) else str(shingle)
        for i in range(n_hashes):
            h = int(hashlib.md5(f"{i}:{s}".encode()).hexdigest(), 16)
            sig[i] = min(sig[i], h)
    return sig


def _minhash_jaccard(sig_a, sig_b):
    """Estimate Jaccard similarity from two MinHash signatures."""
    if not sig_a or not sig_b or len(sig_a) != len(sig_b):
        return 0.0
    return sum(a == b for a, b in zip(sig_a, sig_b)) / len(sig_a)


def _load_minhash_store(path):
    """Load MinHash fingerprint store from JSONL file."""
    entries = []
    if not os.path.exists(path):
        return entries
    with open(path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                entries.append(json.loads(line))
    return entries


def _store_minhash(entries, path):
    """Append MinHash fingerprint entries to JSONL file."""
    with open(path, 'a') as f:
        for entry in entries:
            f.write(json.dumps(entry) + '\n')


def _factor_instructions(shingles, instruction_shingles):
    """Remove instruction-derived shingles from a submission's shingle set."""
    if not instruction_shingles:
        return shingles
    return shingles - instruction_shingles


def analyze_similarity(results, text_map, jaccard_threshold=0.40, struct_threshold=0.90,
                       semantic=False, instruction_shingles=None, similarity_store_path=None):
    """Analyze cross-submission similarity within occupation groups.

    Supports adaptive thresholds (FEAT 11), semantic embeddings (FEAT 12),
    cross-batch MinHash store (FEAT 14), and instruction factoring (FEAT 15).
    """
    by_occ = defaultdict(list)
    for r in results:
        occ = r.get('occupation', '(unknown)')
        by_occ[occ].append(r)

    # Build shingle cache with optional instruction factoring (FEAT 15)
    shingle_cache = {}
    for tid, text in text_map.items():
        s = _word_shingles(text)
        if instruction_shingles:
            s = _factor_instructions(s, instruction_shingles)
        shingle_cache[tid] = s

    # Build embedding cache for semantic similarity (FEAT 12)
    embedding_cache = {}
    if semantic and HAS_SEMANTIC:
        _ensure_semantic()
        all_tids = [r.get('task_id', '') for r in results
                    if r.get('task_id', '') in text_map]
        all_texts = [text_map[tid] for tid in all_tids]
        if all_texts:
            all_vecs = _EMBEDDER.encode(all_texts)
            for tid, vec in zip(all_tids, all_vecs):
                embedding_cache[tid] = vec

    flagged_pairs = []

    for occ, group in by_occ.items():
        if len(group) < 2:
            continue

        # FEAT 11: adaptive thresholds per occupation group
        jac_thresh, str_thresh = _adaptive_thresholds(
            group, shingle_cache, jaccard_threshold, struct_threshold)

        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                r_a, r_b = group[i], group[j]
                att_a = r_a.get('attempter', '').strip().lower()
                att_b = r_b.get('attempter', '').strip().lower()

                if att_a and att_b and att_a == att_b:
                    continue

                tid_a = r_a.get('task_id', '')
                tid_b = r_b.get('task_id', '')

                jac = _jaccard(
                    shingle_cache.get(tid_a, set()),
                    shingle_cache.get(tid_b, set())
                )
                struct = _structural_similarity(r_a, r_b)

                # FEAT 12: semantic embedding similarity
                sem = 0.0
                if semantic and tid_a in embedding_cache and tid_b in embedding_cache:
                    sem = float(_cosine_similarity(
                        [embedding_cache[tid_a]], [embedding_cache[tid_b]])[0][0])

                flags = []
                if jac >= jac_thresh:
                    flags.append('text')
                if struct >= str_thresh:
                    flags.append('structural')
                if sem >= 0.85:
                    flags.append('semantic')

                if flags:
                    flagged_pairs.append({
                        'id_a': tid_a,
                        'id_b': tid_b,
                        'attempter_a': r_a.get('attempter', ''),
                        'attempter_b': r_b.get('attempter', ''),
                        'occupation': occ,
                        'jaccard': jac,
                        'structural': struct,
                        'semantic': sem,
                        'flag_type': '+'.join(flags),
                        'det_a': r_a['determination'],
                        'det_b': r_b['determination'],
                        'adaptive_jac_threshold': jac_thresh,
                        'adaptive_struct_threshold': str_thresh,
                    })

    # FEAT 14: cross-batch MinHash store
    if similarity_store_path:
        stored = _load_minhash_store(similarity_store_path)

        # Compare current batch against stored fingerprints
        for r in results:
            tid = r.get('task_id', '')
            occ = r.get('occupation', '(unknown)')
            att = r.get('attempter', '').strip().lower()
            shingles = shingle_cache.get(tid, set())
            if not shingles:
                continue
            sig = _minhash_signature(shingles)
            for entry in stored:
                if entry.get('occupation', '') != occ:
                    continue
                stored_att = entry.get('attempter', '').strip().lower()
                if stored_att and att and stored_att == att:
                    continue
                est_jac = _minhash_jaccard(sig, entry.get('minhash', []))
                if est_jac >= max(jaccard_threshold * 0.8, 0.12):
                    flagged_pairs.append({
                        'id_a': tid,
                        'id_b': entry.get('task_id', ''),
                        'attempter_a': r.get('attempter', ''),
                        'attempter_b': entry.get('attempter', ''),
                        'occupation': occ,
                        'jaccard': est_jac,
                        'structural': 0.0,
                        'semantic': 0.0,
                        'flag_type': 'cross-batch',
                        'det_a': r.get('determination', ''),
                        'det_b': entry.get('determination', ''),
                        'adaptive_jac_threshold': jaccard_threshold,
                        'adaptive_struct_threshold': struct_threshold,
                    })

        # Store current batch fingerprints
        from datetime import datetime as _dt
        new_entries = []
        for r in results:
            tid = r.get('task_id', '')
            shingles = shingle_cache.get(tid, set())
            if not shingles:
                continue
            new_entries.append({
                'task_id': tid,
                'attempter': r.get('attempter', ''),
                'occupation': r.get('occupation', '(unknown)'),
                'determination': r.get('determination', ''),
                'minhash': _minhash_signature(shingles),
                'timestamp': _dt.now().isoformat(),
            })
        if new_entries:
            _store_minhash(new_entries, similarity_store_path)

    flagged_pairs.sort(key=lambda p: p['jaccard'], reverse=True)
    return flagged_pairs


def apply_similarity_feedback(results, sim_pairs):
    """Post-hoc upgrade YELLOW -> AMBER when similarity confirms template reuse (FEAT 13).

    Returns count of upgraded determinations.
    """
    result_map = {}
    for r in results:
        tid = r.get('task_id', '')
        if tid:
            result_map[tid] = r

    n_upgrades = 0
    for pair in sim_pairs:
        # Only act on meaningful similarity signals
        flag_type = pair.get('flag_type', '')
        has_semantic = 'semantic' in flag_type
        has_text = 'text' in flag_type
        if not has_semantic and not has_text:
            continue

        r_a = result_map.get(pair['id_a'])
        r_b = result_map.get(pair['id_b'])
        if not r_a or not r_b:
            continue

        det_a = r_a.get('determination', '')
        det_b = r_b.get('determination', '')

        # Case 1: one YELLOW paired with RED/AMBER → upgrade YELLOW to AMBER
        if det_a == 'YELLOW' and det_b in ('RED', 'AMBER'):
            r_a['determination'] = 'AMBER'
            r_a['reason'] = r_a.get('reason', '') + f" [upgraded: similarity with {pair['id_b']}]"
            r_a['similarity_upgraded'] = True
            n_upgrades += 1
        if det_b == 'YELLOW' and det_a in ('RED', 'AMBER'):
            r_b['determination'] = 'AMBER'
            r_b['reason'] = r_b.get('reason', '') + f" [upgraded: similarity with {pair['id_a']}]"
            r_b['similarity_upgraded'] = True
            n_upgrades += 1

        # Case 2: both YELLOW and high semantic → upgrade both to AMBER
        if det_a == 'YELLOW' and det_b == 'YELLOW' and pair.get('semantic', 0) >= 0.90:
            r_a['determination'] = 'AMBER'
            r_a['reason'] = r_a.get('reason', '') + f" [upgraded: mutual similarity with {pair['id_b']}]"
            r_a['similarity_upgraded'] = True
            r_b['determination'] = 'AMBER'
            r_b['reason'] = r_b.get('reason', '') + f" [upgraded: mutual similarity with {pair['id_a']}]"
            r_b['similarity_upgraded'] = True
            n_upgrades += 2

    return n_upgrades


def print_similarity_report(pairs):
    """Print cross-submission similarity findings."""
    if not pairs:
        print("\n  No cross-attempter similarity clusters detected.")
        return

    print(f"\n{'='*90}")
    print(f"  SIMILARITY CLUSTERS: {len(pairs)} flagged pairs")
    print(f"{'='*90}")

    for p in pairs:
        icon = 'RED' if p['jaccard'] >= 0.70 else 'AMBER' if p['jaccard'] >= 0.50 else 'YELLOW'
        prefix = '[CROSS-BATCH] ' if p.get('flag_type') == 'cross-batch' else ''
        sem_str = f"  Sem={p['semantic']:.2f}" if p.get('semantic', 0) > 0 else ''
        print(f"\n  {prefix}[{icon}] Jaccard={p['jaccard']:.2f}  Struct={p['structural']:.2f}{sem_str}  [{p['flag_type']}]")
        print(f"     {p['id_a'][:15]:15s} ({p['attempter_a'] or '?':20s}) [{p['det_a']}]")
        print(f"     {p['id_b'][:15]:15s} ({p['attempter_b'] or '?':20s}) [{p['det_b']}]")
        print(f"     Occupation: {p['occupation'][:50]}")

# ==============================================================================
# HISTORICAL MEMORY STORE
# Unified persistence layer for cross-batch detection memory.
# All data lives in a single directory (default .beet/).
# ==============================================================================

from pathlib import Path as _Path
import shutil as _shutil


class MemoryStore:
    """Persistent memory for the BEET detection pipeline.

    Usage:
        store = MemoryStore('.beet/')
        store.record_batch(results, text_map, batch_id='batch_001')
        history = store.get_attempter_history('worker_42')
        cross_matches = store.cross_batch_similarity(results, text_map)
        store.record_confirmation('task_001', 'ai', verified_by='reviewer_A')
    """

    def __init__(self, store_dir='.beet'):
        self.store_dir = _Path(store_dir)
        self.store_dir.mkdir(parents=True, exist_ok=True)
        (self.store_dir / 'calibration_history').mkdir(exist_ok=True)

        self.submissions_path = self.store_dir / 'submissions.jsonl'
        self.fingerprints_path = self.store_dir / 'fingerprints.jsonl'
        self.attempters_path = self.store_dir / 'attempters.jsonl'
        self.confirmed_path = self.store_dir / 'confirmed.jsonl'
        self.calibration_path = self.store_dir / 'calibration.json'
        self.config_path = self.store_dir / 'config.json'

        self._config = self._load_config()

    # ── Config ────────────────────────────────────────────────────

    def _load_config(self):
        if self.config_path.exists():
            with open(self.config_path) as f:
                return json.load(f)
        return {
            'version': '0.65',
            'created': datetime.now().isoformat(),
            'last_updated': datetime.now().isoformat(),
            'total_submissions': 0,
            'total_batches': 0,
            'total_attempters': 0,
            'total_confirmed': 0,
            'occupations': [],
        }

    def _save_config(self):
        self._config['last_updated'] = datetime.now().isoformat()
        with open(self.config_path, 'w') as f:
            json.dump(self._config, f, indent=2)

    # ── Batch Recording ──────────────────────────────────────────

    def record_batch(self, results, text_map, batch_id=None):
        """Record a full batch of pipeline results to memory.

        Updates submissions, fingerprints, attempter profiles, and config.
        """
        if batch_id is None:
            batch_id = f"batch_{datetime.now().strftime('%Y-%m-%d_%H:%M')}"

        timestamp = datetime.now().isoformat()
        n_written = 0

        # Write submissions
        with open(self.submissions_path, 'a') as f:
            for r in results:
                record = {k: r.get(k) for k in _BASELINE_FIELDS}
                record['batch_id'] = batch_id
                record['timestamp'] = timestamp
                record['pipeline_version'] = r.get('audit_trail', {}).get(
                    'pipeline_version', 'unknown')
                record['similarity_partners'] = r.get('similarity_partners', 0)
                record['similarity_max_semantic'] = r.get('similarity_max_semantic', 0.0)
                wc = r.get('word_count', 0)
                if wc < 100:
                    record['length_bin'] = 'short'
                elif wc < 300:
                    record['length_bin'] = 'medium'
                elif wc < 800:
                    record['length_bin'] = 'long'
                else:
                    record['length_bin'] = 'very_long'
                f.write(json.dumps(record, default=str) + '\n')
                n_written += 1

        # Write fingerprints
        self._write_fingerprints(results, text_map, batch_id)

        # Update attempter profiles
        self._update_attempter_profiles(results, batch_id, timestamp)

        # Update config
        self._config['total_submissions'] += n_written
        self._config['total_batches'] += 1
        occs = set(self._config.get('occupations', []))
        for r in results:
            occ = r.get('occupation', '')
            if occ:
                occs.add(occ)
        self._config['occupations'] = sorted(occs)
        self._save_config()

        print(f"  Memory: {n_written} submissions recorded to {self.store_dir}/")
        return n_written

    def _write_fingerprints(self, results, text_map, batch_id):
        """Write MinHash and optional embedding fingerprints."""
        # Pre-compute embeddings if available
        embeddings = {}
        if HAS_SEMANTIC:
            _ensure_semantic()
            texts = []
            tids = []
            for r in results:
                tid = r.get('task_id', '')
                text = text_map.get(tid, '')
                if text:
                    texts.append(text)
                    tids.append(tid)
            if texts:
                raw_embeds = _EMBEDDER.encode(texts)
                for tid, emb in zip(tids, raw_embeds):
                    embeddings[tid] = [round(float(v), 5) for v in emb[:64]]

        with open(self.fingerprints_path, 'a') as f:
            for r in results:
                tid = r.get('task_id', '')
                text = text_map.get(tid, '')
                if not text:
                    continue

                shingles = _word_shingles(text)
                minhash = _minhash_signature(shingles)

                struct_vec = {feat: r.get(feat, 0) for feat in _STRUCT_FEATURES}

                record = {
                    'task_id': tid,
                    'attempter': r.get('attempter', ''),
                    'occupation': r.get('occupation', ''),
                    'batch_id': batch_id,
                    'determination': r.get('determination', ''),
                    'minhash_128': minhash,
                    'structural_vec': struct_vec,
                }

                if tid in embeddings:
                    record['embedding_64'] = embeddings[tid]

                f.write(json.dumps(record, default=str) + '\n')

    # ── Attempter Profiles ───────────────────────────────────────

    def _update_attempter_profiles(self, results, batch_id, timestamp):
        """Update rolling attempter profiles with new batch results."""
        profiles = self._load_attempter_profiles()

        by_att = defaultdict(list)
        for r in results:
            att = r.get('attempter', '').strip()
            if att:
                by_att[att].append(r)

        for att, submissions in by_att.items():
            if att not in profiles:
                profiles[att] = {
                    'attempter': att,
                    'total_submissions': 0,
                    'determinations': {'RED': 0, 'AMBER': 0, 'YELLOW': 0,
                                       'GREEN': 0, 'MIXED': 0, 'REVIEW': 0},
                    'confirmed_ai': 0,
                    'confirmed_human': 0,
                    'occupations': [],
                    'batches': [],
                    'first_seen': timestamp,
                    'feature_sums': {},
                    'feature_counts': 0,
                }

            p = profiles[att]
            p['total_submissions'] += len(submissions)
            p['last_seen'] = timestamp
            p['last_updated'] = timestamp

            if batch_id not in p['batches']:
                p['batches'].append(batch_id)

            for r in submissions:
                det = r.get('determination', 'GREEN')
                p['determinations'][det] = p['determinations'].get(det, 0) + 1

                occ = r.get('occupation', '')
                if occ and occ not in p['occupations']:
                    p['occupations'].append(occ)

                for feat in ['prompt_signature_cfd', 'instruction_density_idi',
                             'voice_dissonance_vsd', 'voice_dissonance_spec_score',
                             'self_similarity_nssi_score']:
                    val = r.get(feat, 0)
                    if val:
                        if feat not in p['feature_sums']:
                            p['feature_sums'][feat] = 0.0
                        p['feature_sums'][feat] += val

                p['feature_counts'] += 1

            # Compute derived fields
            total = p['total_submissions']
            flagged = (p['determinations'].get('RED', 0) +
                       p['determinations'].get('AMBER', 0) +
                       p['determinations'].get('MIXED', 0))
            p['flag_rate'] = round(flagged / max(total, 1), 3)

            if p['feature_counts'] > 0:
                p['mean_features'] = {
                    k: round(v / p['feature_counts'], 3)
                    for k, v in p['feature_sums'].items()
                }

            # Risk tier
            p['risk_tier'] = self._compute_risk_tier(p)

            # Primary detection channel
            channel_counts = Counter()
            for r in submissions:
                if r.get('determination') in ('RED', 'AMBER', 'MIXED'):
                    cd = r.get('channel_details', {})
                    if isinstance(cd, dict):
                        channels = cd.get('channels', {})
                        for ch, info in channels.items():
                            if isinstance(info, dict) and info.get('severity') in ('RED', 'AMBER'):
                                channel_counts[ch] += 1
            if channel_counts:
                p['primary_detection_channel'] = channel_counts.most_common(1)[0][0]

        self._save_attempter_profiles(profiles)
        self._config['total_attempters'] = len(profiles)

    @staticmethod
    def _compute_risk_tier(profile):
        """Compute risk tier from flag rate and confirmation history."""
        flag_rate = profile.get('flag_rate', 0)
        confirmed_ai = profile.get('confirmed_ai', 0)
        if confirmed_ai > 0 and flag_rate > 0.50:
            return 'CRITICAL'
        elif flag_rate > 0.30 or confirmed_ai > 0:
            return 'HIGH'
        elif flag_rate > 0.15:
            return 'ELEVATED'
        else:
            return 'NORMAL'

    def _load_attempter_profiles(self):
        """Load attempter profiles dict."""
        profiles = {}
        if self.attempters_path.exists():
            with open(self.attempters_path) as f:
                for line in f:
                    line = line.strip()
                    if line:
                        try:
                            p = json.loads(line)
                            profiles[p['attempter']] = p
                        except (json.JSONDecodeError, KeyError):
                            continue
        return profiles

    def _save_attempter_profiles(self, profiles):
        """Save attempter profiles (full rewrite)."""
        with open(self.attempters_path, 'w') as f:
            for p in sorted(profiles.values(),
                            key=lambda x: x.get('flag_rate', 0), reverse=True):
                f.write(json.dumps(p, default=str) + '\n')

    # ── Queries ──────────────────────────────────────────────────

    def get_attempter_history(self, attempter):
        """Get full history for a specific attempter."""
        profiles = self._load_attempter_profiles()
        profile = profiles.get(attempter.strip())

        submissions = []
        if self.submissions_path.exists():
            with open(self.submissions_path) as f:
                for line in f:
                    try:
                        rec = json.loads(line.strip())
                        if rec.get('attempter', '').strip() == attempter.strip():
                            submissions.append(rec)
                    except json.JSONDecodeError:
                        continue

        confirmations = []
        if self.confirmed_path.exists():
            with open(self.confirmed_path) as f:
                for line in f:
                    try:
                        rec = json.loads(line.strip())
                        if rec.get('attempter', '').strip() == attempter.strip():
                            confirmations.append(rec)
                    except json.JSONDecodeError:
                        continue

        return {
            'profile': profile,
            'submissions': submissions,
            'confirmations': confirmations,
        }

    def get_attempter_risk_report(self, min_submissions=2):
        """Get all attempters ranked by risk tier and flag rate."""
        profiles = self._load_attempter_profiles()
        tier_rank = {'CRITICAL': 4, 'HIGH': 3, 'ELEVATED': 2, 'NORMAL': 1}
        return sorted(
            [p for p in profiles.values()
             if p.get('total_submissions', 0) >= min_submissions],
            key=lambda p: (-tier_rank.get(p.get('risk_tier', 'NORMAL'), 0),
                           -p.get('flag_rate', 0)),
        )

    def get_occupation_baselines(self, occupation):
        """Get historical feature distributions for an occupation."""
        submissions = []
        if self.submissions_path.exists():
            with open(self.submissions_path) as f:
                for line in f:
                    try:
                        rec = json.loads(line.strip())
                        if rec.get('occupation', '') == occupation:
                            submissions.append(rec)
                    except json.JSONDecodeError:
                        continue
        return submissions

    def pre_batch_context(self, attempter=None, occupation=None):
        """Retrieve historical context before running a batch."""
        context = {}

        if attempter:
            profiles = self._load_attempter_profiles()
            profile = profiles.get(attempter.strip())
            if profile:
                context['attempter_risk_tier'] = profile.get('risk_tier', 'UNKNOWN')
                context['attempter_flag_rate'] = profile.get('flag_rate', 0)
                context['attempter_total'] = profile.get('total_submissions', 0)
                context['attempter_confirmed_ai'] = profile.get('confirmed_ai', 0)

        if occupation:
            subs = self.get_occupation_baselines(occupation)
            if len(subs) >= 5:
                cfd_values = [s.get('prompt_signature_cfd', 0) for s in subs]
                idi_values = [s.get('instruction_density_idi', 0) for s in subs]
                context['occupation_n'] = len(subs)
                context['occupation_median_cfd'] = round(
                    statistics.median(cfd_values), 3)
                context['occupation_median_idi'] = round(
                    statistics.median(idi_values), 3)

        return context

    # ── Cross-Batch Similarity ───────────────────────────────────

    def cross_batch_similarity(self, current_results, text_map,
                               minhash_threshold=0.50):
        """Compare current batch against historical fingerprints."""
        historical = []
        if self.fingerprints_path.exists():
            with open(self.fingerprints_path) as f:
                for line in f:
                    try:
                        historical.append(json.loads(line.strip()))
                    except json.JSONDecodeError:
                        continue

        if not historical:
            return []

        flags = []
        for r in current_results:
            tid = r.get('task_id', '')
            text = text_map.get(tid, '')
            if not text:
                continue

            current_minhash = _minhash_signature(_word_shingles(text))

            for hist in historical:
                if hist.get('task_id') == tid:
                    continue

                att_curr = r.get('attempter', '').strip().lower()
                att_hist = hist.get('attempter', '').strip().lower()
                if att_curr and att_hist and att_curr == att_hist:
                    continue

                mh_sim = _minhash_jaccard(
                    current_minhash, hist.get('minhash_128', []))

                if mh_sim >= minhash_threshold:
                    flags.append({
                        'current_id': tid,
                        'historical_id': hist['task_id'],
                        'current_attempter': r.get('attempter', ''),
                        'historical_attempter': hist.get('attempter', ''),
                        'occupation': r.get('occupation', ''),
                        'minhash_similarity': round(mh_sim, 3),
                        'historical_determination': hist.get('determination', '?'),
                        'historical_batch': hist.get('batch_id', '?'),
                    })

        flags.sort(key=lambda f: f['minhash_similarity'], reverse=True)
        return flags

    # ── Confirmation Feedback ────────────────────────────────────

    def record_confirmation(self, task_id, ground_truth, verified_by='',
                            notes=''):
        """Record a human-verified ground truth label."""
        # Find the original submission
        original = None
        if self.submissions_path.exists():
            with open(self.submissions_path) as f:
                for line in f:
                    try:
                        rec = json.loads(line.strip())
                        if rec.get('task_id') == task_id:
                            original = rec
                            break
                    except json.JSONDecodeError:
                        continue

        record = {
            'task_id': task_id,
            'ground_truth': ground_truth,
            'verified_by': verified_by,
            'verified_at': datetime.now().isoformat(),
            'notes': notes,
        }

        if original:
            record['attempter'] = original.get('attempter', '')
            record['occupation'] = original.get('occupation', '')
            record['pipeline_determination'] = original.get('determination', '')
            record['pipeline_confidence'] = original.get('confidence', 0)

        with open(self.confirmed_path, 'a') as f:
            f.write(json.dumps(record, default=str) + '\n')

        # Update attempter profile
        if original and original.get('attempter'):
            profiles = self._load_attempter_profiles()
            att = original['attempter'].strip()
            if att in profiles:
                if ground_truth == 'ai':
                    profiles[att]['confirmed_ai'] = profiles[att].get(
                        'confirmed_ai', 0) + 1
                else:
                    profiles[att]['confirmed_human'] = profiles[att].get(
                        'confirmed_human', 0) + 1
                profiles[att]['risk_tier'] = self._compute_risk_tier(profiles[att])
                self._save_attempter_profiles(profiles)

        self._config['total_confirmed'] = self._config.get(
            'total_confirmed', 0) + 1
        self._save_config()

        print(f"  Confirmed: {task_id} = {ground_truth} (by {verified_by})")

    # ── Calibration Integration ──────────────────────────────────

    def rebuild_calibration(self):
        """Rebuild calibration table from all confirmed human submissions."""
        confirmed = {}
        if self.confirmed_path.exists():
            with open(self.confirmed_path) as f:
                for line in f:
                    try:
                        rec = json.loads(line.strip())
                        confirmed[rec['task_id']] = rec['ground_truth']
                    except (json.JSONDecodeError, KeyError):
                        continue

        if not confirmed:
            print("  No confirmed labels — cannot rebuild calibration")
            return None

        import tempfile as _tempfile
        with _tempfile.NamedTemporaryFile(mode='w', suffix='.jsonl',
                                          delete=False) as tmp:
            if self.submissions_path.exists():
                with open(self.submissions_path) as f:
                    for line in f:
                        try:
                            rec = json.loads(line.strip())
                            tid = rec.get('task_id', '')
                            if tid in confirmed:
                                rec['ground_truth'] = confirmed[tid]
                                tmp.write(json.dumps(rec, default=str) + '\n')
                        except json.JSONDecodeError:
                            continue
            tmp_path = tmp.name

        cal = calibrate_from_baselines(tmp_path)
        os.unlink(tmp_path)

        if cal is None:
            print("  Insufficient confirmed human data for calibration")
            return None

        # Snapshot current calibration before overwriting
        if self.calibration_path.exists():
            snapshot_name = f"cal_{datetime.now().strftime('%Y-%m-%d_%H%M')}.json"
            snapshot_path = self.store_dir / 'calibration_history' / snapshot_name
            _shutil.copy2(str(self.calibration_path), str(snapshot_path))

        save_calibration(cal, str(self.calibration_path))
        print(f"  Calibration rebuilt: {cal.get('n_calibration', 0)} labeled samples")
        return cal

    # ── Summary ──────────────────────────────────────────────────

    def print_summary(self):
        """Print memory store summary."""
        c = self._config
        print(f"\n  BEET Memory Store: {self.store_dir}/")
        print(f"    Submissions: {c.get('total_submissions', 0)}")
        print(f"    Batches:     {c.get('total_batches', 0)}")
        print(f"    Attempters:  {c.get('total_attempters', 0)}")
        print(f"    Confirmed:   {c.get('total_confirmed', 0)}")
        occs = c.get('occupations', [])
        if occs:
            print(f"    Occupations: {', '.join(occs[:10])}"
                  f"{'...' if len(occs) > 10 else ''}")
        print(f"    Last update: {c.get('last_updated', 'never')}")


def _print_attempter_history(history):
    """Print formatted attempter history from memory store."""
    profile = history.get('profile')
    if not profile:
        print("  No history found for this attempter.")
        return

    print(f"\n  ATTEMPTER: {profile['attempter']}")
    print(f"    Risk tier:   {profile.get('risk_tier', 'UNKNOWN')}")
    print(f"    Flag rate:   {profile.get('flag_rate', 0):.0%}")
    print(f"    Submissions: {profile.get('total_submissions', 0)}")

    dets = profile.get('determinations', {})
    print(f"    R={dets.get('RED', 0)} A={dets.get('AMBER', 0)} "
          f"Y={dets.get('YELLOW', 0)} G={dets.get('GREEN', 0)}")

    if profile.get('confirmed_ai', 0) or profile.get('confirmed_human', 0):
        print(f"    Confirmed: AI={profile.get('confirmed_ai', 0)} "
              f"Human={profile.get('confirmed_human', 0)}")

    occs = profile.get('occupations', [])
    if occs:
        print(f"    Occupations: {', '.join(occs[:5])}")

    print(f"    First seen: {profile.get('first_seen', '?')}")
    print(f"    Last seen:  {profile.get('last_seen', '?')}")
    print(f"    Batches:    {len(profile.get('batches', []))}")

    submissions = history.get('submissions', [])
    if submissions:
        print(f"\n    Recent submissions (last 10):")
        for s in submissions[-10:]:
            print(f"      {s.get('task_id', '?')[:15]:15} "
                  f"[{s.get('determination', '?'):6}] "
                  f"conf={s.get('confidence', 0):.2f} "
                  f"{s.get('occupation', '')[:30]}")

    confirmations = history.get('confirmations', [])
    if confirmations:
        print(f"\n    Confirmations ({len(confirmations)}):")
        for c in confirmations:
            print(f"      {c.get('task_id', '?')[:15]} = {c.get('ground_truth', '?')} "
                  f"(by {c.get('verified_by', '?')})")


# ==============================================================================
# ANALYZER: PREAMBLE
# Preamble detection -- catches LLM output artifacts.
#
# Detects: assistant acknowledgments, artifact delivery frames, first-person
# creation claims, meta-design language, style masking, editorial meta-commentary,
# and Chain-of-Thought leakage from Large Reasoning Models (DeepSeek-R1, o1/o3).
# ==============================================================================
PREAMBLE_PATTERNS = [
    (r"(?i)^\s*[\"']?(got it|sure thing|absolutely|certainly|of course)[.!,\s]", "assistant_ack", "CRITICAL"),
    (r"(?i)^\s*[\"']?here(?:'s| is| are)\s+(your|the|a)\s+(final|updated|revised|complete|rewritten|prompt|task|evaluation)", "artifact_delivery", "CRITICAL"),
    (r"(?i)^\s*[\"']?below is\s+(a\s+)?(rewritten|revised|updated|the|your)", "artifact_delivery", "CRITICAL"),
    (r"(?i)(copy[- ]?paste|ready to use|plug[- ]and[- ]play)", "copy_paste_instruction", "MEDIUM"),
    (r"(?i)(failure[- ]inducing|designed to (test|challenge|trip|catch|induce))", "meta_design", "CRITICAL"),
    (r"(?i)^\s*[\"']?(I'?ve |I have |I'?ll |let me )(created?|drafted?|prepared?|written|designed|built|put together)", "first_person_creation", "CRITICAL"),
    (r"(?i)(natural workplace style|sounds? like a real|human[- ]issued|reads? like a human)", "style_masking", "HIGH"),
    (r"(?i)notes on what I (fixed|changed|cleaned|updated|revised)", "editorial_meta", "HIGH"),
    # Chain-of-thought leakage from Large Reasoning Models (DeepSeek-R1, o1/o3)
    (r"<think>", "cot_leakage", "CRITICAL"),
    (r"</think>", "cot_leakage", "CRITICAL"),
    (r"<reasoning>", "cot_leakage", "CRITICAL"),
    (r"</reasoning>", "cot_leakage", "CRITICAL"),
    (r"(?i)\blet me (?:rethink|reconsider|recalculate|re-examine|verify|double[- ]check|think about this)\b", "cot_reasoning", "HIGH"),
    (r"(?i)\bwait,?\s+(?:actually|no|let me|that'?s not)", "cot_self_correction", "HIGH"),
    (r"(?i)\bhmm,?\s+(?:let me|on second thought|actually)", "cot_self_correction", "HIGH"),
    (r"(?i)\bmy (?:final|revised|updated) answer (?:is|should|would)\b", "cot_conclusion", "HIGH"),
    (r"(?i)\bstep \d+\s*:", "cot_step_numbering", "MEDIUM"),
]


def run_preamble(text):
    """Detect LLM preamble artifacts. Returns (score, severity, hits, spans)."""
    first_500 = text[:500]
    hits = []
    spans = []
    severity = 'NONE'

    for pat, name, sev in PREAMBLE_PATTERNS:
        search_text = first_500 if name in (
            'assistant_ack', 'artifact_delivery', 'first_person_creation', 'cot_leakage',
        ) else text
        match = re.search(pat, search_text)
        if match:
            hits.append((name, sev))
            spans.append({
                'start': match.start(),
                'end': match.end(),
                'text': match.group()[:80],
                'pattern': name,
                'severity': sev,
            })
            if sev == 'CRITICAL':
                severity = 'CRITICAL'
            elif sev == 'HIGH' and severity not in ('CRITICAL',):
                severity = 'HIGH'
            elif sev == 'MEDIUM' and severity == 'NONE':
                severity = 'MEDIUM'

    score = {'CRITICAL': 0.99, 'HIGH': 0.75, 'MEDIUM': 0.50, 'NONE': 0.0}[severity]
    return score, severity, hits, spans

# ==============================================================================
# ANALYZER: FINGERPRINT
# Intrinsic fingerprint detection -- LLM-preferred vocabulary.
# ==============================================================================
FINGERPRINT_WORDS = [
    # Original 27 (ChatGPT-3.5 era, established in v0.51)
    'delve', 'utilize', 'comprehensive', 'streamline', 'leverage', 'robust',
    'facilitate', 'innovative', 'synergy', 'paradigm', 'holistic', 'nuanced',
    'multifaceted', 'spearhead', 'underscore', 'pivotal', 'landscape',
    'cutting-edge', 'actionable', 'seamlessly', 'noteworthy', 'meticulous',
    'endeavor', 'paramount', 'aforementioned', 'furthermore', 'henceforth',
    # v0.63 additions (Kobak et al. 2024 excess vocabulary, Science Advances)
    'tapestry', 'realm', 'embark', 'foster', 'showcasing',
]

_FINGERPRINT_RE = re.compile(
    r'\b(?:' + '|'.join(re.escape(w) for w in FINGERPRINT_WORDS) + r')\b',
    re.IGNORECASE
)


def run_fingerprint(text):
    """Detect LLM fingerprint words. Returns (score, hit_count, rate)."""
    word_count = len(text.split())
    matches = _FINGERPRINT_RE.findall(text)
    hits = len(matches)
    rate = hits / max(word_count / 1000, 1)
    score = min(rate / 5.0, 1.0)
    return score, hits, rate


def collect_spans(text):
    """Collect character-level spans from all regex-based detectors.

    Returns a sorted list of dicts, each with:
        start, end     – character offsets into *text*
        text           – the matched substring
        layer          – which detector family fired
        pattern        – pattern identifier
        severity       – CRITICAL / HIGH / MEDIUM / LOW
        weight         – numeric weight (matches scoring weights where applicable)

    Spans are purely diagnostic; they do not affect scoring.
    """
    spans = []

    # --- Preamble patterns ---
    first_500_names = {
        'assistant_ack', 'artifact_delivery', 'first_person_creation', 'cot_leakage',
    }
    for pat_str, name, sev in PREAMBLE_PATTERNS:
        search_text = text[:500] if name in first_500_names else text
        for m in re.finditer(pat_str, search_text):
            spans.append({
                'start': m.start(),
                'end': m.end(),
                'text': m.group(),
                'layer': 'preamble',
                'pattern': name,
                'severity': sev,
                'weight': {'CRITICAL': 0.99, 'HIGH': 0.75, 'MEDIUM': 0.50}.get(sev, 0.0),
            })

    # --- Fingerprint words ---
    for m in _FINGERPRINT_RE.finditer(text):
        spans.append({
            'start': m.start(),
            'end': m.end(),
            'text': m.group(),
            'layer': 'fingerprint',
            'pattern': 'fingerprint_word',
            'severity': 'LOW',
            'weight': 1.0,
        })

    # --- Formulaic academic phrases ---
    for compiled_pat, weight in _FORMULAIC_PATTERNS:
        for m in compiled_pat.finditer(text):
            spans.append({
                'start': m.start(),
                'end': m.end(),
                'text': m.group(),
                'layer': 'formulaic',
                'pattern': compiled_pat.pattern[:60],
                'severity': 'MEDIUM',
                'weight': weight,
            })

    # --- Power adjectives ---
    for m in _POWER_ADJ.finditer(text):
        spans.append({
            'start': m.start(),
            'end': m.end(),
            'text': m.group(),
            'layer': 'power_adj',
            'pattern': 'power_adjective',
            'severity': 'LOW',
            'weight': 1.0,
        })

    # --- Transition connectors ---
    for m in _TRANSITION.finditer(text):
        spans.append({
            'start': m.start(),
            'end': m.end(),
            'text': m.group(),
            'layer': 'transition',
            'pattern': 'transition_connector',
            'severity': 'LOW',
            'weight': 1.0,
        })

    # --- Demonstrative phrases ---
    for m in _DEMONSTRATIVE.finditer(text):
        spans.append({
            'start': m.start(),
            'end': m.end(),
            'text': m.group(),
            'layer': 'demonstrative',
            'pattern': 'demonstrative_phrase',
            'severity': 'LOW',
            'weight': 1.0,
        })

    spans.sort(key=lambda s: s['start'])
    return spans


# ==============================================================================
# ANALYZER: PROMPT SIGNATURE
# ==============================================================================
"""Prompt-engineering signature detection -- CFD, MFSR, numbered criteria."""


CONSTRAINT_FRAMES = [
    r'must account for', r'should be visible', r'at least \d+[%$]?',
    r'at or below', r'no more than', r'no \w+ may', r'must have',
    r'should address', r'should be delivered', r'within \d+%',
    r'or higher', r'or lower', r'instead of', r'without exceeding',
    r'about \d+[-–]\d+', r'strictly on',
    r'must include\b', r'must address\b', r'must be \w+',
    r'you may not\b',
    r'may not (?:be|introduce|omit|use|exceed|include)\b',
    r'in this exact\b', r'with exactly \d+',
    r'every \w+ must\b', r'all \w+ must\b',
    r'clearly (?:list|state|describe|identify|document)',
    r'(?:document|report|response) must\b',
    r'(?:following|these) sections',
    r'use \w+ formatting', r'plain language',
    r'no \w+[- ]only\b',
]

META_DESIGN_PATTERNS = [
    r'(?i)workflows? tested',
    r'(?i)acceptance (checklist|criteria)',
    r'(?i)(used for|for) grading',
    r'(?i)SOC \d{2}-?\d{4}',
    r'(?i)expected effort:?\s*\d',
    r'(?i)deliberate (anomalies|errors|issues)',
    r'(?i)checkable artifacts',
    r'(?i)authoritative source of truth',
    r'(?i)scenario anchor date',
    r'(?i)avoid vague language',
    r'(?i)explicit non-functional',
    r'(?i)grounded in\b',
]


def run_prompt_signature(text):
    """Detect prompt-engineering signatures. Returns dict of metrics."""
    sents = get_sentences(text)
    n_sents = max(len(sents), 1)
    word_count = len(text.split())

    total_frames = 0
    distinct_pats = set()
    for pat in CONSTRAINT_FRAMES:
        matches = re.findall(pat, text, re.IGNORECASE)
        if matches:
            total_frames += len(matches)
            distinct_pats.add(pat)
    cfd = total_frames / n_sents

    multi_frame = 0
    for sent in sents:
        ct = sum(1 for pat in CONSTRAINT_FRAMES if re.search(pat, sent, re.IGNORECASE))
        if ct >= 2:
            multi_frame += 1
    mfsr = multi_frame / n_sents

    has_role = bool(re.search(r'you (are|work|supervise|manage|serve|lead|oversee)', text[:600], re.IGNORECASE))
    has_deliverable = bool(re.search(r'(submit|deliver|present|provide|create|produce|prepare|generate)\s+(your|the|a|an|exactly)', text, re.IGNORECASE))
    has_closing = bool(re.search(r'(final|should be delivered|all conclusions|base all|submission|deliverable)', text[-300:], re.IGNORECASE))
    fc = int(has_role) + int(has_deliverable) + int(has_closing)

    cond_count = len(re.findall(r'\bif\b[^.]*?,', text, re.IGNORECASE))
    cond_count += len(re.findall(r'\bwhen\b[^.]*?,', text, re.IGNORECASE))
    cond_count += len(re.findall(r'\bunless\b', text, re.IGNORECASE))
    cond_density = cond_count / n_sents

    meta_hits = [pat for pat in META_DESIGN_PATTERNS if re.search(pat, text)]

    contractions = len(re.findall(r"\b\w+'(?:t|re|ve|s|d|ll|m)\b", text, re.IGNORECASE))

    must_count = len(re.findall(r'\bmust\b', text, re.IGNORECASE))
    must_rate = must_count / n_sents

    numbered_criteria = len(re.findall(r'^\s*\d{1,2}[.)]\s+.{20,}', text, re.MULTILINE))

    composite = 0.0
    if cfd >= 0.50: composite += 0.40
    elif cfd >= 0.30: composite += 0.25
    elif cfd >= 0.15: composite += 0.10
    if len(distinct_pats) >= 8: composite += 0.20
    elif len(distinct_pats) >= 5: composite += 0.12
    elif len(distinct_pats) >= 3: composite += 0.05
    if len(meta_hits) >= 3: composite += 0.20
    elif len(meta_hits) >= 1: composite += 0.08
    if fc == 3: composite += 0.10
    if fc >= 2 and len(distinct_pats) >= 8:
        composite += 0.15
    if numbered_criteria >= 15: composite += 0.15
    elif numbered_criteria >= 10: composite += 0.08
    if contractions == 0 and word_count > 500: composite += 0.05

    return {
        'cfd': cfd,
        'distinct_frames': len(distinct_pats),
        'mfsr': mfsr,
        'framing_completeness': fc,
        'conditional_density': cond_density,
        'meta_design_hits': len(meta_hits),
        'meta_design_details': meta_hits,
        'contractions': contractions,
        'must_count': must_count,
        'must_rate': must_rate,
        'numbered_criteria': numbered_criteria,
        'composite': min(composite, 1.0),
    }


# ==============================================================================
# ANALYZER: VOICE DISSONANCE
# Voice-Specification Dissonance (VSD) + Sterile Spec Index detection.
# ==============================================================================
CASUAL_MARKERS = [
    'hey', 'ok so', 'ok,', 'dont', 'wont', 'cant', 'gonna', 'gotta',
    'thx', 'pls', 'gimme', 'lemme', 'kinda', 'sorta', 'tho', 'btw',
    'fyi', 'alot', 'ya', 'yep', 'nah', 'nope', 'lol', 'haha',
]

MANUFACTURED_TYPOS = [
    'atached', 'alot', 'recieved', 'seperate', 'occured', 'wierd',
    'definately', 'accomodate', 'occurence', 'independant', 'noticable',
    'occassion', 'tommorow', 'calender', 'begining', 'acheive', 'untill',
    'beleive', 'existance', 'grammer', 'arguement', 'commited',
    'maintainance', 'necesary', 'occuring', 'persue', 'prefered',
    'recomend', 'refered', 'succesful', 'suprise',
]


def _build_marker_pattern(marker):
    tokens = marker.split()
    if len(tokens) > 1:
        escaped = r'\s+'.join(re.escape(t) for t in tokens)
    else:
        escaped = re.escape(marker)
    first_char = tokens[0][0]
    if first_char.isalnum() or first_char == '_':
        leading = r'\b'
    else:
        leading = r'(?<!\w)'
    last_char = tokens[-1][-1]
    if last_char.isalnum() or last_char == '_':
        trailing = r'\b'
    else:
        trailing = r'(?!\w)'
    return leading + escaped + trailing


_CASUAL_RE = [re.compile(_build_marker_pattern(m), re.IGNORECASE) for m in CASUAL_MARKERS]
_TYPO_RE = [re.compile(r'\b' + re.escape(t) + r'\b', re.IGNORECASE) for t in MANUFACTURED_TYPOS]


def run_voice_dissonance(text):
    """Detect voice-specification dissonance. Returns dict of metrics."""
    words = text.split()
    n_words = len(words)
    per100 = max(n_words / 100, 1)

    casual_count = sum(len(pat.findall(text)) for pat in _CASUAL_RE)
    misspelling_count = sum(len(pat.findall(text)) for pat in _TYPO_RE)

    contractions = len(re.findall(r"\b\w+'(?:t|re|ve|s|d|ll|m)\b", text, re.IGNORECASE))

    em_dashes = len(re.findall(r'(?<!\d)\s?[—–]\s?(?!\d)', text))
    em_dashes += len(re.findall(r' - ', text))

    lowercase_starts = sum(1 for line in text.split('\n') if line.strip() and line.strip()[0].islower())

    voice_score = (casual_count * 5 + misspelling_count * 1 + contractions * 1.5
                   + em_dashes * 1 + lowercase_starts * 0.5) / per100

    camel_cols = len(re.findall(r'[A-Z][a-z]+_[A-Z][a-z_]+', text))
    filenames = len(set(re.findall(
        r'\w+\.(?:csv|xlsx|xls|tsv|json|xml|pdf|docx|doc|pptx|ppt|txt|md|html|py|zip|png|jpg|jpeg|gif|mp4)\b',
        text, re.IGNORECASE)))
    calcs = len(re.findall(
        r'(calculated?|computed?|deriv|formula|multiply|divid|subtract|sum\b|average|ratio|percent|\bnet\b.*[-=])',
        text, re.IGNORECASE))
    tabs = len(re.findall(r'(?i)(tab \d|\btab\b.*[:—-]|sheet \d)', text))
    col_listings = len(re.findall(r'(?:columns?|fields?)\s*[:]\s*\w', text, re.IGNORECASE))
    tech_parens = len(re.findall(
        r'\([^)]*(?:\.\w{2,4}|%|\d+[kKmM]?\b|formula|column|tab)[^)]*\)', text))

    spec_score = (camel_cols * 1.5 + filenames * 2 + calcs * 2 + tabs * 3
                  + col_listings * 3 + tech_parens * 1) / per100

    vsd = voice_score * spec_score

    hedges = len(re.findall(
        r'\b(pretty sure|i think|probably|maybe|might be|seems like|sort of|kind of|'
        r'not sure|i guess|iirc|afaik|if i recall|i believe)\b', text, re.IGNORECASE))

    return {
        'voice_score': voice_score,
        'spec_score': spec_score,
        'vsd': vsd,
        'voice_gated': voice_score > 2.0,
        'casual_markers': casual_count,
        'misspellings': misspelling_count,
        'contractions': contractions,
        'em_dashes': em_dashes,
        'camel_cols': camel_cols,
        'filenames': filenames,
        'calcs': calcs,
        'tabs': tabs,
        'col_listings': col_listings,
        'hedges': hedges,
    }

# ==============================================================================
# ANALYZER: INSTRUCTION DENSITY
# Instruction Density Index (IDI) -- catches formal-exhaustive LLM output.
# ==============================================================================
def run_instruction_density(text):
    """Compute instruction density index. Returns dict of metrics."""
    words = text.split()
    n_words = len(words)
    per100 = max(n_words / 100, 1)

    imp_keywords = ['must', 'include', 'create', 'load', 'set', 'show', 'use', 'derive', 'treat', 'mark']
    imperatives = sum(len(re.findall(r'\b' + kw + r'\b', text, re.IGNORECASE)) for kw in imp_keywords)

    cond_keywords = ['if', 'otherwise', 'when', 'unless']
    conditionals = sum(len(re.findall(r'\b' + kw + r'\b', text, re.IGNORECASE)) for kw in cond_keywords)

    binary_specs = len(re.findall(r'\b(?:Yes|No)\b', text))
    missing_handling = len(re.findall(r'\bMISSING\b', text))
    flag_count = len(re.findall(r'\b[Ff]lag\b', text))

    idi = (imperatives * 1.0 + conditionals * 2.0 + binary_specs * 1.5 +
           missing_handling * 3.0 + flag_count * 2.0) / per100

    return {
        'idi': idi,
        'imperatives': imperatives,
        'imp_rate': imperatives / per100,
        'conditionals': conditionals,
        'cond_rate': conditionals / per100,
        'binary_specs': binary_specs,
        'missing_refs': missing_handling,
        'flag_count': flag_count,
    }

# ==============================================================================
# ANALYZER: SEMANTIC RESONANCE
# ==============================================================================
"""Semantic Resonance -- embedding proximity to AI/human archetype centroids.

Ref: Mitchell et al. (2023) "DetectGPT" -- semantic density as AI signal.
"""



def run_semantic_resonance(text):
    """Measure semantic similarity to AI vs human archetypes.

    Returns dict with semantic scores, delta, determination, and confidence.
    """
    _ensure_semantic()
    if not HAS_SEMANTIC:
        return {
            'semantic_ai_score': 0.0,
            'semantic_human_score': 0.0,
            'semantic_delta': 0.0,
            'determination': None,
            'confidence': 0.0,
            'reason': 'Semantic layer unavailable (sentence-transformers not installed)',
        }

    words = text.split()
    if len(words) < 30:
        return {
            'semantic_ai_score': 0.0,
            'semantic_human_score': 0.0,
            'semantic_delta': 0.0,
            'determination': None,
            'confidence': 0.0,
            'reason': 'Semantic layer: text too short',
        }

    chunk_size = 200
    chunks = []
    for i in range(0, len(words), chunk_size):
        chunk = ' '.join(words[i:i + chunk_size])
        if len(chunk.split()) >= 30:
            chunks.append(chunk)

    if not chunks:
        chunks = [text]

    vecs = _EMBEDDER.encode(chunks)

    ai_sims = _cosine_similarity(vecs, _AI_CENTROIDS)
    max_ai_sim = float(ai_sims.max())
    mean_ai_sim = float(ai_sims.max(axis=1).mean())

    human_sims = _cosine_similarity(vecs, _HUMAN_CENTROIDS)
    max_human_sim = float(human_sims.max())
    mean_human_sim = float(human_sims.max(axis=1).mean())

    semantic_delta = mean_ai_sim - mean_human_sim

    if mean_ai_sim >= 0.65 and semantic_delta >= 0.15:
        det = 'AMBER'
        conf = min(0.60, mean_ai_sim)
        reason = f"Semantic resonance: AI={mean_ai_sim:.2f}, delta={semantic_delta:.2f}"
    elif mean_ai_sim >= 0.50 and semantic_delta >= 0.08:
        det = 'YELLOW'
        conf = min(0.35, mean_ai_sim)
        reason = f"Semantic resonance: AI={mean_ai_sim:.2f}, delta={semantic_delta:.2f}"
    else:
        det = None
        conf = 0.0
        reason = 'Semantic resonance: below threshold'

    return {
        'semantic_ai_score': round(max_ai_sim, 4),
        'semantic_human_score': round(max_human_sim, 4),
        'semantic_ai_mean': round(mean_ai_sim, 4),
        'semantic_human_mean': round(mean_human_sim, 4),
        'semantic_delta': round(semantic_delta, 4),
        'determination': det,
        'confidence': conf,
        'reason': reason,
    }


def run_token_cohesiveness(text, n_copies=10, deletion_rate=0.015, seed=42):
    """Token Cohesiveness (TOCSIN) — semantic distance under random word deletion.

    AI-generated text tends to maintain higher semantic cohesion when words
    are randomly deleted, because LLM outputs are more redundant and
    semantically smooth. Human text degrades faster under deletion.

    Returns dict with cohesiveness score, std, determination, confidence, reason.
    Requires sentence-transformers (HAS_SEMANTIC).
    """
    _ensure_semantic()
    _tocsin_empty = {
        'cohesiveness': 0.0,
        'cohesiveness_std': 0.0,
        'n_rounds': 0,
        'determination': None,
        'confidence': 0.0,
        'reason': '',
    }
    if not HAS_SEMANTIC:
        _tocsin_empty['reason'] = 'TOCSIN unavailable (sentence-transformers not installed)'
        return _tocsin_empty

    words = text.split()
    if len(words) < 40:
        _tocsin_empty['reason'] = 'TOCSIN: text too short'
        return _tocsin_empty

    import random as _random_mod
    rng = _random_mod.Random(seed)
    n_delete = max(1, int(len(words) * deletion_rate))

    # Encode original
    original_vec = _EMBEDDER.encode([text])

    sims = []
    for _ in range(n_copies):
        indices = list(range(len(words)))
        rng.shuffle(indices)
        delete_set = set(indices[:n_delete])
        perturbed_words = [w for i, w in enumerate(words) if i not in delete_set]
        perturbed_text = ' '.join(perturbed_words)
        perturbed_vec = _EMBEDDER.encode([perturbed_text])
        sim = float(_cosine_similarity(original_vec, perturbed_vec)[0][0])
        sims.append(sim)

    cohesiveness = sum(sims) / len(sims)
    if len(sims) >= 2:
        coh_std = (sum((s - cohesiveness) ** 2 for s in sims) / (len(sims) - 1)) ** 0.5
    else:
        coh_std = 0.0

    # High cohesiveness = more AI-like (text is semantically redundant)
    if cohesiveness >= 0.995 and len(words) >= 100:
        det = 'YELLOW'
        conf = 0.25
        reason = f'TOCSIN: high cohesiveness {cohesiveness:.4f}'
    else:
        det = None
        conf = 0.0
        reason = f'TOCSIN: cohesiveness {cohesiveness:.4f}'

    return {
        'cohesiveness': round(cohesiveness, 6),
        'cohesiveness_std': round(coh_std, 6),
        'n_rounds': n_copies,
        'determination': det,
        'confidence': conf,
        'reason': reason,
    }


# ==============================================================================
# ANALYZER: SELF-SIMILARITY (NSSI)
# N-Gram Self-Similarity Index (NSSI) -- offline statistical fingerprinting.  
# ==============================================================================
# -- Formulaic Academic Phrases (pre-compiled at import time) --
_FORMULAIC_PATTERNS = [
    (re.compile(r'\bthis\s+(?:report|analysis|paper|study|section|document)\s+(?:provides?|presents?|examines?|dissects?|identifies?|evaluates?|proposes?|outlines?)\b', re.I), 1.5),
    (re.compile(r'\b(?:it\s+is\s+(?:worth|important|imperative|crucial|essential|critical)\s+(?:noting|to\s+note|to\s+acknowledge|to\s+emphasize|to\s+recognize))\b', re.I), 2.0),
    (re.compile(r'\b(?:to\s+address\s+this\s+(?:gap|issue|problem|challenge|limitation|deficiency|concern|shortcoming))\b', re.I), 1.5),
    (re.compile(r'\b(?:perhaps\s+the\s+most\s+(?:\w+\s+)?(?:damning|significant|important|critical|notable|striking|concerning|alarming))\b', re.I), 2.0),
    (re.compile(r'\b(?:(?:while|although)\s+(?:theoretically|conceptually|technically)\s+(?:sound|elegant|promising|valid|robust|appealing))\b', re.I), 2.0),
    (re.compile(r'\b(?:the\s+(?:analysis|evidence|data|results?|findings?)\s+(?:suggests?|reveals?|indicates?|shows?|demonstrates?|confirms?)\s+that)\b', re.I), 1.0),
    (re.compile(r'\b(?:this\s+(?:creates?|represents?|highlights?|underscores?|reveals?|illustrates?|exemplifies?)\s+(?:a|the|an))\b', re.I), 1.5),
    (re.compile(r'\b(?:the\s+(?:primary|core|fundamental|critical|key|central|overarching)\s+(?:challenge|issue|problem|question|limitation|concern|insight|takeaway))\b', re.I), 1.0),
    (re.compile(r'\b(?:in\s+(?:layman.s\s+terms|other\s+words|practical\s+terms|simple\s+terms|real.world\s+(?:terms|scenarios|situations)))\b', re.I), 1.5),
    (re.compile(r'\b(?:defense\s+in\s+depth)\b', re.I), 1.0),
    (re.compile(r'\b(?:arms?\s+race)\b', re.I), 0.5),
    (re.compile(r'\b(?:the\s+(?:era|age|dawn)\s+of)\b', re.I), 0.5),
    (re.compile(r'\b(?:a\s+(?:paradigm|fundamental|seismic|tectonic)\s+shift)\b', re.I), 2.0),
    (re.compile(r'\b(?:the\s+(?:elephant|gorilla)\s+in\s+the\s+room)\b', re.I), 1.5),
    (re.compile(r'\b(?:a\s+double.edged\s+sword)\b', re.I), 1.5),
    (re.compile(r'\b(?:in\s+(?:conclusion|summary|closing),?)\b', re.I), 0.5),
    (re.compile(r'\b(?:the\s+path\s+forward\s+(?:is|requires|demands|involves))\b', re.I), 1.5),
    (re.compile(r'\b(?:(?:unless|until)\s+the\s+(?:community|industry|field|sector)\s+(?:adopts?|embraces?|commits?))\b', re.I), 2.0),
    (re.compile(r'\b(?:the\s+(?:immediate|long.term|strategic)\s+(?:future|imperative|priority|solution)\s+(?:belongs?\s+to|lies?\s+in|requires?))\b', re.I), 2.0),
]

# -- Power Adjectives --
_POWER_ADJ = re.compile(
    r'\b(?:comprehensive|exhaustive|rigorous|robust|holistic|systemic|'
    r'fundamental|critical|profound|decisive|catastrophic|perilous|'
    r'unprecedented|groundbreaking|transformative|paradigmatic|'
    r'monumental|pivotal|seminal|nascent|burgeoning|'
    r'overarching|multifaceted|nuanced|granular|bespoke|'
    r'actionable|scalable|tractable|non-trivial|intractable)\b',
    re.I
)

# -- Discourse Scaffolding --
_SCARE_QUOTE = re.compile(r'[\u201c\u201d][^\u201c\u201d]{2,40}[\u201c\u201d]|"[^"]{2,40}"')
_EM_DASH = re.compile(r'\u2014|--')
_PAREN = re.compile(r'\([^)]{12,}\)')
_COLON_EXPLAIN = re.compile(r':\s+[A-Z]')

# -- Demonstrative Monotony --
_DEMONSTRATIVE = re.compile(
    r'\bthis\s+(?:approach|method|framework|analysis|issue|mechanism|assumption|'
    r'limitation|strategy|technique|variant|disparity|metric|paradigm|'
    r'architecture|pipeline|deficiency|vulnerability|solution|concept|'
    r'pattern|signal|feature|constraint|observation|phenomenon|'
    r'suggests?|indicates?|creates?|ensures?|effectively|underscores?|'
    r'highlights?|represents?|reveals?|demonstrates?|means?|implies?|'
    r'raises?|poses?|necessitates?)\b', re.I
)

# -- Transition Connector Density --
_TRANSITION = re.compile(
    r'\b(?:however|furthermore|consequently|moreover|nevertheless|'
    r'additionally|specifically|crucially|ultimately|conversely|'
    r'notably|importantly|interestingly|remarkably|significantly|'
    r'simultaneously|correspondingly|paradoxically)\b', re.I
)

# -- Causal Reasoning Deficit --
_CAUSAL = re.compile(
    r'\b(?:because|since|'
    r'so\b(?!\s+(?:that|much|many|far|long|called))|'
    r'if|but|although|though|unless|whereas|'
    r'while(?=\s+\w+\s+(?:is|was|are|were|has|had|do|does|did|can|could|would|should|might|may))|'
    r'therefore|hence|thus|'
    r'think|believe|feel|know|suspect|doubt|wonder|guess|suppose|reckon|'
    r'maybe|perhaps|probably|apparently|presumably)\b', re.I
)


def _get_sentences(text):
    """Split text into sentences (regex fallback for NSSI)."""
    sents = re.split(r'(?<=[.!?])\s+(?=[A-Z"\'])', text)
    return [s.strip() for s in sents if len(s.strip()) > 10]


def run_self_similarity(text):
    """Compute N-Gram Self-Similarity Index (NSSI).

    Returns dict with individual feature scores and composite NSSI.
    """
    words = text.split()
    word_count = len(words)
    sentences = _get_sentences(text)
    n_sents = max(len(sentences), 1)

    if word_count < 200:
        return {
            'nssi_score': 0.0, 'nssi_signals': 0, 'nssi_active': [],
            'determination': None, 'confidence': 0.0,
            'reason': 'NSSI: text too short for analysis',
            'formulaic_density': 0.0, 'formulaic_weighted': 0.0,
            'power_adj_density': 0.0, 'scare_quote_density': 0.0,
            'emdash_density': 0.0, 'parenthetical_density': 0.0,
            'colon_density': 0.0, 'demonstrative_density': 0.0,
            'transition_density': 0.0, 'causal_density': 0.0,
            'causal_ratio': 0.0, 'this_the_start_rate': 0.0,
            'section_depth': 0, 'sent_length_cv': 0.0,
            'comp_ratio': 0.0, 'hapax_ratio': 0.0,
            'hapax_count': 0, 'unique_words': 0,
            'word_count': word_count, 'sentence_count': n_sents,
            'shuffled_comp_ratio': 0.0, 'structural_compression_delta': 0.0,
        }

    # 1. Formulaic phrase density
    formulaic_raw = 0
    formulaic_weighted = 0.0
    for compiled_pat, weight in _FORMULAIC_PATTERNS:
        hits = len(compiled_pat.findall(text))
        formulaic_raw += hits
        formulaic_weighted += hits * weight
    formulaic_density = formulaic_raw / n_sents
    formulaic_w_density = formulaic_weighted / n_sents

    # 2. Power adjective density
    power_hits = len(_POWER_ADJ.findall(text))
    power_density = power_hits / n_sents

    # 3. Discourse scaffolding
    scare_quotes = len(_SCARE_QUOTE.findall(text))
    emdashes = len(_EM_DASH.findall(text))
    parentheticals = len(_PAREN.findall(text))
    colon_explains = len(_COLON_EXPLAIN.findall(text))

    scare_density = scare_quotes / n_sents
    emdash_density = emdashes / n_sents
    paren_density = parentheticals / n_sents
    colon_density = colon_explains / n_sents

    # 4. Demonstrative monotony
    demo_hits = len(_DEMONSTRATIVE.findall(text))
    demo_density = demo_hits / n_sents

    # 5. Transition connector density
    trans_hits = len(_TRANSITION.findall(text))
    trans_density = trans_hits / n_sents

    # 5b. Causal reasoning deficit
    causal_hits = len(_CAUSAL.findall(text))
    causal_density = causal_hits / n_sents
    causal_ratio = (trans_hits + 1) / (causal_hits + 1)

    # 6. Sentence-start monotony
    starts = [s.split()[0].lower() for s in sentences if s.split()]
    this_the_starts = sum(1 for s in starts if s in ('this', 'the', 'these', 'those'))
    this_the_rate = this_the_starts / n_sents

    # 7. Section hierarchy depth
    headers = re.findall(r'^(\d+(?:\.\d+)*)\s+', text, re.M)
    section_depth = max((h.count('.') + 1 for h in headers), default=0)

    # 8. Sentence length CV
    sent_lens = [len(s.split()) for s in sentences]
    if len(sent_lens) > 2:
        sent_cv = statistics.stdev(sent_lens) / max(statistics.mean(sent_lens), 1)
    else:
        sent_cv = 0.5

    # -- Composite NSSI (12-signal convergence) --
    signals = []

    s1 = min(formulaic_w_density / 0.25, 1.0) if formulaic_w_density >= 0.04 else 0.0
    if s1 > 0: signals.append(('formulaic', s1))

    s2 = min(power_density / 0.30, 1.0) if power_density >= 0.08 else 0.0
    if s2 > 0: signals.append(('power_adj', s2))

    s3 = min(scare_density / 0.40, 1.0) if scare_density >= 0.08 else 0.0
    if s3 > 0: signals.append(('scare_quotes', s3))

    s4 = min(demo_density / 0.12, 1.0) if demo_density >= 0.03 else 0.0
    if s4 > 0: signals.append(('demonstratives', s4))

    s5 = min(trans_density / 0.20, 1.0) if trans_density >= 0.05 else 0.0
    if s5 > 0: signals.append(('transitions', s5))

    scaffold = emdash_density + paren_density + colon_density
    s6 = min(scaffold / 0.60, 1.0) if scaffold >= 0.15 else 0.0
    if s6 > 0: signals.append(('scaffolding', s6))

    s7 = min(this_the_rate / 0.35, 1.0) if this_the_rate >= 0.20 else 0.0
    if s7 > 0: signals.append(('start_monotony', s7))

    s8 = min(section_depth / 4.0, 1.0) if section_depth >= 3 else 0.0
    if s8 > 0: signals.append(('hierarchy', s8))

    s9 = 0.0
    if trans_hits >= 2 and causal_ratio >= 1.5:
        s9 = min((causal_ratio - 1.0) / 3.0, 1.0) * 0.5
    if s9 > 0: signals.append(('causal_deficit', s9))

    # s10: Burstiness
    s10 = 0.0
    if n_sents >= 4 and sent_cv <= 0.35:
        s10 = min((0.35 - sent_cv) / 0.15, 1.0)
    if s10 > 0: signals.append(('low_burstiness', s10))

    # s11: Zlib compression entropy
    text_bytes = text.encode('utf-8')
    original_len = max(len(text_bytes), 1)
    compressed_len = len(zlib.compress(text_bytes))
    comp_ratio = compressed_len / original_len

    s11 = 0.0
    if comp_ratio <= 0.42 and word_count >= 150:
        s11 = min((0.42 - comp_ratio) / 0.08, 1.0)
    if s11 > 0: signals.append(('high_compressibility', s11))

    # s12: Hapax legomena deficit
    clean_words = [w.strip('.,!?"\'():;').lower() for w in words]
    clean_words = [w for w in clean_words if w]
    word_freqs = Counter(clean_words)
    hapax_count = sum(1 for count in word_freqs.values() if count == 1)
    unique_words = len(word_freqs)
    hapax_ratio = hapax_count / unique_words if unique_words > 0 else 0.0

    s12 = 0.0
    if hapax_ratio <= 0.45 and word_count >= 150:
        s12 = min((0.45 - hapax_ratio) / 0.15, 1.0)
    if s12 > 0: signals.append(('hapax_deficit', s12))

    # s13: Structural compression delta (original vs shuffled)
    # Isolates structural redundancy from lexical redundancy.
    # Human text loses more compressibility when shuffled (structural patterns break).
    # LLM text retains compressibility (redundancy is lexical/uniform).
    import random as _random
    shuffled_words = list(clean_words)
    _random.Random(42).shuffle(shuffled_words)  # Deterministic for reproducibility
    shuffled_text = ' '.join(shuffled_words)
    shuffled_bytes = shuffled_text.encode('utf-8')
    shuffled_comp_ratio = len(zlib.compress(shuffled_bytes)) / max(len(shuffled_bytes), 1)
    structural_compression_delta = shuffled_comp_ratio - comp_ratio

    s13 = 0.0
    if structural_compression_delta < 0.03 and word_count >= 150:
        s13 = min((0.03 - structural_compression_delta) / 0.02, 1.0)
    if s13 > 0: signals.append(('low_structural_delta', s13))

    # -- Convergence scoring --
    n_active = len(signals)
    if n_active == 0:
        nssi_score = 0.0
    else:
        mean_strength = sum(s for _, s in signals) / n_active
        convergence = min(n_active / 5.5, 1.0)
        nssi_score = mean_strength * convergence
        if n_active >= 8:
            nssi_score = min(nssi_score * 1.3, 1.0)

    # -- Determination --
    if nssi_score >= 0.70 and n_active >= 7:
        det = 'RED'
        conf = min(0.85, nssi_score)
        reason = f"NSSI convergence (score={nssi_score:.2f}, {n_active} signals)"
    elif nssi_score >= 0.45 and n_active >= 5:
        det = 'AMBER'
        conf = min(0.65, nssi_score)
        reason = f"Elevated NSSI (score={nssi_score:.2f}, {n_active} signals)"
    elif nssi_score >= 0.25 and n_active >= 4:
        det = 'YELLOW'
        conf = min(0.40, nssi_score)
        reason = f"Moderate NSSI (score={nssi_score:.2f}, {n_active} signals)"
    else:
        det = None
        conf = 0.0
        reason = 'NSSI: insufficient signal convergence'

    return {
        'nssi_score': round(nssi_score, 4), 'nssi_signals': n_active,
        'nssi_active': [(name, round(val, 3)) for name, val in signals],
        'determination': det, 'confidence': round(conf, 4), 'reason': reason,
        'formulaic_density': round(formulaic_density, 4),
        'formulaic_weighted': round(formulaic_w_density, 4),
        'power_adj_density': round(power_density, 4),
        'scare_quote_density': round(scare_density, 4),
        'emdash_density': round(emdash_density, 4),
        'parenthetical_density': round(paren_density, 4),
        'colon_density': round(colon_density, 4),
        'demonstrative_density': round(demo_density, 4),
        'transition_density': round(trans_density, 4),
        'causal_density': round(causal_density, 4),
        'causal_ratio': round(causal_ratio, 4),
        'this_the_start_rate': round(this_the_rate, 4),
        'section_depth': section_depth,
        'sent_length_cv': round(sent_cv, 4),
        'comp_ratio': round(comp_ratio, 4),
        'hapax_ratio': round(hapax_ratio, 4),
        'hapax_count': hapax_count,
        'unique_words': unique_words,
        'word_count': word_count, 'sentence_count': n_sents,
        'shuffled_comp_ratio': round(shuffled_comp_ratio, 4),
        'structural_compression_delta': round(structural_compression_delta, 4),
    }

# ==============================================================================
# ANALYZER: CONTINUATION API (DNA-GPT)
# DNA-GPT Divergent Continuation Analysis via LLM API.  Truncates candidate te
# ==============================================================================
def _dna_ngrams(tokens, n):
    """Generate n-grams from token list."""
    return [tuple(tokens[i:i+n]) for i in range(len(tokens) - n + 1)]


def _dna_bscore(original_tokens, regenerated_tokens, ns=(2, 3, 4), weights=(0.25, 0.50, 0.25)):
    """Compute DNA-GPT BScore: weighted n-gram overlap."""
    scores = []
    for n, w in zip(ns, weights):
        orig_ng = set(_dna_ngrams(original_tokens, n))
        regen_ng = set(_dna_ngrams(regenerated_tokens, n))
        if not orig_ng or not regen_ng:
            scores.append(0.0)
            continue
        overlap = len(orig_ng & regen_ng)
        precision = overlap / len(regen_ng) if regen_ng else 0
        recall = overlap / len(orig_ng) if orig_ng else 0
        if precision + recall > 0:
            f1 = 2 * precision * recall / (precision + recall)
        else:
            f1 = 0.0
        scores.append(f1 * w)
    return sum(scores)


def _dna_truncate_text(text, ratio=0.5):
    """Truncate text at sentence boundary. Returns (prefix, continuation)."""
    sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z"\'])', text)
    sentences = [s.strip() for s in sentences if len(s.strip()) > 5]
    if len(sentences) < 4:
        words = text.split()
        mid = int(len(words) * ratio)
        return ' '.join(words[:mid]), ' '.join(words[mid:])
    cut = max(2, int(len(sentences) * ratio))
    return ' '.join(sentences[:cut]), ' '.join(sentences[cut:])


def _dna_call_anthropic(prefix, continuation_length, api_key,
                        model='claude-sonnet-4-20250514', n_samples=3, temperature=0.7):
    """Generate continuations using Anthropic API."""
    try:
        import anthropic
    except ImportError:
        raise ImportError("pip install anthropic  (required for continuation analysis with Anthropic)")
    client = anthropic.Anthropic(api_key=api_key)
    continuations = []
    max_tokens = min(max(continuation_length * 2, 200), 4096)
    for _ in range(n_samples):
        msg = client.messages.create(
            model=model, max_tokens=max_tokens, temperature=temperature,
            messages=[{"role": "user",
                       "content": f"Continue the following text naturally, maintaining the same style, tone, and topic. Do not add any preamble or meta-commentary — just continue writing:\n\n{prefix}"}]
        )
        continuations.append(msg.content[0].text if msg.content else "")
    return continuations


DNA_GPT_STORED_PROMPT_ID = 'pmpt_69a8ff3fd48081938b2de58954245ebf0f4f01733906fee0'


def _dna_call_openai(prefix, continuation_length, api_key,
                     model='gpt-4o-mini', n_samples=3, temperature=0.7):
    """Generate continuations using OpenAI Responses API with stored prompt."""
    try:
        import openai
    except ImportError:
        raise ImportError("pip install openai  (required for continuation analysis with OpenAI)")
    client = openai.OpenAI(api_key=api_key)
    continuations = []
    max_tokens = min(max(continuation_length * 2, 200), 4096)
    for _ in range(n_samples):
        resp = client.responses.create(
            model=model,
            max_output_tokens=max_tokens,
            temperature=temperature,
            instructions={
                "type": "stored_prompt",
                "id": DNA_GPT_STORED_PROMPT_ID,
            },
            input=prefix,
        )
        continuations.append(resp.output_text or "")
    return continuations


def run_continuation_api(text, api_key=None, provider='anthropic', model=None,
                         truncation_ratio=0.5, n_samples=3, temperature=0.7):
    """DNA-GPT divergent continuation analysis via LLM API."""
    word_count = len(text.split())

    if word_count < 150:
        return {'bscore': 0.0, 'bscore_samples': [], 'determination': None,
                'confidence': 0.0, 'reason': 'DNA-GPT: insufficient text',
                'n_samples': 0, 'truncation_ratio': truncation_ratio, 'word_count': word_count}

    if not api_key:
        return {'bscore': 0.0, 'bscore_samples': [], 'determination': None,
                'confidence': 0.0, 'reason': 'DNA-GPT: no API key provided',
                'n_samples': 0, 'truncation_ratio': truncation_ratio, 'word_count': word_count}

    prefix, original_continuation = _dna_truncate_text(text, truncation_ratio)
    if len(original_continuation.split()) < 30:
        return {'bscore': 0.0, 'bscore_samples': [], 'determination': None,
                'confidence': 0.0, 'reason': 'DNA-GPT: continuation too short after truncation',
                'n_samples': 0, 'truncation_ratio': truncation_ratio, 'word_count': word_count}

    orig_tokens = original_continuation.lower().split()
    continuation_word_count = len(orig_tokens)

    if model is None:
        model = 'claude-sonnet-4-20250514' if provider == 'anthropic' else 'gpt-4o-mini'

    try:
        if provider == 'anthropic':
            continuations = _dna_call_anthropic(prefix, continuation_word_count, api_key,
                                                model, n_samples, temperature)
        elif provider == 'openai':
            continuations = _dna_call_openai(prefix, continuation_word_count, api_key,
                                             model, n_samples, temperature)
        else:
            raise ValueError(f"Unknown provider: {provider}. Use 'anthropic' or 'openai'.")
    except Exception as e:
        return {'bscore': 0.0, 'bscore_samples': [], 'determination': None,
                'confidence': 0.0, 'reason': f'DNA-GPT: API call failed ({e})',
                'n_samples': 0, 'truncation_ratio': truncation_ratio, 'word_count': word_count}

    sample_scores = []
    for regen_text in continuations:
        regen_tokens = regen_text.lower().split()
        if len(regen_tokens) < 10:
            continue
        regen_tokens = regen_tokens[:int(len(orig_tokens) * 1.5)]
        bs = _dna_bscore(orig_tokens, regen_tokens)
        sample_scores.append(round(bs, 4))

    if not sample_scores:
        return {'bscore': 0.0, 'bscore_samples': [], 'determination': None,
                'confidence': 0.0, 'reason': 'DNA-GPT: all regenerations failed or too short',
                'n_samples': 0, 'truncation_ratio': truncation_ratio, 'word_count': word_count}

    bscore = statistics.mean(sample_scores)
    bscore_max = max(sample_scores)

    if bscore >= 0.20 and bscore_max >= 0.22:
        det, conf = 'RED', min(0.90, 0.60 + bscore)
        reason = f"DNA-GPT: high continuation overlap (BScore={bscore:.3f}, max={bscore_max:.3f})"
    elif bscore >= 0.12:
        det, conf = 'AMBER', min(0.70, 0.40 + bscore)
        reason = f"DNA-GPT: elevated continuation overlap (BScore={bscore:.3f})"
    elif bscore >= 0.08:
        det, conf = 'YELLOW', min(0.40, 0.20 + bscore)
        reason = f"DNA-GPT: moderate continuation overlap (BScore={bscore:.3f})"
    else:
        det, conf = 'GREEN', max(0.0, 0.10 - bscore)
        reason = f"DNA-GPT: low continuation overlap (BScore={bscore:.3f}) -- likely human"

    return {
        'bscore': round(bscore, 4), 'bscore_max': round(bscore_max, 4),
        'bscore_samples': sample_scores, 'determination': det,
        'confidence': round(conf, 4), 'reason': reason,
        'n_samples': len(sample_scores), 'truncation_ratio': truncation_ratio,
        'continuation_words': continuation_word_count, 'word_count': word_count,
    }

# ==============================================================================
# ANALYZER: CONTINUATION LOCAL (DNA-GPT PROXY)
# ==============================================================================
"""DNA-GPT Proxy -- zero-LLM divergent continuation analysis.

Uses a backoff n-gram language model as surrogate for LLM regeneration.
Ref: Yang et al. (2024) "DNA-GPT" (ICLR 2024)
Ref: Li et al. (2004) "The Similarity Metric" (NCD theory)
"""


_TOKEN_RE = re.compile(r'\w+|[^\w\s]')


def _proxy_tokenize(text):
    """Tokenize for n-gram LM. Returns lowercased word/punct tokens."""
    return _TOKEN_RE.findall(text.lower())


class _BackoffNGramLM:
    """Simple backoff n-gram language model for DNA-GPT proxy regeneration."""

    def __init__(self, order=5, alpha=0.1):
        self.order = max(order, 1)
        self.alpha = alpha
        self.tables = [defaultdict(Counter) for _ in range(self.order)]
        self.vocab = set()

    def fit(self, texts):
        """Train on an iterable of text strings."""
        bos = ['<s>'] * (self.order - 1)
        for text in texts:
            toks = bos + _proxy_tokenize(text) + ['</s>']
            self.vocab.update(toks)
            for i in range(self.order - 1, len(toks)):
                for ctx_len in range(self.order):
                    ctx = tuple(toks[i - ctx_len:i]) if ctx_len else ()
                    self.tables[ctx_len][ctx][toks[i]] += 1

    def _counts(self, context):
        """Get counts for context with backoff."""
        max_ctx = min(len(context), self.order - 1)
        for ctx_len in range(max_ctx, -1, -1):
            ctx = tuple(context[-ctx_len:]) if ctx_len else ()
            counts = self.tables[ctx_len].get(ctx)
            if counts:
                return counts
        return Counter({t: 1 for t in self.vocab}) if self.vocab else Counter({'</s>': 1})

    def sample_next(self, context):
        """Sample a single next token given context."""
        import random as _random
        counts = self._counts(context)
        items = list(counts.items())
        total = sum(c for _, c in items)
        r = _random.random() * total
        acc = 0.0
        for tok, c in items:
            acc += c
            if acc >= r:
                return tok
        return items[-1][0]

    def logprob(self, token, context):
        """Log-probability of token given context (with Laplace smoothing)."""
        counts = self._counts(context)
        total = sum(counts.values())
        vocab_size = max(len(self.vocab), 1)
        p = (counts.get(token, 0) + self.alpha) / (total + self.alpha * vocab_size)
        return math.log(p)

    def sample_suffix(self, prefix_tokens, length):
        """Generate a continuation of `length` tokens from prefix context."""
        ctx = ['<s>'] * (self.order - 1) + list(prefix_tokens)
        out = []
        for _ in range(length):
            tok = self.sample_next(ctx)
            if tok == '</s>':
                break
            out.append(tok)
            ctx.append(tok)
        return out


def _calculate_ncd(prefix, suffix):
    """Normalized Compression Distance between prefix and suffix."""
    x = prefix.encode('utf-8')
    y = suffix.encode('utf-8')
    xy = x + b' ' + y

    c_x = len(zlib.compress(x))
    c_y = len(zlib.compress(y))
    c_xy = len(zlib.compress(xy))

    denom = max(c_x, c_y)
    if denom == 0:
        return 0.0
    return (c_xy - min(c_x, c_y)) / denom


def _internal_ngram_overlap(prefix_tokens, suffix_tokens, ns=(3, 4)):
    """Fraction of suffix n-grams that appear in prefix (echo effect)."""
    if not suffix_tokens:
        return 0.0

    total_weight = 0.0
    weighted_overlap = 0.0

    for n in ns:
        pfx_ng = set(_dna_ngrams(prefix_tokens, n))
        sfx_ng = set(_dna_ngrams(suffix_tokens, n))
        if not sfx_ng:
            continue
        w = n * math.log(n) if n > 1 else 1.0
        overlap = len(pfx_ng & sfx_ng) / len(sfx_ng)
        weighted_overlap += w * overlap
        total_weight += w

    return weighted_overlap / total_weight if total_weight else 0.0


def _repeated_ngram_rate(tokens, n=4):
    """Fraction of n-grams that are repetitions (monotonicity signal)."""
    count = max(0, len(tokens) - n + 1)
    if count == 0:
        return 0.0
    grams = [tuple(tokens[i:i + n]) for i in range(count)]
    return 1.0 - len(set(grams)) / len(grams)


def _conditional_surprisal(lm, prefix_tokens, suffix_tokens):
    """Mean negative log-probability of suffix given prefix under LM."""
    ctx = ['<s>'] * (lm.order - 1) + list(prefix_tokens)
    total = 0.0
    for tok in suffix_tokens:
        total -= lm.logprob(tok, ctx)
        ctx.append(tok)
    return total / max(1, len(suffix_tokens))


def _type_token_ratio(tokens):
    """Type-Token Ratio: vocabulary richness."""
    if not tokens:
        return 0.0
    return len(set(tokens)) / len(tokens)


def _multi_segment_ncd(text, n_segments=4):
    """Compute NCD between all pairs of text segments.

    Low variance = all segments are similarly redundant = AI signal.
    """
    sentences = re.split(r'(?<=[.!?])\s+', text)
    if len(sentences) < n_segments * 2:
        return {'ncd_mean': 0.0, 'ncd_variance': 0.0, 'ncd_min': 0.0, 'n_pairs': 0}

    seg_size = len(sentences) // n_segments
    segments = []
    for i in range(n_segments):
        start = i * seg_size
        end = start + seg_size if i < n_segments - 1 else len(sentences)
        segments.append(' '.join(sentences[start:end]))

    ncds = []
    for i in range(len(segments)):
        for j in range(i + 1, len(segments)):
            ncd = _calculate_ncd(segments[i], segments[j])
            ncds.append(ncd)

    if not ncds:
        return {'ncd_mean': 0.0, 'ncd_variance': 0.0, 'ncd_min': 0.0, 'n_pairs': 0}

    return {
        'ncd_mean': round(statistics.mean(ncds), 4),
        'ncd_variance': round(statistics.variance(ncds) if len(ncds) >= 2 else 0.0, 6),
        'ncd_min': round(min(ncds), 4),
        'n_pairs': len(ncds),
    }


def _surprisal_improvement_curve(lm, full_tokens, splits=(0.25, 0.50, 0.75)):
    """Measure how conditional surprisal changes with increasing prefix length.

    Human text: improvement_rate ~0.15-0.40 (surprisal drops with more context).
    AI text: improvement_rate ~0.00-0.10 (already predictable from any prefix).
    """
    n = len(full_tokens)
    if n < 40:
        return {'surprisal_curve': [], 'improvement_rate': 0.0}

    tail_start = int(n * 0.75)
    tail_tokens = full_tokens[tail_start:]

    surprisals = []
    for split in splits:
        prefix_end = int(n * split)
        if prefix_end < 10:
            continue
        prefix = full_tokens[:prefix_end]
        surp = _conditional_surprisal(lm, prefix, tail_tokens)
        surprisals.append((split, round(surp, 4)))

    if len(surprisals) >= 2:
        first = surprisals[0][1]
        last = surprisals[-1][1]
        improvement_rate = (first - last) / max(first, 1e-6)
    else:
        improvement_rate = 0.0

    return {
        'surprisal_curve': surprisals,
        'improvement_rate': round(improvement_rate, 4),
    }


def run_continuation_local(text, gamma=0.5, K=32, order=5):
    """Zero-LLM DNA-GPT proxy via backoff n-gram language model."""
    word_count = len(text.split())

    if word_count < 80:
        return {
            'bscore': 0.0, 'bscore_samples': [], 'determination': None,
            'confidence': 0.0, 'reason': 'DNA-GPT-Local: insufficient text (<80 words)',
            'n_samples': 0, 'truncation_ratio': gamma, 'word_count': word_count,
            'proxy_features': {},
        }

    prefix_text, suffix_text = _dna_truncate_text(text, gamma)
    prefix_tokens = _proxy_tokenize(prefix_text)
    suffix_tokens = _proxy_tokenize(suffix_text)

    if len(suffix_tokens) < 20:
        return {
            'bscore': 0.0, 'bscore_samples': [], 'determination': None,
            'confidence': 0.0, 'reason': 'DNA-GPT-Local: suffix too short after split',
            'n_samples': 0, 'truncation_ratio': gamma, 'word_count': word_count,
            'proxy_features': {},
        }

    lm = _BackoffNGramLM(order=order)
    lm.fit([prefix_text])

    sample_scores = []
    for _ in range(K):
        regen = lm.sample_suffix(prefix_tokens, len(suffix_tokens))
        if len(regen) < 10:
            continue
        bs = _dna_bscore(suffix_tokens, regen)
        sample_scores.append(round(bs, 4))

    if not sample_scores:
        sample_scores = [0.0]

    bscore = statistics.mean(sample_scores)
    bscore_max = max(sample_scores)

    ncd = _calculate_ncd(prefix_text, suffix_text)
    internal_overlap = _internal_ngram_overlap(prefix_tokens, suffix_tokens)
    cond_surp = _conditional_surprisal(lm, prefix_tokens, suffix_tokens)
    repeat4 = _repeated_ngram_rate(suffix_tokens, 4)
    ttr = _type_token_ratio(suffix_tokens)

    proxy_features = {
        'ncd': round(ncd, 4),
        'internal_overlap': round(internal_overlap, 4),
        'cond_surprisal': round(cond_surp, 4),
        'repeat4': round(repeat4, 4),
        'ttr': round(ttr, 4),
    }

    # Composite scoring
    ncd_signal = max(0.0, (1.0 - ncd) / 0.15)
    overlap_signal = max(0.0, min(1.0, (internal_overlap - 0.05) / 0.30))
    repeat_signal = max(0.0, min(1.0, repeat4 / 0.15))
    ttr_signal = max(0.0, min(1.0, (0.55 - ttr) / 0.20))
    bscore_signal = min(1.0, bscore / 0.15)

    composite = (
        0.30 * bscore_signal +
        0.25 * ncd_signal +
        0.20 * overlap_signal +
        0.10 * repeat_signal +
        0.10 * ttr_signal +
        0.05 * max(0.0, min(1.0, (5.0 - cond_surp) / 3.0))
    )

    proxy_features['composite'] = round(composite, 4)

    # FEAT 6: Multi-segment NCD matrix (diagnostic only)
    multi_ncd = _multi_segment_ncd(text)
    proxy_features['ncd_matrix_mean'] = multi_ncd['ncd_mean']
    proxy_features['ncd_matrix_variance'] = multi_ncd['ncd_variance']
    proxy_features['ncd_matrix_min'] = multi_ncd['ncd_min']

    # FEAT 2: Cross-prefix surprisal improvement curve (diagnostic only)
    all_tokens = _proxy_tokenize(text)
    surp_curve = _surprisal_improvement_curve(lm, all_tokens)
    proxy_features['surprisal_curve'] = surp_curve['surprisal_curve']
    proxy_features['improvement_rate'] = surp_curve['improvement_rate']

    if composite >= 0.60 and (ncd_signal >= 0.4 or overlap_signal >= 0.5):
        det = 'RED'
        conf = min(0.80, 0.50 + composite * 0.30)
        reason = (f"DNA-GPT-Local: high self-consistency "
                  f"(composite={composite:.2f}, NCD={ncd:.3f}, "
                  f"overlap={internal_overlap:.3f})")
    elif composite >= 0.40:
        det = 'AMBER'
        conf = min(0.60, 0.30 + composite * 0.30)
        reason = (f"DNA-GPT-Local: elevated predictability "
                  f"(composite={composite:.2f}, NCD={ncd:.3f})")
    elif composite >= 0.25:
        det = 'YELLOW'
        conf = min(0.35, 0.15 + composite * 0.20)
        reason = (f"DNA-GPT-Local: moderate self-consistency "
                  f"(composite={composite:.2f})")
    else:
        det = None
        conf = 0.0
        reason = (f"DNA-GPT-Local: low predictability "
                  f"(composite={composite:.2f}) -- likely human")

    return {
        'bscore': round(bscore, 4),
        'bscore_max': round(bscore_max, 4),
        'bscore_samples': sample_scores,
        'determination': det,
        'confidence': round(conf, 4),
        'reason': reason,
        'n_samples': len(sample_scores),
        'truncation_ratio': gamma,
        'continuation_words': len(suffix_tokens),
        'word_count': word_count,
        'proxy_features': proxy_features,
    }


def run_continuation_local_multi(text, gammas=(0.3, 0.5, 0.7), K=16, order=5):
    """Multi-truncation DNA-GPT local proxy.

    Runs continuation analysis at multiple truncation ratios and measures
    stability of the composite score. High stability (low variance) across
    truncation points is an AI signal (TDT, West et al., 2025).
    """
    composites = []
    full_result = None

    for gamma in gammas:
        result = run_continuation_local(text, gamma=gamma, K=K, order=order)
        comp = result.get('proxy_features', {}).get('composite', 0.0)
        composites.append(comp)
        if gamma == 0.5:
            full_result = result

    if full_result is None:
        full_result = result

    if len(composites) >= 2:
        comp_mean = statistics.mean(composites)
        comp_var = statistics.variance(composites)
        # Normalize: variance of [0,1]-bounded values rarely exceeds 0.08
        stability = max(0.0, 1.0 - (comp_var / 0.08))
    else:
        comp_mean = composites[0] if composites else 0.0
        comp_var = 0.0
        stability = 0.0

    full_result['proxy_features']['multi_composites'] = [round(c, 4) for c in composites]
    full_result['proxy_features']['composite_variance'] = round(comp_var, 6)
    full_result['proxy_features']['composite_stability'] = round(stability, 4)

    # Stability boosts composite when it agrees with the primary signal
    if stability >= 0.75 and comp_mean >= 0.30:
        boosted = min(full_result['proxy_features']['composite'] + 0.10, 1.0)
        full_result['proxy_features']['composite'] = round(boosted, 4)

    return full_result


# ==============================================================================
# ANALYZER: PERPLEXITY
# ==============================================================================
"""Local perplexity scoring via distilgpt2.

AI text has low perplexity (< 20); human text typically > 35.
Ref: GLTR (Gehrmann et al. 2019), DetectGPT (Mitchell et al. 2023)
"""



def run_perplexity(text):
    """Calculate perplexity with DivEye variance and volatility decay.

    Extended with DivEye surprisal diversity (Basani & Chen, ICML 2025)
    and late-stage volatility decay (Sun et al., arXiv:2601.04833).

    Returns dict with perplexity, surprisal_variance, volatility_decay,
    determination, and confidence.
    """
    _ensure_perplexity()
    _ppl_empty = {
        'perplexity': 0.0, 'determination': None, 'confidence': 0.0,
        'surprisal_variance': 0.0, 'first_half_variance': 0.0,
        'second_half_variance': 0.0, 'volatility_decay': 1.0, 'n_tokens': 0,
        'comp_ratio': 0.0, 'zlib_normalized_ppl': 0.0, 'comp_ppl_ratio': 0.0,
        '_token_losses': [],
    }
    if not HAS_PERPLEXITY:
        return {**_ppl_empty, 'reason': 'Perplexity scoring unavailable (transformers/torch not installed)'}

    words = text.split()
    if len(words) < 50:
        return {**_ppl_empty, 'reason': 'Perplexity: text too short'}

    encodings = _PPL_TOKENIZER(text, return_tensors='pt', truncation=True,
                                max_length=1024)
    input_ids = encodings.input_ids

    if input_ids.size(1) < 10:
        return {**_ppl_empty, 'reason': 'Perplexity: too few tokens after encoding'}

    with _torch.no_grad():
        outputs = _PPL_MODEL(input_ids, labels=input_ids)

    # ── Per-token surprisal extraction ──────────────────────────────────
    # Compute per-token cross-entropy for the full surprisal distribution.
    # The mean gives us perplexity; the variance and decay give DivEye features.
    shift_logits = outputs.logits[..., :-1, :].contiguous()
    shift_labels = input_ids[..., 1:].contiguous()
    token_losses = _torch.nn.functional.cross_entropy(
        shift_logits.view(-1, shift_logits.size(-1)),
        shift_labels.view(-1),
        reduction='none',
    ).float().cpu()

    n_tok = token_losses.size(0)
    mean_loss = token_losses.mean().item()
    ppl = _torch.exp(_torch.tensor(mean_loss)).item()

    # ── DivEye: surprisal variance (Basani & Chen, ICML 2025) ──────────
    surprisal_var = token_losses.std().item() if n_tok > 1 else 0.0

    # ── Volatility decay (Sun et al., arXiv:2601.04833) ────────────────
    # AI text "settles" — first half is noisier than second half.
    if n_tok >= 20:
        first_half = token_losses[:n_tok // 2]
        second_half = token_losses[n_tok // 2:]
        first_half_var = first_half.std().item() if first_half.numel() > 1 else 0.0
        second_half_var = second_half.std().item() if second_half.numel() > 1 else 0.0
        volatility_decay = first_half_var / max(second_half_var, 1e-6)
    else:
        first_half_var = 0.0
        second_half_var = 0.0
        volatility_decay = 1.0

    # ── Determination: Layer 1 — original perplexity thresholds ────────
    if ppl <= 15.0:
        det = 'AMBER'
        conf = min(0.65, (20.0 - ppl) / 20.0)
        reason = f"Low perplexity ({ppl:.1f}): highly predictable text"
    elif ppl <= 25.0:
        det = 'YELLOW'
        conf = min(0.35, (30.0 - ppl) / 30.0)
        reason = f"Moderate perplexity ({ppl:.1f}): somewhat predictable"
    else:
        det = None
        conf = 0.0
        reason = f"Normal perplexity ({ppl:.1f}): consistent with human text"

    # ── Determination: Layer 2 — DivEye + Volatility can upgrade ───────
    # Low variance = uniform token selection (AI signature).
    # High decay = text stabilizes in second half (AI signature).
    # Thresholds are provisional — must recalibrate in v0.64 data pass.
    diveye_signal = surprisal_var < 2.0 and n_tok >= 30
    volatility_signal = volatility_decay > 1.5 and n_tok >= 40

    if diveye_signal and volatility_signal:
        # Both variance and decay signal — strong compound evidence
        if det is None:
            det = 'YELLOW'
            conf = min(0.40, 0.20 + (2.0 - surprisal_var) * 0.05
                       + (volatility_decay - 1.0) * 0.05)
            reason = (f"Surprisal uniformity (var={surprisal_var:.2f}, "
                      f"decay={volatility_decay:.2f}): machine rhythm detected")
        elif det == 'YELLOW':
            det = 'AMBER'
            conf = min(0.65, conf + 0.15)
            reason += (f" + DivEye(var={surprisal_var:.2f}, "
                       f"decay={volatility_decay:.2f})")
        elif det == 'AMBER':
            conf = min(0.80, conf + 0.10)
            reason += (f" + DivEye(var={surprisal_var:.2f}, "
                       f"decay={volatility_decay:.2f})")
    elif diveye_signal or volatility_signal:
        # One signal alone — supporting evidence only
        if det is not None:
            conf = min(conf + 0.05, 0.70)
            if diveye_signal:
                reason += f" + low_variance({surprisal_var:.2f})"
            else:
                reason += f" + volatility_decay({volatility_decay:.2f})"

    # ── FEAT 7: Compression-perplexity divergence ──────────────────────
    # Zlib-normalized perplexity (Carlini et al. 2021, Shi et al. 2024):
    # low value = text is both predictable AND compressible = AI zone.
    text_bytes = text.encode('utf-8')
    comp_len = len(zlib.compress(text_bytes))
    comp_ratio_ppl = comp_len / max(len(text_bytes), 1)
    zlib_normalized_ppl = ppl * comp_ratio_ppl
    comp_ppl_ratio = comp_ratio_ppl / max(ppl / 100.0, 0.01)

    zlib_ppl_signal = zlib_normalized_ppl < 8.0 and n_tok >= 30
    if zlib_ppl_signal:
        if det is None:
            det = 'YELLOW'
            conf = min(0.35, 0.15 + (8.0 - zlib_normalized_ppl) * 0.02)
            reason = f"Zlib-normalized PPL ({zlib_normalized_ppl:.1f}): predictable and compressible"
        elif det in ('YELLOW', 'AMBER'):
            conf = min(conf + 0.05, 0.80)
            reason += f" + zlib_ppl({zlib_normalized_ppl:.1f})"

    return {
        'perplexity': round(ppl, 2),
        'surprisal_variance': round(surprisal_var, 4),
        'first_half_variance': round(first_half_var, 4),
        'second_half_variance': round(second_half_var, 4),
        'volatility_decay': round(volatility_decay, 4),
        'n_tokens': n_tok,
        'comp_ratio': round(comp_ratio_ppl, 4),
        'zlib_normalized_ppl': round(zlib_normalized_ppl, 2),
        'comp_ppl_ratio': round(comp_ppl_ratio, 4),
        'determination': det,
        'confidence': round(conf, 4),
        'reason': reason,
        '_token_losses': token_losses.tolist(),
    }


# ==============================================================================
# ANALYZER: STYLOMETRY
# ==============================================================================
"""Topic-scrubbed stylometric feature extraction.

Masks topical content before computing style features to reduce topic leakage.
"""


# Topic masking patterns
_TOPIC_URL_RE = re.compile(r'https?://\S+|www\.\S+', re.I)
_TOPIC_EMAIL_RE = re.compile(r'\b[\w.+-]+@[\w-]+\.[\w.-]+\b')
_TOPIC_DATE_RE = re.compile(
    r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b'
    r'|\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b'
    r'|\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*[\s,]+\d{1,2},?\s*\d{4}\b',
    re.I,
)
_TOPIC_FILENAME_RE = re.compile(r'\b\w+\.\w{2,4}\b')
_TOPIC_VERSION_RE = re.compile(r'\bv?\d+\.\d+(?:\.\d+)*\b', re.I)
_TOPIC_NUMBER_RE = re.compile(r'\b\d{3,}\b')
_TOPIC_CAMELCASE_RE = re.compile(r'\b[a-z]+[A-Z]\w+\b|\b[A-Z][a-z]+[A-Z]\w*\b')
_TOPIC_ALLCAPS_RE = re.compile(r'\b[A-Z]{2,}\b')


def mask_topical_content(text):
    """Replace topical tokens with placeholders. Returns (masked_text, mask_count)."""
    count = 0
    for pattern, repl in [
        (_TOPIC_URL_RE, ' _URL_ '),
        (_TOPIC_EMAIL_RE, ' _EMAIL_ '),
        (_TOPIC_DATE_RE, ' _DATE_ '),
        (_TOPIC_FILENAME_RE, ' _FILE_ '),
        (_TOPIC_VERSION_RE, ' _VER_ '),
        (_TOPIC_NUMBER_RE, ' _NUM_ '),
        (_TOPIC_CAMELCASE_RE, ' _IDENT_ '),
        (_TOPIC_ALLCAPS_RE, ' _ACRO_ '),
    ]:
        hits = len(pattern.findall(text))
        if hits:
            text = pattern.sub(repl, text)
            count += hits
    return text, count


def extract_stylometric_features(text, masked_text=None):
    """Extract topic-invariant stylometric features.

    Returns dict with char n-gram profile, function word ratio, punctuation
    bigrams, sentence length dispersion, type-token ratio, etc.
    """
    if masked_text is None:
        masked_text, _ = mask_topical_content(text)

    words = masked_text.lower().split()
    n_words = max(len(words), 1)

    # Character 4-grams
    lower_masked = masked_text.lower()
    char4 = Counter()
    for i in range(len(lower_masked) - 3):
        gram = lower_masked[i:i+4]
        if not gram.startswith('_'):
            char4[gram] += 1
    total_4grams = max(sum(char4.values()), 1)
    char_ngram_profile = {g: c / total_4grams for g, c in char4.most_common(50)}

    # Function word ratio
    fw_count = sum(1 for w in words if w in ENGLISH_FUNCTION_WORDS)
    function_word_ratio = fw_count / n_words

    # Punctuation bigrams
    punct_chars = re.findall(r'[^\w\s]', text)
    punct_bigrams = Counter()
    for i in range(len(punct_chars) - 1):
        punct_bigrams[punct_chars[i] + punct_chars[i+1]] += 1

    # Sentence length dispersion
    sentences = get_sentences(text)
    sent_lengths = [len(s.split()) for s in sentences if s.strip()]
    if len(sent_lengths) >= 2:
        mean_sl = statistics.mean(sent_lengths)
        std_sl = statistics.stdev(sent_lengths)
        sent_length_dispersion = std_sl / max(mean_sl, 1)
    else:
        sent_length_dispersion = 0.0

    # Type-token ratio
    orig_words = re.findall(r'\w+', text.lower())
    n_orig = max(len(orig_words), 1)
    type_token_ratio = len(set(orig_words)) / n_orig

    # Average word length
    word_lengths = [len(w) for w in orig_words]
    avg_word_length = statistics.mean(word_lengths) if word_lengths else 0

    # Short word ratio
    short_words = sum(1 for w in orig_words if len(w) <= 3)
    short_word_ratio = short_words / n_orig

    return {
        'char_ngram_profile': char_ngram_profile,
        'function_word_ratio': round(function_word_ratio, 4),
        'punct_bigrams': dict(punct_bigrams.most_common(20)),
        'sent_length_dispersion': round(sent_length_dispersion, 4),
        'type_token_ratio': round(type_token_ratio, 4),
        'avg_word_length': round(avg_word_length, 2),
        'short_word_ratio': round(short_word_ratio, 4),
    }


# ==============================================================================
# ANALYZER: WINDOWING
# ==============================================================================
"""Windowed scoring -- detect mixed human+AI content via per-window analysis.

Ref: M4GT-Bench (Wang et al. 2024) -- mixed detection as separate task.
"""



def detect_changepoint(feature_sequence, threshold=3.0):
    """CUSUM changepoint detection on a 1D feature sequence.

    Returns dict with changepoint location and effect size, or None.
    """
    if len(feature_sequence) < 6:
        return None

    n = len(feature_sequence)
    mean_all = statistics.mean(feature_sequence)

    cusum = [0.0]
    for val in feature_sequence:
        cusum.append(cusum[-1] + (val - mean_all))

    max_dev = 0.0
    best_idx = None
    for i in range(1, n):
        dev = abs(cusum[i])
        if dev > max_dev:
            max_dev = dev
            best_idx = i

    if best_idx is None or best_idx < 2 or best_idx > n - 2:
        return None

    before = feature_sequence[:best_idx]
    after = feature_sequence[best_idx:]
    if len(before) < 2 or len(after) < 2:
        return None

    mean_before = statistics.mean(before)
    mean_after = statistics.mean(after)
    pooled_std = statistics.stdev(feature_sequence)

    if pooled_std < 1e-6:
        return None

    effect_size = abs(mean_after - mean_before) / pooled_std
    if effect_size < threshold:
        return None

    return {
        'changepoint_sentence': best_idx,
        'effect_size': round(effect_size, 3),
        'mean_before': round(mean_before, 4),
        'mean_after': round(mean_after, 4),
    }


def score_windows(text, window_size=5, stride=2):
    """Score text in overlapping sentence windows.

    Returns dict with per-window scores, max/mean/variance, hot span, mixed signal.
    """
    sentences = get_sentences(text)
    if len(sentences) < window_size:
        return {
            'windows': [],
            'max_window_score': 0.0,
            'mean_window_score': 0.0,
            'window_variance': 0.0,
            'hot_span_length': 0,
            'n_windows': 0,
            'mixed_signal': False,
            'fw_trajectory_cv': 0.0,
            'comp_trajectory_mean': 0.0,
            'comp_trajectory_cv': 0.0,
            'changepoint': None,
        }

    windows = []
    fw_ratios = []
    comp_ratios = []
    for start in range(0, len(sentences) - window_size + 1, stride):
        end = start + window_size
        window_text = ' '.join(sentences[start:end])
        window_words = window_text.split()
        n_w = max(len(window_words), 1)

        formulaic_count = sum(
            len(compiled_pat.findall(window_text))
            for compiled_pat, _weight in _FORMULAIC_PATTERNS
        )
        formulaic_density = formulaic_count / (n_w / 100)

        trans_hits = len(_TRANSITION.findall(window_text))
        trans_density = trans_hits / (n_w / 100)

        power_hits = len(_POWER_ADJ.findall(window_text))
        power_density = power_hits / (n_w / 100)

        fw = sum(1 for w in window_words if w.lower() in ENGLISH_FUNCTION_WORDS)
        fw_ratio = fw / n_w
        fw_ratios.append(fw_ratio)

        # FEAT 4: Per-window compression ratio
        window_bytes = window_text.encode('utf-8')
        if len(window_bytes) > 20:
            window_comp = len(zlib.compress(window_bytes)) / len(window_bytes)
        else:
            window_comp = 0.5
        comp_ratios.append(window_comp)

        w_sent_lengths = [len(s.split()) for s in sentences[start:end] if s.strip()]
        if len(w_sent_lengths) >= 2:
            w_mean = statistics.mean(w_sent_lengths)
            w_std = statistics.stdev(w_sent_lengths)
            w_cv = w_std / max(w_mean, 1)
        else:
            w_cv = 0.5

        ai_indicators = 0.0
        if formulaic_density > 2.0:
            ai_indicators += min(formulaic_density / 5.0, 0.3)
        if trans_density > 3.0:
            ai_indicators += min(trans_density / 8.0, 0.2)
        if power_density > 1.5:
            ai_indicators += min(power_density / 4.0, 0.2)
        if w_cv < 0.25 and len(w_sent_lengths) >= 3:
            ai_indicators += 0.15
        if fw_ratio < 0.12:
            ai_indicators += 0.15

        window_score = min(ai_indicators, 1.0)

        windows.append({
            'start': start,
            'end': end,
            'score': round(window_score, 3),
            'formulaic': round(formulaic_density, 2),
            'transitions': round(trans_density, 2),
            'sent_cv': round(w_cv, 3),
        })

    scores = [w['score'] for w in windows]
    max_score = max(scores) if scores else 0.0
    mean_score = statistics.mean(scores) if scores else 0.0
    variance = statistics.variance(scores) if len(scores) >= 2 else 0.0

    hot_threshold = 0.30
    hot_span = 0
    current_span = 0
    for s in scores:
        if s >= hot_threshold:
            current_span += 1
            hot_span = max(hot_span, current_span)
        else:
            current_span = 0

    mixed_signal = variance >= 0.02 and max_score >= 0.30 and mean_score < 0.50

    # FEAT 3: Function word trajectory CV (diagnostic only)
    if len(fw_ratios) >= 3:
        fw_trajectory_cv = statistics.stdev(fw_ratios) / max(statistics.mean(fw_ratios), 0.01)
    else:
        fw_trajectory_cv = 0.0

    # FEAT 4: Windowed compression profile (diagnostic only)
    if len(comp_ratios) >= 3:
        comp_trajectory_cv = statistics.stdev(comp_ratios) / max(statistics.mean(comp_ratios), 0.01)
        comp_trajectory_mean = statistics.mean(comp_ratios)
    else:
        comp_trajectory_cv = 0.0
        comp_trajectory_mean = 0.0

    # FEAT 9: Changepoint detection via CUSUM
    changepoint = detect_changepoint(scores) if len(scores) >= 6 else None

    return {
        'windows': windows,
        'max_window_score': round(max_score, 3),
        'mean_window_score': round(mean_score, 3),
        'window_variance': round(variance, 4),
        'hot_span_length': hot_span,
        'n_windows': len(windows),
        'mixed_signal': mixed_signal,
        'fw_trajectory_cv': round(fw_trajectory_cv, 4),
        'comp_trajectory_mean': round(comp_trajectory_mean, 4),
        'comp_trajectory_cv': round(comp_trajectory_cv, 4),
        'changepoint': changepoint,
    }


def score_surprisal_windows(token_losses, window_size=64, stride=32):
    """Windowed surprisal trajectory analysis (FEAT 10).

    Computes how mean surprisal and its variance evolve across the text.
    Returns trajectory_cv (coefficient of variation of window means),
    var_of_var (variation of per-window variance), and a composite
    stationarity_score.
    """
    if not token_losses or len(token_losses) < window_size:
        return {
            'surprisal_trajectory_cv': 0.0,
            'surprisal_var_of_var': 0.0,
            'surprisal_stationarity': 0.0,
            'n_surprisal_windows': 0,
        }

    window_means = []
    window_vars = []
    for start in range(0, len(token_losses) - window_size + 1, stride):
        chunk = token_losses[start:start + window_size]
        w_mean = sum(chunk) / len(chunk)
        w_var = sum((x - w_mean) ** 2 for x in chunk) / len(chunk)
        window_means.append(w_mean)
        window_vars.append(w_var)

    if len(window_means) < 2:
        return {
            'surprisal_trajectory_cv': 0.0,
            'surprisal_var_of_var': 0.0,
            'surprisal_stationarity': 0.0,
            'n_surprisal_windows': len(window_means),
        }

    mean_of_means = sum(window_means) / len(window_means)
    std_of_means = (sum((m - mean_of_means) ** 2 for m in window_means) / len(window_means)) ** 0.5
    trajectory_cv = std_of_means / max(mean_of_means, 1e-6)

    mean_of_vars = sum(window_vars) / len(window_vars)
    std_of_vars = (sum((v - mean_of_vars) ** 2 for v in window_vars) / len(window_vars)) ** 0.5
    var_of_var = std_of_vars / max(mean_of_vars, 1e-6)

    stationarity = (1 - min(trajectory_cv, 1.0)) * (1 - min(var_of_var, 1.0))

    return {
        'surprisal_trajectory_cv': round(trajectory_cv, 4),
        'surprisal_var_of_var': round(var_of_var, 4),
        'surprisal_stationarity': round(stationarity, 4),
        'n_surprisal_windows': len(window_means),
    }


# ==============================================================================
# LEXICON PACKS
# ==============================================================================
"""
BEET Lexicon Packs v1.0
═══════════════════════
Externalized, versioned detection vocabulary organized into typed families
with independent weights and caps per family.

Design rationale (from roadmap):
  "Flat word bags are where detectors go to die."
  Each pack is a named family with:
    - A semantic category (obligation, schema, gherkin, etc.)
    - Patterns (regex or literal keywords)
    - A weight (contribution strength per hit)
    - A cap (maximum contribution from this family alone)
    - A target layer (which pipeline layer consumes it)
    - Mode eligibility (task_prompt, generic_aigt, or both)

Usage:
    from lexicon_packs import PACK_REGISTRY, get_packs_for_layer, score_packs

    # Get all packs feeding prompt_signature layer
    ps_packs = get_packs_for_layer('prompt_signature')

    # Score text against packs
    results = score_packs(text, ps_packs)
    # → {'obligation': {'hits': 5, 'score': 0.35, 'capped': 0.30, 'matches': [...]}, ...}
"""


_LEXICON_VERSION = '1.0.0'
__pack_date__ = '2026-03-04'


# ══════════════════════════════════════════════════════════════════════════════
# PACK DATA STRUCTURE
# ══════════════════════════════════════════════════════════════════════════════

@dataclass(frozen=True)
class LexiconPack:
    """A single vocabulary family with scoring parameters.

    Attributes:
        name:       Unique identifier (e.g., 'obligation', 'schema_json')
        category:   Semantic grouping ('constraint', 'schema', 'exec_spec',
                    'instruction', 'format', 'discourse')
        target_layer: Which pipeline layer consumes this ('prompt_signature',
                      'voice_dissonance', 'instruction_density', 'self_similarity', 'multi')
        mode:       'task_prompt', 'generic_aigt', or 'both'
        patterns:   List of (regex_string, per_hit_weight) tuples.
                    Regexes are compiled with re.IGNORECASE by default.
        keywords:   Frozenset of literal lowercase keywords (fast path).
                    Checked via word-boundary match.
        uppercase_keywords: Frozenset of UPPERCASE keywords where case matters
                    (e.g., BCP 14 normative forms). Matched case-sensitively.
        family_weight: Global multiplier applied to this pack's total score.
        family_cap:   Maximum contribution (0.0–1.0) from this pack alone.
        description:  Human-readable purpose.
        source_refs:  Citation keys for provenance.
        version:      Pack-level version string.
    """
    name: str
    category: str
    target_layer: str  # e.g. 'prompt_signature', 'voice_dissonance', 'instruction_density'
    mode: str = 'both'
    patterns: Tuple[Tuple[str, float], ...] = ()
    keywords: FrozenSet[str] = frozenset()
    uppercase_keywords: FrozenSet[str] = frozenset()
    family_weight: float = 1.0
    family_cap: float = 1.0
    description: str = ''
    source_refs: Tuple[str, ...] = ()
    version: str = '1.0.0'


# ══════════════════════════════════════════════════════════════════════════════
# PRIORITY 1: CONSTRAINT_FRAMES EXPANSION
# BCP 14 / RFC 2119 + EARS + Cardinality Operators
# ══════════════════════════════════════════════════════════════════════════════

PACK_OBLIGATION = LexiconPack(
    name='obligation',
    category='constraint',
    target_layer='prompt_signature',
    mode='task_prompt',
    description='RFC 2119 / BCP 14 obligation operators. Uppercase forms carry '
                'normative meaning per RFC 8174.',
    source_refs=('RFC2119', 'RFC8174', 'BCP14'),
    uppercase_keywords=frozenset([
        'MUST', 'REQUIRED', 'SHALL',
    ]),
    patterns=(
        (r'\bmust\s+(?:always|ensure|verify|contain|produce|return|handle)\b', 1.5),
        (r'\b(?:is|are)\s+required\s+to\b', 1.2),
        (r'\bshall\s+(?:be|not|ensure|provide|include|comply)\b', 1.2),
        (r'\bmandatory\b', 1.0),
        (r'\bobligatory\b', 0.8),
        (r'\bnon-?negotiable\b', 1.0),
    ),
    keywords=frozenset([
        'must', 'required', 'shall', 'mandatory',
    ]),
    family_weight=1.2,
    family_cap=0.35,
)

PACK_PROHIBITION = LexiconPack(
    name='prohibition',
    category='constraint',
    target_layer='prompt_signature',
    mode='task_prompt',
    description='RFC 2119 prohibition operators plus common task-prompt negation.',
    source_refs=('RFC2119', 'RFC8174'),
    uppercase_keywords=frozenset([
        'MUST NOT', 'SHALL NOT',
    ]),
    patterns=(
        (r'\bmust\s+not\b', 1.5),
        (r'\bshall\s+not\b', 1.2),
        (r'\bdo\s+not\b(?:\s+(?:include|use|add|modify|change|assume|omit|skip|invent))', 1.2),
        (r'\bnever\s+(?:use|include|add|assume|omit|modify|generate|invent|fabricate)\b', 1.3),
        (r'\b(?:is|are)\s+(?:not\s+)?(?:prohibited|forbidden|disallowed)\b', 1.0),
        (r'\bavoid\s+(?:using|including|adding|creating)\b', 0.8),
        (r'\bunder\s+no\s+circumstances?\b', 1.5),
        (r'\byou\s+may\s+not\b', 1.2),
        (r'\bmay\s+not\s+(?:be|introduce|omit|use|exceed|include|modify|alter)\b', 1.2),
    ),
    keywords=frozenset([
        'prohibited', 'forbidden', 'disallowed',
    ]),
    family_weight=1.2,
    family_cap=0.30,
)

PACK_RECOMMENDATION = LexiconPack(
    name='recommendation',
    category='constraint',
    target_layer='prompt_signature',
    mode='task_prompt',
    description='RFC 2119 recommendation-tier operators (SHOULD/RECOMMENDED/MAY/OPTIONAL).',
    source_refs=('RFC2119', 'RFC8174'),
    uppercase_keywords=frozenset([
        'SHOULD', 'SHOULD NOT', 'RECOMMENDED', 'NOT RECOMMENDED',
        'MAY', 'OPTIONAL',
    ]),
    patterns=(
        (r'\bshould\s+(?:include|address|ensure|consider|provide|use|be)\b', 0.8),
        (r'\bshould\s+not\b', 0.9),
        (r'\b(?:it\s+is\s+)?recommended\s+(?:that|to)\b', 0.8),
        (r'\bnot\s+recommended\b', 0.8),
        (r'\boptional(?:ly)?\b', 0.6),
        (r'\b(?:may|can)\s+(?:optionally|also|additionally)\b', 0.5),
        (r'\bprefer(?:red|ably)\b', 0.6),
        (r'\bideally\b', 0.5),
    ),
    keywords=frozenset([
        'recommended', 'optional', 'preferably', 'ideally',
    ]),
    family_weight=0.8,
    family_cap=0.20,
)

PACK_CONDITIONAL = LexiconPack(
    name='conditional',
    category='constraint',
    target_layer='prompt_signature',
    mode='task_prompt',
    description='EARS-derived event/state/conditional scaffolds for task specifications.',
    source_refs=('EARS', 'Mavin2009'),
    patterns=(
        # EARS: When [event], [system] shall...
        (r'\bwhen\s+(?:the|a|an)\s+\w+\s+(?:is|are|has|have|does|occurs?|triggers?|receives?|detects?)\b', 1.5),
        (r'\bwhen\s+\w+ing\b', 0.8),
        # EARS: While [state], [system] shall...
        (r'\bwhile\s+(?:the|a|an)\s+\w+\s+(?:is|are|remains?)\b', 1.3),
        (r'\bwhile\s+\w+ing\b', 0.8),
        # EARS: If [condition] then...
        (r'\bif\s+[^.]{5,60},?\s+then\b', 1.5),
        (r'\bif\s+(?:the|a|an|any|no)\s+\w+\s+(?:is|are|has|does|contains?|exceeds?|falls?|matches?)\b', 1.2),
        (r'\bif\s+present\b', 1.2),
        (r'\bif\s+absent\b', 1.2),
        (r'\bif\s+(?:not\s+)?(?:provided|specified|available|applicable|empty|null|blank|missing)\b', 1.3),
        # EARS: Where [feature] is included...
        (r'\bwhere\s+(?:the|a)\s+\w+\s+(?:is|are)\s+(?:included|enabled|supported|present|active)\b', 1.0),
        # General conditionals
        (r'\botherwise\b', 0.8),
        (r'\bunless\s+(?:the|a|an|otherwise|explicitly|specifically)\b', 1.0),
        (r'\bin\s+(?:the\s+)?(?:case|event)\s+(?:of|that|where)\b', 0.8),
        (r'\bprovided\s+that\b', 0.8),
        (r'\bgiven\s+that\b', 0.7),
    ),
    keywords=frozenset([
        'otherwise', 'unless', 'whereas', 'whenever',
    ]),
    family_weight=1.0,
    family_cap=0.30,
)

PACK_CARDINALITY = LexiconPack(
    name='cardinality',
    category='constraint',
    target_layer='prompt_signature',
    mode='task_prompt',
    description='Cardinality and quantification operators for task constraints.',
    source_refs=('RFC2119', 'EARS'),
    patterns=(
        (r'\bexactly\s+\d+\b', 1.5),
        (r'\bat\s+least\s+\d+\b', 1.2),
        (r'\bat\s+most\s+\d+\b', 1.2),
        (r'\bno\s+more\s+than\s+\d+\b', 1.2),
        (r'\bno\s+fewer\s+than\s+\d+\b', 1.2),
        (r'\bno\s+less\s+than\b', 1.0),
        (r'\bbetween\s+\d+\s+and\s+\d+\b', 1.0),
        (r'\b(?:one|two|three|four|five)\s+(?:of\s+the\s+following|or\s+more|to\s+\w+)\b', 0.8),
        (r'\bone\s+of\b', 0.7),
        (r'\beach\s+(?:of\s+the|row|column|entry|item|record|field|section|task)\b', 1.0),
        (r'\bevery\s+(?:row|column|entry|item|record|field|section|task)\b', 1.0),
        (r'\ball\s+(?:of\s+the|rows|columns|entries|items|records|fields|sections|tasks)\b', 0.8),
        (r'\bonly\s+(?:one|the|if|when|those|items?|records?)\b', 0.8),
        (r'\bup\s+to\s+\d+\b', 0.8),
        (r'\bminimum\s+(?:of\s+)?\d+\b', 1.0),
        (r'\bmaximum\s+(?:of\s+)?\d+\b', 1.0),
        (r'\b(?:a\s+)?single\s+(?:row|column|entry|item|value|file|output)\b', 0.8),
        (r'\bper\s+(?:row|column|entry|item|record|patient|case|task)\b', 0.8),
    ),
    keywords=frozenset([
        'exactly', 'minimum', 'maximum',
    ]),
    family_weight=1.0,
    family_cap=0.25,
)

PACK_STATE = LexiconPack(
    name='state',
    category='constraint',
    target_layer='prompt_signature',
    mode='task_prompt',
    description='State and lifecycle operators for task specifications.',
    source_refs=('EARS',),
    patterns=(
        (r'\b(?:initial|default|starting|baseline)\s+(?:state|value|condition|setting)\b', 1.0),
        (r'\b(?:final|end|terminal|completed?)\s+(?:state|value|condition|output)\b', 1.0),
        (r'\b(?:before|after|during|upon)\s+(?:processing|completion|submission|loading|saving|execution)\b', 0.8),
        (r'\b(?:on|upon)\s+(?:success|failure|error|timeout|completion|receipt)\b', 1.0),
        (r'\b(?:transition|change|switch|move)\s+(?:to|from|between)\s+(?:state|mode|phase)\b', 1.0),
        (r'\b(?:enabled?|disabled?|active|inactive|locked|unlocked|open|closed)\s+(?:state|mode|by default)\b', 0.8),
    ),
    keywords=frozenset([
        'initialized', 'finalized', 'pending', 'completed',
    ]),
    family_weight=0.8,
    family_cap=0.20,
)


# ══════════════════════════════════════════════════════════════════════════════
# PRIORITY 2: SCHEMA / STRUCTURED-OUTPUT LEXICON
# JSON Schema + OpenAPI + Tabular Specs
# ══════════════════════════════════════════════════════════════════════════════

PACK_SCHEMA_JSON = LexiconPack(
    name='schema_json',
    category='schema',
    target_layer='voice_dissonance',
    mode='task_prompt',
    description='JSON Schema and data serialization vocabulary.',
    source_refs=('JSONSchema2020', 'OpenAPI3.1'),
    patterns=(
        (r'\bjson\s*schema\b', 2.0),
        (r'\b(?:json|yaml|xml|csv)\s+(?:output|format|file|object|response|payload)\b', 1.2),
        (r'\bschema\s+(?:for|of|with|object|definition|validation)\b', 1.5),
        (r'\b(?:request|response)\s+(?:body|schema|format|payload|object)\b', 1.2),
        (r'\b(?:query|path|header)\s+param(?:eter)?s?\b', 1.0),
        (r'\b(?:200|201|400|401|403|404|500)\s+(?:response|status|error)\b', 1.0),
        (r'\bendpoint\b', 0.8),
        (r'\bapi\s+(?:call|request|response|endpoint|contract|spec)\b', 1.2),
        (r'\bhttp\s+(?:get|post|put|patch|delete|method)\b', 1.0),
        (r'\brest(?:ful)?\s+api\b', 1.0),
    ),
    keywords=frozenset([
        'schema', 'endpoint', 'payload', 'serialization',
        'deserialization', 'marshalling', 'unmarshalling',
    ]),
    family_weight=1.0,
    family_cap=0.30,
)

PACK_SCHEMA_TYPES = LexiconPack(
    name='schema_types',
    category='schema',
    target_layer='voice_dissonance',
    mode='task_prompt',
    description='JSON Schema type system and property-control keywords.',
    source_refs=('JSONSchema2020',),
    patterns=(
        (r'\btype\s*:\s*["\']?(?:string|integer|number|boolean|array|object|null)\b', 1.5),
        (r'\b(?:required|optional)\s+(?:field|property|parameter|column|attribute)s?\b', 1.2),
        (r'\b(?:additional|pattern)\s*properties\b', 1.5),
        (r'\benum\s*:\s*\[', 1.5),
        (r'\b(?:enum|enumerat(?:ed?|ion))\s+(?:of|values?|type|field)\b', 1.2),
        (r'\b(?:min|max)(?:imum|Length|Items|Properties)\b', 1.0),
        (r'\bdefault\s*:\s', 1.0),
        (r'\bnullable\b', 1.0),
        (r'\b\$ref\b', 1.5),
        (r'\boneOf|anyOf|allOf\b', 1.5),
    ),
    keywords=frozenset([
        'nullable', 'properties', 'additionalproperties',
    ]),
    family_weight=1.0,
    family_cap=0.25,
)

PACK_DATA_FIELDS = LexiconPack(
    name='data_fields',
    category='schema',
    target_layer='voice_dissonance',
    mode='task_prompt',
    description='Generic data-structure vocabulary: field, key, value, record, etc.',
    source_refs=('JSONSchema2020', 'RFC4180', 'Frictionless'),
    patterns=(
        (r'\b(?:field|column|attribute)\s+(?:name|type|description|value|definition)s?\b', 1.0),
        (r'\bkey[- ]?value\s+pair\b', 1.2),
        (r'\bprimary\s+key\b', 1.0),
        (r'\bforeign\s+key\b', 1.0),
        (r'\bdata\s+(?:type|model|structure|format|contract|dictionary)\b', 1.0),
        (r'\b(?:input|output)\s+(?:field|column|schema|format|parameter)s?\b', 1.0),
        (r'\breturn\s+(?:type|value|format|schema)\b', 0.8),
    ),
    keywords=frozenset([
        'field', 'key', 'value', 'record', 'attribute',
    ]),
    family_weight=0.8,
    family_cap=0.20,
)

PACK_TABULAR = LexiconPack(
    name='tabular',
    category='schema',
    target_layer='voice_dissonance',
    mode='task_prompt',
    description='Tabular data and spreadsheet vocabulary (RFC 4180, Frictionless).',
    source_refs=('RFC4180', 'Frictionless'),
    patterns=(
        (r'\bcsv\s+(?:file|format|output|input|data|export|import)\b', 1.0),
        (r'\bdelimiter\b', 1.2),
        (r'\bheader\s+row\b', 1.2),
        (r'\b(?:first|top)\s+row\s+(?:is|contains?|should|as)\s+(?:the\s+)?header\b', 1.2),
        (r'\bworksheet\b', 1.0),
        (r'\btabular\s+(?:data|format|output|resource)\b', 1.2),
        (r'\bpivot\s+table\b', 1.0),
        (r'\bvlookup\b', 1.0),
        (r'\bspreadsheet\b', 0.8),
        (r'\btext/csv\b', 1.5),
        (r'\btsv\b', 0.8),
    ),
    keywords=frozenset([
        'delimiter', 'worksheet', 'spreadsheet', 'tabular',
    ]),
    family_weight=0.8,
    family_cap=0.20,
)


# ══════════════════════════════════════════════════════════════════════════════
# PRIORITY 3: EXECUTABLE-SPEC / RUBRIC / GHERKIN
# ══════════════════════════════════════════════════════════════════════════════

PACK_GHERKIN = LexiconPack(
    name='gherkin',
    category='exec_spec',
    target_layer='prompt_signature',
    mode='task_prompt',
    description='Gherkin BDD specification keywords. These structure executable '
                'specifications — very close to the prompt genre being detected.',
    source_refs=('Cucumber', 'Gherkin'),
    patterns=(
        (r'^\s*Feature:\s', 2.0),
        (r'^\s*Scenario(?:\s+Outline)?:\s', 2.0),
        (r'^\s*Given\s+', 1.5),
        (r'^\s*When\s+', 1.2),
        (r'^\s*Then\s+', 1.5),
        (r'^\s*And\s+', 0.5),
        (r'^\s*But\s+', 0.5),
        (r'^\s*Examples:\s', 1.5),
        (r'^\s*Background:\s', 1.5),
        (r'\b(?:given|when|then)\s+(?:the|a|an)\s+\w+\s+(?:is|are|has|does|should|shall)\b', 1.0),
    ),
    keywords=frozenset(),  # Gherkin keywords are positional — use patterns only
    family_weight=1.3,
    family_cap=0.30,
)

PACK_RUBRIC = LexiconPack(
    name='rubric',
    category='exec_spec',
    target_layer='prompt_signature',
    mode='task_prompt',
    description='Rubric, evaluation, and grading vocabulary for assessment tasks.',
    source_refs=('GDPval', 'Prometheus'),
    patterns=(
        (r'\brubric\b', 1.5),
        (r'\bscoring\s+(?:criteria|rubric|guide|matrix|scale)\b', 1.5),
        (r'\bpass(?:ing)?[/\\]fail(?:ing)?\b', 1.5),
        (r'\bgrader?\b', 1.0),
        (r'\bchecklist\b', 1.0),
        (r'\bverification\s+(?:step|criteria|check|point)\b', 1.2),
        (r'\bvalidation\s+(?:step|criteria|check|rule)\b', 1.0),
        (r'\btest\s+case\b', 1.2),
        (r'\bedge\s+case\b', 1.2),
        (r'\bcorner\s+case\b', 1.0),
        (r'\bexpected\s+(?:output|result|response|behavior|value|answer)\b', 1.5),
        (r'\bgolden\s+(?:answer|response|output|standard|reference)\b', 1.5),
        (r'\bsource\s+of\s+truth\b', 1.5),
        (r'\bground\s+truth\b', 1.5),
        (r'\bgrounded\s+in\b', 1.0),
        (r'\bcite\s+(?:source|evidence|reference)\b', 1.0),
        (r'\bevidence[- ]based\b', 0.8),
        (r'\bscenario\b', 0.6),
        (r'\bexamples?\s*:\s', 0.8),
        (r'\bsample\s+(?:input|output|response|answer)\b', 1.2),
    ),
    keywords=frozenset([
        'rubric', 'checklist', 'grader', 'verification', 'validation',
    ]),
    family_weight=1.2,
    family_cap=0.30,
)

PACK_ACCEPTANCE = LexiconPack(
    name='acceptance',
    category='exec_spec',
    target_layer='prompt_signature',
    mode='task_prompt',
    description='Acceptance criteria and definition-of-done vocabulary.',
    source_refs=('Cucumber', 'AgileAlliance'),
    patterns=(
        (r'\bacceptance\s+criteria\b', 2.0),
        (r'\bdefinition\s+of\s+done\b', 1.5),
        (r'\bdone\s+(?:when|criteria|definition)\b', 1.0),
        (r'\bexit\s+criteria\b', 1.2),
        (r'\bentry\s+criteria\b', 1.0),
        (r'\bsuccess\s+(?:criteria|metric|condition|measure)\b', 1.2),
        (r'\bcompletion\s+criteria\b', 1.2),
        (r'\bdeliverable\s+(?:must|should|criteria|requirement)\b', 1.0),
        (r'\buser\s+story\b', 0.8),
        (r'\bas\s+a\s+\w+,?\s+i\s+want\b', 1.5),
    ),
    keywords=frozenset([
        'deliverable', 'milestone', 'criterion', 'criteria',
    ]),
    family_weight=1.0,
    family_cap=0.25,
)


# ══════════════════════════════════════════════════════════════════════════════
# PRIORITY 4 (PARTIAL): IDI TYPED INSTRUCTION OPERATORS
# Task verbs, value-domain, formatting verbs
# ══════════════════════════════════════════════════════════════════════════════

PACK_TASK_VERBS = LexiconPack(
    name='task_verbs',
    category='instruction',
    target_layer='instruction_density',
    mode='task_prompt',
    description='Bloom-taxonomy action verbs for task specifications. Noisy alone; '
                'gains signal when paired with constraint/schema operators.',
    source_refs=('BloomTaxonomy', 'Utica'),
    patterns=(
        # Higher-order verbs (more signal)
        (r'\b(?:classify|categorize)\s+(?:the|each|all|every)\b', 1.2),
        (r'\b(?:evaluate|assess|critique|analyze)\s+(?:the|each|whether|how)\b', 1.2),
        (r'\b(?:design|architect|propose)\s+(?:a|an|the)\b', 1.0),
        (r'\b(?:justify|defend|argue|explain\s+why)\b', 1.0),
        # Mid-level verbs
        (r'\b(?:compare|contrast|differentiate)\s+(?:the|between|across)\b', 1.0),
        (r'\b(?:summarize|synthesize|consolidate)\s+(?:the|all|key)\b', 0.8),
        (r'\b(?:rewrite|revise|edit|refactor)\s+(?:the|this|each|to)\b', 0.8),
        (r'\b(?:translate|convert|transform)\s+(?:the|each|all|from|into|to)\b', 1.0),
        # Extraction / manipulation verbs
        (r'\b(?:extract|identify|locate|find|detect)\s+(?:the|all|each|any|every)\b', 1.0),
        (r'\b(?:label|tag|annotate|mark)\s+(?:the|each|all|every)\b', 1.0),
        (r'\b(?:rank|sort|order|prioritize)\s+(?:the|by|based|according)\b', 1.0),
        (r'\b(?:populate|fill|complete)\s+(?:the|each|all|every|a)\b', 0.8),
        (r'\b(?:validate|verify|check|confirm)\s+(?:that|the|each|whether|all)\b', 1.0),
        (r'\b(?:normalize|standardize|clean)\s+(?:the|all|each)\b', 1.0),
        (r'\b(?:parse|tokenize|split)\s+(?:the|each|into)\b', 1.0),
        (r'\b(?:map|associate|link|cross-reference)\s+(?:the|each|to|between)\b', 0.8),
        (r'\b(?:generate|produce|create|output)\s+(?:a|an|the|your)\b', 0.8),
        (r'\b(?:format|structure|organize)\s+(?:the|your|as|into|according)\b', 0.8),
    ),
    keywords=frozenset([
        'classify', 'identify', 'extract', 'label', 'compare', 'evaluate',
        'rewrite', 'translate', 'summarize', 'justify', 'rank', 'design',
        'generate', 'format', 'populate', 'validate', 'convert', 'normalize',
        'parse', 'map', 'annotate', 'synthesize', 'categorize', 'prioritize',
    ]),
    family_weight=0.7,  # Low weight alone — gains via pairing
    family_cap=0.20,
)

PACK_VALUE_DOMAIN = LexiconPack(
    name='value_domain',
    category='instruction',
    target_layer='instruction_density',
    mode='task_prompt',
    description='Value-domain and control operators (null handling, defaults, allowed values).',
    source_refs=('JSONSchema2020', 'RFC2119'),
    patterns=(
        (r'\b(?:true|false)\b', 0.5),
        (r'\b(?:yes|no)\b(?!\s+(?:longer|one|matter|idea))', 0.5),
        (r'\bnull\b', 1.0),
        (r'\bnone\b(?=\s|$|[,;.])', 0.5),
        (r'\bunknown\b', 0.5),
        (r'\bleave\s+blank\b', 1.5),
        (r'\bdefault\s+(?:to|value|is|of)\b', 1.2),
        (r'\bfallback\s+(?:to|value|is)\b', 1.2),
        (r'\ballowed\s+values?\b', 1.5),
        (r'\bvalid\s+values?\b', 1.5),
        (r'\bacceptable\s+values?\b', 1.2),
        (r'\bpermitted\s+values?\b', 1.2),
        (r'\bone\s+of\s*[:\[({]', 1.5),
        (r'\breturn\s+as\b', 1.2),
        (r'\boutput\s+as\b', 1.0),
        (r'\bformat\s+as\b', 1.0),
        (r'\bMISSING\b', 2.0),
        (r'\bN/?A\b', 0.5),
        (r'\bempty\s+(?:string|value|cell|field)\b', 1.0),
        (r'\bplaceholder\b', 0.8),
        (r'\bsentinel\s+value\b', 1.5),
    ),
    keywords=frozenset([
        'null', 'fallback', 'default', 'placeholder', 'sentinel',
    ]),
    family_weight=1.0,
    family_cap=0.25,
)


# ══════════════════════════════════════════════════════════════════════════════
# PRIORITY 5 (PARTIAL): FORMAT / MARKUP SUBLEXICON
# ══════════════════════════════════════════════════════════════════════════════

PACK_FORMAT_MARKUP = LexiconPack(
    name='format_markup',
    category='format',
    target_layer='voice_dissonance',
    mode='task_prompt',
    description='Markdown, code fences, tables, task lists — output-shape markers '
                'that occur constantly in modern task prompts.',
    source_refs=('CommonMark', 'GFM'),
    patterns=(
        (r'```\w*\n', 1.5),   # Fenced code block
        (r'~~~\w*\n', 1.2),   # Tilde code fence
        (r'\bcode\s+(?:block|fence|snippet|example)\b', 1.0),
        (r'\bmarkdown\s+(?:format|table|output|syntax)\b', 1.2),
        (r'\b(?:bullet|numbered|ordered|unordered)\s+list\b', 0.8),
        (r'\btask\s+list\b', 1.0),
        (r'\bchecklist\s+(?:format|style|item)\b', 1.0),
        (r'\[[ x]\]', 1.5),   # GitHub task list syntax
        (r'\bheading\s+(?:level|format|style)\b', 0.8),
        (r'\btable\s+(?:format|with|header|row|column)\b', 1.0),
        (r'\bheader\s+row\b', 1.0),
        (r'\bpipe[- ]?delimited\b', 1.2),
        (r'\binline\s+code\b', 0.8),
        (r'\bformatted?\s+(?:as|in|using)\s+(?:markdown|json|yaml|csv|xml|html)\b', 1.2),
        (r'\boutput\s+format\s*:', 1.5),
        (r'\bresponse\s+format\s*:', 1.5),
    ),
    keywords=frozenset([
        'markdown', 'backtick', 'codeblock',
    ]),
    family_weight=0.9,
    family_cap=0.20,
)


# ══════════════════════════════════════════════════════════════════════════════
# PACK REGISTRY
# ══════════════════════════════════════════════════════════════════════════════

PACK_REGISTRY: Dict[str, LexiconPack] = {
    # Priority 1: Constraint families
    'obligation': PACK_OBLIGATION,
    'prohibition': PACK_PROHIBITION,
    'recommendation': PACK_RECOMMENDATION,
    'conditional': PACK_CONDITIONAL,
    'cardinality': PACK_CARDINALITY,
    'state': PACK_STATE,
    # Priority 2: Schema / structured output
    'schema_json': PACK_SCHEMA_JSON,
    'schema_types': PACK_SCHEMA_TYPES,
    'data_fields': PACK_DATA_FIELDS,
    'tabular': PACK_TABULAR,
    # Priority 3: Executable spec / rubric
    'gherkin': PACK_GHERKIN,
    'rubric': PACK_RUBRIC,
    'acceptance': PACK_ACCEPTANCE,
    # Priority 4: IDI typed operators
    'task_verbs': PACK_TASK_VERBS,
    'value_domain': PACK_VALUE_DOMAIN,
    # Priority 5: Format / markup
    'format_markup': PACK_FORMAT_MARKUP,
}


# ══════════════════════════════════════════════════════════════════════════════
# SCORING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def _compile_pack(pack: LexiconPack):
    """Compile regex patterns for a pack. Returns list of (compiled_re, weight)."""
    compiled = []
    flags = re.IGNORECASE | re.MULTILINE
    for pat_str, weight in pack.patterns:
        try:
            compiled.append((re.compile(pat_str, flags), weight))
        except re.error as e:
            print(f"WARNING: Bad regex in pack '{pack.name}': {pat_str} — {e}")
    return compiled


# Pre-compile all packs at import time
_COMPILED_PACKS: Dict[str, list] = {}
_KEYWORD_RES: Dict[str, re.Pattern] = {}
_UPPERCASE_RES: Dict[str, re.Pattern] = {}

for _name, _pack in PACK_REGISTRY.items():
    _COMPILED_PACKS[_name] = _compile_pack(_pack)

    if _pack.keywords:
        _kw_pattern = r'\b(?:' + '|'.join(re.escape(k) for k in sorted(_pack.keywords)) + r')\b'
        _KEYWORD_RES[_name] = re.compile(_kw_pattern, re.IGNORECASE)

    if _pack.uppercase_keywords:
        # Case-sensitive matching for normative uppercase forms
        _uc_parts = []
        for kw in sorted(_pack.uppercase_keywords, key=len, reverse=True):
            _uc_parts.append(r'\b' + re.escape(kw) + r'\b')
        _UPPERCASE_RES[_name] = re.compile('|'.join(_uc_parts))


@dataclass
class PackScore:
    """Scoring result for a single pack against a text."""
    pack_name: str
    category: str
    raw_hits: int = 0
    weighted_hits: float = 0.0
    keyword_hits: int = 0
    uppercase_hits: int = 0
    raw_score: float = 0.0
    capped_score: float = 0.0
    matches: List[str] = field(default_factory=list)
    spans: List[dict] = field(default_factory=list)


def score_pack(text: str, pack_name: str, n_sentences: int = 1) -> PackScore:
    """Score a single pack against text.

    Args:
        text: Normalized text to score.
        pack_name: Key in PACK_REGISTRY.
        n_sentences: Sentence count for density normalization.

    Returns:
        PackScore with hits, weighted score (density-normalized), and capped score.
    """
    pack = PACK_REGISTRY[pack_name]
    compiled = _COMPILED_PACKS[pack_name]
    n_sents = max(n_sentences, 1)

    result = PackScore(pack_name=pack_name, category=pack.category)

    # Pattern matching (finditer for span capture)
    for compiled_re, weight in compiled:
        for m in compiled_re.finditer(text):
            result.raw_hits += 1
            result.weighted_hits += weight
            if len(result.matches) < 3:
                result.matches.append(m.group())
            result.spans.append({
                'start': m.start(),
                'end': m.end(),
                'text': m.group()[:80],
                'pack': pack_name,
                'weight': weight,
            })

    # Keyword matching (finditer for span capture)
    kw_re = _KEYWORD_RES.get(pack_name)
    if kw_re:
        for m in kw_re.finditer(text):
            result.keyword_hits += 1
            result.spans.append({
                'start': m.start(),
                'end': m.end(),
                'text': m.group(),
                'pack': pack_name,
                'weight': 0.0,
                'type': 'keyword',
            })

    # Uppercase keyword matching (case-sensitive, finditer for span capture)
    uc_re = _UPPERCASE_RES.get(pack_name)
    if uc_re:
        for m in uc_re.finditer(text):
            result.uppercase_hits += 1
            result.weighted_hits += 2.0
            result.spans.append({
                'start': m.start(),
                'end': m.end(),
                'text': m.group(),
                'pack': pack_name,
                'weight': 2.0,
                'type': 'uppercase',
            })

    # Density-normalized score: weighted_hits per sentence × family_weight
    density = result.weighted_hits / n_sents
    result.raw_score = density * pack.family_weight
    result.capped_score = min(result.raw_score, pack.family_cap)

    return result


def score_packs(text: str, pack_names: Optional[List[str]] = None,
                n_sentences: int = 1) -> Dict[str, PackScore]:
    """Score multiple packs against text.

    Args:
        text: Normalized text.
        pack_names: List of pack names to score. If None, scores all packs.
        n_sentences: Sentence count for density normalization.

    Returns:
        Dict mapping pack_name → PackScore.
    """
    names = pack_names or list(PACK_REGISTRY.keys())
    return {name: score_pack(text, name, n_sentences) for name in names}


def get_packs_for_layer(target_layer: str) -> List[str]:
    """Get pack names that feed a specific pipeline layer."""
    return [name for name, pack in PACK_REGISTRY.items()
            if pack.target_layer == target_layer or pack.target_layer == 'multi']


def get_packs_for_mode(mode: str) -> List[str]:
    """Get pack names eligible for a detection mode."""
    return [name for name, pack in PACK_REGISTRY.items()
            if pack.mode == mode or pack.mode == 'both']


def get_category_score(pack_scores: Dict[str, PackScore], category: str) -> float:
    """Aggregate capped scores across all packs in a category."""
    return sum(
        ps.capped_score for ps in pack_scores.values()
        if ps.category == category
    )


def get_total_constraint_score(pack_scores: Dict[str, PackScore]) -> float:
    """Sum of all constraint-family capped scores (Priority 1)."""
    return get_category_score(pack_scores, 'constraint')


def get_total_schema_score(pack_scores: Dict[str, PackScore]) -> float:
    """Sum of all schema-family capped scores (Priority 2)."""
    return get_category_score(pack_scores, 'schema')


def get_total_exec_spec_score(pack_scores: Dict[str, PackScore]) -> float:
    """Sum of all exec_spec-family capped scores (Priority 3)."""
    return get_category_score(pack_scores, 'exec_spec')


# ══════════════════════════════════════════════════════════════════════════════
# INTEGRATION HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def compute_pack_enhanced_cfd(text: str, n_sentences: int,
                               legacy_cfd: float = 0.0) -> dict:
    """Compute pack-enhanced Constraint Frame Density for prompt_signature.

    Combines legacy CONSTRAINT_FRAMES patterns with the new typed packs.
    Returns dict with enhanced_cfd, pack breakdown, and legacy comparison.
    """
    constraint_packs = get_packs_for_layer('prompt_signature')
    constraint_packs = [p for p in constraint_packs
                        if PACK_REGISTRY[p].category in ('constraint', 'exec_spec')]

    scores = score_packs(text, constraint_packs, n_sentences)

    total_constraint = get_total_constraint_score(scores)
    total_exec_spec = get_total_exec_spec_score(scores)

    # Enhanced CFD: legacy + new packs (with diminishing returns)
    pack_boost = total_constraint * 0.6 + total_exec_spec * 0.4
    enhanced_cfd = legacy_cfd + pack_boost

    return {
        'enhanced_cfd': round(enhanced_cfd, 4),
        'legacy_cfd': legacy_cfd,
        'pack_constraint_score': round(total_constraint, 4),
        'pack_exec_spec_score': round(total_exec_spec, 4),
        'pack_boost': round(pack_boost, 4),
        'active_packs': {name: {
            'hits': s.raw_hits,
            'weighted': round(s.weighted_hits, 2),
            'capped': round(s.capped_score, 4),
            'uc_hits': s.uppercase_hits,
        } for name, s in scores.items() if s.raw_hits > 0},
        'distinct_pack_families': sum(1 for s in scores.values() if s.raw_hits > 0),
    }


def compute_pack_enhanced_spec(text: str, n_sentences: int,
                                legacy_spec_score: float = 0.0) -> dict:
    """Compute pack-enhanced spec_score for voice_dissonance.

    Adds schema/structured-output vocabulary to the legacy spreadsheet-focused spec_score.
    """
    schema_packs = get_packs_for_layer('voice_dissonance')
    schema_packs = [p for p in schema_packs
                    if PACK_REGISTRY[p].category in ('schema', 'format')]

    scores = score_packs(text, schema_packs, n_sentences)

    total_schema = get_total_schema_score(scores)
    format_score = get_category_score(scores, 'format')

    pack_boost = total_schema * 0.7 + format_score * 0.3
    enhanced_spec = legacy_spec_score + pack_boost

    return {
        'enhanced_spec': round(enhanced_spec, 4),
        'legacy_spec': legacy_spec_score,
        'pack_schema_score': round(total_schema, 4),
        'pack_format_score': round(format_score, 4),
        'pack_boost': round(pack_boost, 4),
        'active_packs': {name: {
            'hits': s.raw_hits,
            'weighted': round(s.weighted_hits, 2),
            'capped': round(s.capped_score, 4),
        } for name, s in scores.items() if s.raw_hits > 0},
    }


def compute_pack_enhanced_idi(text: str, n_words: int,
                               legacy_idi: float = 0.0) -> dict:
    """Compute pack-enhanced IDI for instruction_density.

    Adds typed task-verb and value-domain operators to legacy imperative counting.
    Key insight from roadmap: action verbs alone are noisy, but action verbs
    PLUS constraint/schema operators are strong signal.
    """
    idi_packs = get_packs_for_layer('instruction_density')
    per100 = max(n_words / 100, 1)

    scores = score_packs(text, idi_packs, n_sentences=1)

    # Raw per-100-word density for IDI compatibility
    task_verb_density = scores.get('task_verbs', PackScore('task_verbs', 'instruction')).weighted_hits / per100
    value_domain_density = scores.get('value_domain', PackScore('value_domain', 'instruction')).weighted_hits / per100

    # Pairing bonus: verbs get extra weight when constraint/schema packs also fired
    # (Computed by caller who has access to L2.5 and L2.6 results)
    pack_idi_contribution = (task_verb_density * 1.0 + value_domain_density * 2.0)
    enhanced_idi = legacy_idi + pack_idi_contribution

    return {
        'enhanced_idi': round(enhanced_idi, 2),
        'legacy_idi': legacy_idi,
        'task_verb_density': round(task_verb_density, 3),
        'value_domain_density': round(value_domain_density, 3),
        'pack_contribution': round(pack_idi_contribution, 2),
        'active_packs': {name: {
            'hits': s.raw_hits,
            'weighted': round(s.weighted_hits, 2),
            'capped': round(s.capped_score, 4),
        } for name, s in scores.items() if s.raw_hits > 0},
    }


# ══════════════════════════════════════════════════════════════════════════════
# DIAGNOSTICS
# ══════════════════════════════════════════════════════════════════════════════

def pack_summary() -> str:
    """Print a summary table of all registered packs."""
    lines = [
        f"{'Pack':<20} {'Category':<12} {'Layer':<6} {'Mode':<14} "
        f"{'Patterns':<9} {'Keywords':<9} {'UC Keys':<8} {'Weight':<7} {'Cap':<5}",
        '-' * 100,
    ]
    for name, pack in PACK_REGISTRY.items():
        lines.append(
            f"{name:<20} {pack.category:<12} {pack.target_layer:<6} {pack.mode:<14} "
            f"{len(pack.patterns):<9} {len(pack.keywords):<9} {len(pack.uppercase_keywords):<8} "
            f"{pack.family_weight:<7.1f} {pack.family_cap:<5.2f}"
        )
    return '\n'.join(lines)


def diagnose_text(text: str, n_sentences: int = 1) -> str:
    """Score all packs and return a diagnostic report."""
    scores = score_packs(text, n_sentences=n_sentences)

    lines = [f"BEET Lexicon Pack Diagnostic (v{_LEXICON_VERSION})", '=' * 70]

    for cat in ['constraint', 'schema', 'exec_spec', 'instruction', 'format']:
        cat_packs = [(n, s) for n, s in scores.items() if s.category == cat]
        if not cat_packs:
            continue

        cat_total = sum(s.capped_score for _, s in cat_packs)
        lines.append(f"\n{cat.upper()} (total capped: {cat_total:.3f})")
        lines.append('-' * 50)

        for name, s in sorted(cat_packs, key=lambda x: -x[1].capped_score):
            if s.raw_hits > 0:
                lines.append(
                    f"  {name:<20} hits={s.raw_hits:<3} kw={s.keyword_hits:<3} "
                    f"uc={s.uppercase_hits:<2} raw={s.raw_score:.3f} "
                    f"cap={s.capped_score:.3f}  {s.matches[:2]}"
                )

    return '\n'.join(lines)


# ==============================================================================
# LEXICON INTEGRATION
# ==============================================================================
"""
Lexicon Packs Integration
=========================
Enhanced layer wrappers that augment the base analyzers with the externalized
versioned lexicon packs.

Usage:
        run_prompt_signature_enhanced,
        run_voice_dissonance_enhanced,
        run_instruction_density_enhanced,
    )
    result = run_prompt_signature_enhanced(text)
"""

def run_prompt_signature_enhanced(text: str, base_result: Optional[dict] = None) -> dict:
    """Enhanced prompt_signature with lexicon pack integration.

    Runs the legacy prompt_signature first (or accepts pre-computed result),
    then augments with Priority 1 (constraint families) and Priority 3
    (exec-spec/rubric/Gherkin) packs.
    """
    if base_result is None:
        base_result = run_prompt_signature(text)

    sents = base_result.get('_sentences', None)
    if sents is None:
        sents = get_sentences(text)
    n_sents = max(len(sents), 1)

    constraint_names = [n for n in get_packs_for_layer('prompt_signature')
                        if PACK_REGISTRY[n].category == 'constraint']
    constraint_scores = score_packs(text, constraint_names, n_sents)
    total_constraint = get_total_constraint_score(constraint_scores)

    exec_spec_names = [n for n in get_packs_for_layer('prompt_signature')
                       if PACK_REGISTRY[n].category == 'exec_spec']
    exec_spec_scores = score_packs(text, exec_spec_names, n_sents)
    total_exec_spec = get_total_exec_spec_score(exec_spec_scores)

    all_pack_scores = {**constraint_scores, **exec_spec_scores}
    active_families = sum(1 for s in all_pack_scores.values() if s.raw_hits > 0)

    legacy_composite = base_result.get('composite', 0.0)

    pack_boost = 0.0
    if total_constraint >= 0.40:
        pack_boost += 0.20
    elif total_constraint >= 0.20:
        pack_boost += 0.12
    elif total_constraint >= 0.08:
        pack_boost += 0.05

    if total_exec_spec >= 0.30:
        pack_boost += 0.15
    elif total_exec_spec >= 0.15:
        pack_boost += 0.08
    elif total_exec_spec >= 0.05:
        pack_boost += 0.03

    if active_families >= 6:
        pack_boost += 0.15
    elif active_families >= 4:
        pack_boost += 0.08
    elif active_families >= 2:
        pack_boost += 0.03

    uc_total = sum(s.uppercase_hits for s in all_pack_scores.values())
    if uc_total >= 3:
        pack_boost += 0.10
    elif uc_total >= 1:
        pack_boost += 0.04

    enhanced_composite = min(legacy_composite + pack_boost, 1.0)

    result = dict(base_result)
    result.update({
        'composite': enhanced_composite,
        'legacy_composite': legacy_composite,
        'pack_boost': round(pack_boost, 4),
        'pack_constraint_score': round(total_constraint, 4),
        'pack_exec_spec_score': round(total_exec_spec, 4),
        'pack_active_families': active_families,
        'pack_uc_hits': uc_total,
        'pack_details': {
            name: {
                'hits': s.raw_hits,
                'capped': round(s.capped_score, 4),
                'uc': s.uppercase_hits,
            }
            for name, s in all_pack_scores.items()
            if s.raw_hits > 0
        },
        'pack_spans': sorted(
            [sp for s in all_pack_scores.values() for sp in s.spans],
            key=lambda x: x['start'],
        ),
    })

    return result


def run_voice_dissonance_enhanced(text: str, base_result: Optional[dict] = None) -> dict:
    """Enhanced voice_dissonance with schema/structured-output vocabulary."""
    if base_result is None:
        base_result = run_voice_dissonance(text)

    words = text.split()
    n_words = len(words)
    per100 = max(n_words / 100, 1)

    schema_names = [n for n in get_packs_for_layer('voice_dissonance')
                    if PACK_REGISTRY[n].category == 'schema']
    schema_scores = score_packs(text, schema_names, n_sentences=1)
    total_schema = get_total_schema_score(schema_scores)

    format_names = [n for n in get_packs_for_layer('voice_dissonance')
                    if PACK_REGISTRY[n].category == 'format']
    format_scores = score_packs(text, format_names, n_sentences=1)
    total_format = get_category_score(format_scores, 'format')

    all_pack_scores = {**schema_scores, **format_scores}

    legacy_spec = base_result.get('spec_score', 0.0)
    schema_per100 = sum(s.weighted_hits for s in schema_scores.values()) / per100
    format_per100 = sum(s.weighted_hits for s in format_scores.values()) / per100
    pack_spec_boost = schema_per100 * 2.0 + format_per100 * 1.0
    enhanced_spec = legacy_spec + pack_spec_boost

    voice_score = base_result.get('voice_score', 0.0)
    enhanced_vsd = voice_score * enhanced_spec

    ssi_spec_threshold = 5.0 if base_result.get('contractions', 0) == 0 else 7.0
    enhanced_ssi = (
        enhanced_spec >= ssi_spec_threshold
        and voice_score < 0.5
        and base_result.get('hedges', 0) == 0
        and n_words >= 150
    )

    result = dict(base_result)
    result.update({
        'spec_score': round(enhanced_spec, 2),
        'legacy_spec_score': legacy_spec,
        'vsd': round(enhanced_vsd, 1),
        'legacy_vsd': base_result.get('vsd', 0.0),
        'ssi_enhanced': enhanced_ssi,
        'pack_schema_score': round(total_schema, 4),
        'pack_format_score': round(total_format, 4),
        'pack_schema_per100': round(schema_per100, 3),
        'pack_format_per100': round(format_per100, 3),
        'pack_spec_boost': round(pack_spec_boost, 3),
        'pack_details': {
            name: {
                'hits': s.raw_hits,
                'capped': round(s.capped_score, 4),
            }
            for name, s in all_pack_scores.items()
            if s.raw_hits > 0
        },
        'pack_spans': sorted(
            [sp for s in all_pack_scores.values() for sp in s.spans],
            key=lambda x: x['start'],
        ),
    })

    return result


def run_instruction_density_enhanced(text: str, base_result: Optional[dict] = None,
                                     constraint_active: bool = False,
                                     schema_active: bool = False) -> dict:
    """Enhanced instruction_density with typed task-verb and value-domain operators."""
    if base_result is None:
        base_result = run_instruction_density(text)

    words = text.split()
    n_words = len(words)
    per100 = max(n_words / 100, 1)

    idi_names = get_packs_for_layer('instruction_density')
    idi_scores = score_packs(text, idi_names, n_sentences=1)

    task_verb_score = idi_scores.get('task_verbs', PackScore('task_verbs', 'instruction'))
    value_domain_score = idi_scores.get('value_domain', PackScore('value_domain', 'instruction'))

    tv_per100 = task_verb_score.weighted_hits / per100
    vd_per100 = value_domain_score.weighted_hits / per100

    if constraint_active or schema_active:
        tv_weight = 1.0
        pairing_label = 'paired'
    else:
        tv_weight = 0.5
        pairing_label = 'unpaired'

    pack_idi_boost = (tv_per100 * tv_weight * 1.0) + (vd_per100 * 2.0)

    legacy_idi = base_result.get('idi', 0.0)
    enhanced_idi = legacy_idi + pack_idi_boost

    result = dict(base_result)
    result.update({
        'idi': round(enhanced_idi, 1),
        'legacy_idi': legacy_idi,
        'pack_idi_boost': round(pack_idi_boost, 2),
        'pack_tv_per100': round(tv_per100, 3),
        'pack_vd_per100': round(vd_per100, 3),
        'pack_tv_pairing': pairing_label,
        'pack_tv_weight': tv_weight,
        'pack_details': {
            name: {
                'hits': s.raw_hits,
                'capped': round(s.capped_score, 4),
            }
            for name, s in idi_scores.items()
            if s.raw_hits > 0
        },
        'pack_spans': sorted(
            [sp for s in idi_scores.values() for sp in s.spans],
            key=lambda x: x['start'],
        ),
    })

    return result


# ==============================================================================
# CHANNEL: BASE CLASS
# ==============================================================================
"""Evidence fusion channels."""


class ChannelResult:
    """Result from a single detection channel."""
    __slots__ = ('channel', 'score', 'severity', 'explanation',
                 'mode_eligibility', 'sub_signals')

    SEVERITIES = ('GREEN', 'YELLOW', 'AMBER', 'RED')
    SEV_ORDER = {'GREEN': 0, 'YELLOW': 1, 'AMBER': 2, 'RED': 3}

    def __init__(self, channel, score=0.0, severity='GREEN', explanation='',
                 mode_eligibility=None, sub_signals=None):
        self.channel = channel
        self.score = score
        self.severity = severity
        self.explanation = explanation
        self.mode_eligibility = mode_eligibility or ['task_prompt', 'generic_aigt']
        self.sub_signals = sub_signals or {}

    @property
    def sev_level(self):
        return self.SEV_ORDER.get(self.severity, 0)

    def __repr__(self):
        return f"CH:{self.channel}={self.severity}({self.score:.2f})"


# ==============================================================================
# CHANNEL: PROMPT STRUCTURE
# ==============================================================================
"""Channel 1: Prompt-structure signals (task_prompt primary).

Combines preamble, prompt signature, voice dissonance, instruction density, SSI.
"""



def score_prompt_structure(preamble_score, preamble_severity, prompt_sig, voice_dis, instr_density, word_count):
    """Score prompt-structure channel. Returns ChannelResult."""
    sub = {}
    score = 0.0
    severity = 'GREEN'
    parts = []

    # Preamble
    if preamble_severity == 'CRITICAL':
        return ChannelResult(
            'prompt_structure', 0.99, 'RED',
            'Preamble detection (critical hit)',
            mode_eligibility=['task_prompt', 'generic_aigt'],
            sub_signals={'preamble': 0.99},
        )
    if preamble_score >= 0.50:
        sub['preamble'] = preamble_score
        score = max(score, preamble_score)
        parts.append(f"preamble={preamble_score:.2f}")

    # Prompt signature
    comp = prompt_sig['composite']
    sub['prompt_signature'] = comp
    if comp >= 0.60:
        score = max(score, comp)
        severity = 'RED'
        parts.append(f"prompt_sig={comp:.2f}(RED)")
    elif comp >= 0.40:
        score = max(score, comp)
        severity = max(severity, 'AMBER', key=lambda s: ChannelResult.SEV_ORDER.get(s, 0))
        parts.append(f"prompt_sig={comp:.2f}(AMBER)")
    elif comp >= 0.20:
        score = max(score, comp * 0.7)
        severity = max(severity, 'YELLOW', key=lambda s: ChannelResult.SEV_ORDER.get(s, 0))
        parts.append(f"prompt_sig={comp:.2f}(YELLOW)")

    # Voice dissonance (voice-gated)
    if voice_dis['voice_gated']:
        vsd = voice_dis['vsd']
        sub['vsd_gated'] = vsd
        if vsd >= 50:
            score = max(score, 0.90)
            severity = 'RED'
            parts.append(f"VSD={vsd:.0f}(RED)")
        elif vsd >= 21:
            score = max(score, 0.70)
            if ChannelResult.SEV_ORDER.get(severity, 0) < 2:
                severity = 'AMBER'
            parts.append(f"VSD={vsd:.0f}(AMBER)")

    # Instruction density
    if instr_density:
        idi = instr_density['idi']
        sub['idi'] = idi
        if idi >= 12:
            score = max(score, 0.85)
            severity = 'RED'
            parts.append(f"IDI={idi:.0f}(RED)")
        elif idi >= 8:
            score = max(score, 0.65)
            if ChannelResult.SEV_ORDER.get(severity, 0) < 2:
                severity = 'AMBER'
            parts.append(f"IDI={idi:.0f}(AMBER)")

    # SSI (sterile specification)
    ssi_spec_threshold = 5.0 if voice_dis['contractions'] == 0 else 7.0
    ssi_triggered = (
        voice_dis['spec_score'] >= ssi_spec_threshold
        and voice_dis['voice_score'] < 0.5
        and voice_dis['hedges'] == 0
        and word_count >= 150
    )
    if ssi_triggered:
        sub['ssi'] = voice_dis['spec_score']
        if voice_dis['spec_score'] >= 8.0:
            score = max(score, 0.70)
            if ChannelResult.SEV_ORDER.get(severity, 0) < 2:
                severity = 'AMBER'
            parts.append(f"SSI={voice_dis['spec_score']:.0f}(AMBER)")
        else:
            score = max(score, 0.45)
            if ChannelResult.SEV_ORDER.get(severity, 0) < 1:
                severity = 'YELLOW'
            parts.append(f"SSI={voice_dis['spec_score']:.0f}(YELLOW)")

    # VSD ungated (very high)
    if not voice_dis['voice_gated'] and voice_dis['vsd'] >= 100:
        sub['vsd_ungated'] = voice_dis['vsd']
        score = max(score, 0.60)
        if ChannelResult.SEV_ORDER.get(severity, 0) < 2:
            severity = 'AMBER'
        parts.append(f"VSD_ungated={voice_dis['vsd']:.0f}")
    elif not voice_dis['voice_gated'] and voice_dis['vsd'] >= 21:
        sub['vsd_ungated'] = voice_dis['vsd']
        score = max(score, 0.30)
        if ChannelResult.SEV_ORDER.get(severity, 0) < 1:
            severity = 'YELLOW'

    explanation = f"Prompt-structure: {', '.join(parts)}" if parts else 'Prompt-structure: no signals'

    return ChannelResult(
        'prompt_structure', score, severity, explanation,
        mode_eligibility=['task_prompt', 'generic_aigt'],
        sub_signals=sub,
    )


# ==============================================================================
# CHANNEL: STYLOMETRIC
# ==============================================================================
"""Channel 2: Stylometric signals (generic_aigt primary).

Combines NSSI, semantic resonance, perplexity, and fingerprints.
"""



def score_stylometric(fingerprint_score, self_sim, voice_dis=None, semantic=None, ppl=None):
    """Score stylometric channel. Returns ChannelResult."""
    sub = {}
    score = 0.0
    severity = 'GREEN'
    parts = []

    # Fingerprints: supporting-only
    if fingerprint_score > 0:
        sub['fingerprints'] = fingerprint_score

    # NSSI: primary stylometric signal
    if self_sim and self_sim.get('determination'):
        nssi_det = self_sim['determination']
        nssi_score = self_sim.get('nssi_score', 0)
        nssi_signals = self_sim.get('nssi_signals', 0)
        sub['nssi_score'] = nssi_score
        sub['nssi_signals'] = nssi_signals

        if nssi_det == 'RED':
            score = max(score, min(0.85, self_sim.get('confidence', 0.80)))
            severity = 'RED'
            parts.append(f"NSSI={nssi_score:.2f}/{nssi_signals}sig(RED)")
        elif nssi_det == 'AMBER':
            score = max(score, min(0.65, self_sim.get('confidence', 0.60)))
            severity = 'AMBER'
            parts.append(f"NSSI={nssi_score:.2f}/{nssi_signals}sig(AMBER)")
        elif nssi_det == 'YELLOW':
            score = max(score, min(0.40, self_sim.get('confidence', 0.30)))
            severity = 'YELLOW'
            parts.append(f"NSSI={nssi_score:.2f}/{nssi_signals}sig(YELLOW)")

    # Semantic resonance: supporting signal
    if semantic and semantic.get('determination'):
        sem_det = semantic['determination']
        sem_delta = semantic.get('semantic_delta', 0)
        sub['semantic_ai_score'] = semantic.get('semantic_ai_mean', 0)
        sub['semantic_delta'] = sem_delta

        if sem_det == 'AMBER':
            if severity in ('RED', 'AMBER'):
                score = min(score + 0.10, 1.0)
                parts.append(f"Sem=AMBER(delta={sem_delta:.2f},boost)")
            else:
                score = max(score, semantic.get('confidence', 0.55))
                severity = max(severity, 'AMBER',
                               key=lambda s: ChannelResult.SEV_ORDER.get(s, 0))
                parts.append(f"Sem=AMBER(delta={sem_delta:.2f})")
        elif sem_det == 'YELLOW':
            if severity != 'GREEN':
                score = min(score + 0.05, 1.0)
                parts.append(f"Sem=YELLOW(delta={sem_delta:.2f},supporting)")
            else:
                score = max(score, semantic.get('confidence', 0.30))
                severity = 'YELLOW'
                parts.append(f"Sem=YELLOW(delta={sem_delta:.2f})")

    # Perplexity: supporting signal
    if ppl and ppl.get('determination'):
        ppl_det = ppl['determination']
        ppl_val = ppl.get('perplexity', 0)
        sub['perplexity'] = ppl_val

        if ppl_det == 'AMBER':
            if severity in ('RED', 'AMBER'):
                score = min(score + 0.10, 1.0)
                parts.append(f"PPL={ppl_val:.0f}(AMBER,boost)")
            else:
                score = max(score, ppl.get('confidence', 0.55))
                severity = max(severity, 'AMBER',
                               key=lambda s: ChannelResult.SEV_ORDER.get(s, 0))
                parts.append(f"PPL={ppl_val:.0f}(AMBER)")
        elif ppl_det == 'YELLOW':
            if severity != 'GREEN':
                score = min(score + 0.05, 1.0)
                parts.append(f"PPL={ppl_val:.0f}(YELLOW,supporting)")
            else:
                score = max(score, ppl.get('confidence', 0.30))
                severity = 'YELLOW'
                parts.append(f"PPL={ppl_val:.0f}(YELLOW)")

    # DivEye surprisal variance & volatility decay: pass through as sub-signals.
    # Severity promotion is handled inside run_perplexity() itself; the channel
    # only records the values for audit trail and downstream baselines.
    if ppl:
        sub['surprisal_variance'] = ppl.get('surprisal_variance', 0.0)
        sub['volatility_decay'] = ppl.get('volatility_decay', 1.0)

    # Fingerprints add supporting weight if any stylometric signal is active
    if fingerprint_score >= 0.30 and severity != 'GREEN':
        score = min(score + 0.10, 1.0)
        parts.append(f"fingerprint={fingerprint_score:.2f}(supporting)")

    explanation = f"Stylometry: {', '.join(parts)}" if parts else 'Stylometry: no signals'

    return ChannelResult(
        'stylometry', score, severity, explanation,
        mode_eligibility=['generic_aigt'],
        sub_signals=sub,
    )


# ==============================================================================
# CHANNEL: CONTINUATION
# ==============================================================================
"""Channel 3: Continuation-based detection (DNA-GPT / DNA-GPT-Local)."""



def score_continuation(cont_result):
    """Score continuation channel. Returns ChannelResult."""
    sub = {}
    score = 0.0
    severity = 'GREEN'
    parts = []

    if cont_result and cont_result.get('determination'):
        dna_det = cont_result['determination']
        bscore = cont_result.get('bscore', 0)
        sub['bscore'] = bscore

        proxy = cont_result.get('proxy_features')
        if proxy:
            sub['ncd'] = proxy.get('ncd', 0)
            sub['internal_overlap'] = proxy.get('internal_overlap', 0)
            sub['composite'] = proxy.get('composite', 0)
            label = 'Local'
        else:
            label = 'API'

        if dna_det == 'RED':
            score = min(0.90, cont_result.get('confidence', 0.80))
            severity = 'RED'
            parts.append(f"BScore={bscore:.3f}({label},RED)")
        elif dna_det == 'AMBER':
            score = min(0.70, cont_result.get('confidence', 0.60))
            severity = 'AMBER'
            parts.append(f"BScore={bscore:.3f}({label},AMBER)")
        elif dna_det == 'YELLOW':
            score = min(0.40, cont_result.get('confidence', 0.30))
            severity = 'YELLOW'
            parts.append(f"BScore={bscore:.3f}({label},YELLOW)")

    explanation = f"Continuation: {', '.join(parts)}" if parts else 'Continuation: no signals'

    return ChannelResult(
        'continuation', score, severity, explanation,
        mode_eligibility=['task_prompt', 'generic_aigt'],
        sub_signals=sub,
    )


# ==============================================================================
# CHANNEL: WINDOWED
# ==============================================================================
"""Channel 4: Sentence-window scoring for mixed content detection."""



def score_windowed(window_result=None):
    """Score windowed channel. Returns ChannelResult."""
    if window_result is None or window_result.get('n_windows', 0) == 0:
        return ChannelResult(
            'windowing', 0.0, 'GREEN',
            'Windowing: insufficient text for windows',
            mode_eligibility=['generic_aigt'],
            sub_signals={},
        )

    sub = {
        'max_window': window_result['max_window_score'],
        'mean_window': window_result['mean_window_score'],
        'variance': window_result['window_variance'],
        'hot_span': window_result['hot_span_length'],
        'n_windows': window_result['n_windows'],
        'mixed_signal': window_result['mixed_signal'],
    }

    score = 0.0
    severity = 'GREEN'
    parts = []

    max_w = window_result['max_window_score']
    hot_span = window_result['hot_span_length']
    variance = window_result['window_variance']
    mixed = window_result['mixed_signal']

    if max_w >= 0.60 and hot_span >= 3:
        score = max(score, 0.75)
        severity = 'RED'
        parts.append(f"hot_span={hot_span}(max={max_w:.2f})")
    elif max_w >= 0.45 and hot_span >= 2:
        score = max(score, 0.55)
        severity = 'AMBER'
        parts.append(f"hot_span={hot_span}(max={max_w:.2f})")
    elif max_w >= 0.30:
        score = max(score, 0.30)
        severity = 'YELLOW'
        parts.append(f"max_window={max_w:.2f}")

    if mixed and severity != 'GREEN':
        parts.append(f"MIXED(var={variance:.3f})")

    explanation = f"Windowing: {', '.join(parts)}" if parts else 'Windowing: no signals'

    return ChannelResult(
        'windowing', score, severity, explanation,
        mode_eligibility=['generic_aigt'],
        sub_signals=sub,
    )


# ==============================================================================
# EVIDENCE FUSION
# ==============================================================================
"""Evidence fusion -- combines channel results into final determination."""


def _detect_mode(prompt_sig, instr_density, self_sim, word_count):
    """Auto-detect whether text is a task prompt or generic AI text."""
    prompt_signal = 0.0
    if prompt_sig['composite'] >= 0.15:
        prompt_signal += prompt_sig['composite']
    if instr_density and instr_density.get('idi', 0) >= 5:
        prompt_signal += 0.3
    if prompt_sig.get('framing_completeness', 0) >= 2:
        prompt_signal += 0.2

    generic_signal = 0.0
    if self_sim and self_sim.get('nssi_signals', 0) >= 3:
        generic_signal += 0.4
    if word_count >= 400:
        generic_signal += 0.2

    if prompt_signal > generic_signal + 0.1:
        return 'task_prompt'
    elif generic_signal > prompt_signal + 0.1:
        return 'generic_aigt'
    else:
        return 'task_prompt'


def determine(preamble_score, preamble_severity, prompt_sig, voice_dis,
              instr_density=None, word_count=0,
              self_sim=None, cont_result=None, lang_gate=None, norm_report=None,
              mode='auto', fingerprint_score=0.0, semantic=None, ppl=None, **kwargs):
    """Evidence fusion with channel-based corroboration.

    Returns (determination, reason, confidence, channel_details).
    """
    # Mode detection
    if mode == 'auto':
        mode = _detect_mode(prompt_sig, instr_density, self_sim, word_count)

    # Score all channels
    ch_prompt = score_prompt_structure(preamble_score, preamble_severity, prompt_sig, voice_dis, instr_density, word_count)
    ch_style = score_stylometric(fingerprint_score, self_sim, voice_dis, semantic=semantic, ppl=ppl)
    ch_cont = score_continuation(cont_result)
    ch_window = score_windowed(window_result=kwargs.get('window_result'))

    channels = [ch_prompt, ch_style, ch_cont, ch_window]

    # Detect channel disagreement: any channel at AMBER+ while another is GREEN
    active_sevs = [ch.severity for ch in channels if ch.severity != 'GREEN']
    green_count = sum(1 for ch in channels if ch.severity == 'GREEN')
    has_strong = any(ch.sev_level >= 2 for ch in channels)
    disagreement = has_strong and green_count >= 2
    agreement_label = 'DISAGREE' if disagreement else 'AGREE'

    channel_details = {
        'mode': mode,
        'channel_agreement': agreement_label,
        'channels': {ch.channel: {
            'score': ch.score, 'severity': ch.severity,
            'explanation': ch.explanation, 'mode_eligible': mode in ch.mode_eligibility,
        } for ch in channels},
    }

    # Fairness severity cap
    severity_cap = None
    if lang_gate and lang_gate.get('support_level') == 'UNSUPPORTED':
        severity_cap = 'YELLOW'
    elif lang_gate and lang_gate.get('support_level') == 'REVIEW':
        severity_cap = 'AMBER'

    def _apply_cap(det, reason, conf):
        if severity_cap is None:
            return det, reason, conf
        sev_order = {'GREEN': 0, 'YELLOW': 1, 'REVIEW': 1, 'AMBER': 2, 'RED': 3}
        if sev_order.get(det, 0) > sev_order.get(severity_cap, 3):
            gate_reason = lang_gate.get('reason', 'language support gate')
            return severity_cap, f"{reason} [capped from {det}: {gate_reason}]", min(conf, 0.40)
        return det, reason, conf

    # L0 CRITICAL: instant RED
    if ch_prompt.sub_signals.get('preamble') == 0.99 and preamble_severity == 'CRITICAL':
        det, reason, conf = _apply_cap('RED', ch_prompt.explanation, 0.99)
        return det, reason, conf, channel_details

    # Mode-aware channel filtering
    if mode == 'task_prompt':
        primary_channels = [ch for ch in channels if 'task_prompt' in ch.mode_eligibility]
        supporting_channels = [ch for ch in channels if 'task_prompt' not in ch.mode_eligibility]
    else:
        primary_channels = channels
        supporting_channels = []

    # Evidence fusion
    all_active = sorted(
        [ch for ch in channels if ch.severity != 'GREEN'],
        key=lambda c: c.sev_level, reverse=True,
    )
    primary_active = sorted(
        [ch for ch in primary_channels if ch.severity != 'GREEN'],
        key=lambda c: c.sev_level, reverse=True,
    )
    support_active = [ch for ch in supporting_channels if ch.severity != 'GREEN']

    n_red = sum(1 for ch in all_active if ch.severity == 'RED')
    n_amber_plus = sum(1 for ch in all_active if ch.sev_level >= 2)
    n_yellow_plus = sum(1 for ch in all_active if ch.sev_level >= 1)
    n_primary_red = sum(1 for ch in primary_active if ch.severity == 'RED')
    n_primary_amber = sum(1 for ch in primary_active if ch.sev_level >= 2)
    n_primary_yellow_plus = sum(1 for ch in primary_active if ch.sev_level >= 1)

    top_explanations = [ch.explanation for ch in all_active[:3]]
    combined_reason = ' + '.join(top_explanations) if top_explanations else 'No significant signals'
    top_score = max((ch.score for ch in all_active), default=0.0)

    # RED: strong primary + supporting, or two AMBER+ channels
    if n_primary_red >= 1 and n_yellow_plus >= 2:
        det, reason, conf = _apply_cap('RED', combined_reason, top_score)
        return det, reason, conf, channel_details

    if n_primary_amber >= 2:
        det, reason, conf = _apply_cap('RED', combined_reason, min(top_score, 0.85))
        return det, reason, conf, channel_details

    if mode == 'task_prompt' and n_primary_red >= 1 and n_yellow_plus == 1:
        demote_tag = '[single-channel, demoted from RED]'
        if disagreement:
            demote_tag = '[single-channel, demoted from RED, channels disagree]'
        det, reason, conf = _apply_cap('AMBER', f"{combined_reason} {demote_tag}", min(top_score, 0.75))
        return det, reason, conf, channel_details

    if mode == 'generic_aigt' and n_red >= 1:
        if n_yellow_plus >= 2:
            det, reason, conf = _apply_cap('RED', combined_reason, top_score)
        else:
            det, reason, conf = _apply_cap('RED', f"{combined_reason} [single-channel]", min(top_score, 0.75))
        return det, reason, conf, channel_details

    # AMBER: one channel at AMBER, or two at YELLOW+
    if n_primary_amber >= 1:
        det, reason, conf = _apply_cap('AMBER', combined_reason, min(top_score, 0.70))
        if ch_window.sub_signals.get('mixed_signal') and ch_window.severity != 'GREEN':
            return 'MIXED', f"{reason} [windowed variance suggests hybrid text]", min(conf, 0.60), channel_details
        return det, reason, conf, channel_details

    if mode == 'task_prompt':
        convergence_count = n_primary_yellow_plus + min(1, len(support_active))
    else:
        convergence_count = n_yellow_plus

    if convergence_count >= 2:
        det, reason, conf = _apply_cap('AMBER', f"{combined_reason} [multi-channel convergence]", min(top_score, 0.60))
        if ch_window.sub_signals.get('mixed_signal') and ch_window.severity != 'GREEN':
            return 'MIXED', f"{reason} [windowed variance suggests hybrid text]", min(conf, 0.55), channel_details
        return det, reason, conf, channel_details

    # Supporting channels at AMBER in task_prompt mode
    if mode == 'task_prompt' and any(ch.sev_level >= 2 for ch in support_active):
        support_expl = [ch.explanation for ch in support_active if ch.sev_level >= 2]
        det, reason, conf = _apply_cap('AMBER', f"{' + '.join(support_expl)} [supporting channel]", 0.55)
        return det, reason, conf, channel_details

    # YELLOW: one channel at YELLOW+
    if n_yellow_plus >= 1:
        det, reason, conf = _apply_cap('YELLOW', combined_reason, min(top_score, 0.45))
        if ch_window.sub_signals.get('mixed_signal') and ch_window.severity != 'GREEN':
            return 'MIXED', f"{reason} [windowed variance suggests hybrid text]", min(conf, 0.50), channel_details
        return det, reason, conf, channel_details

    # Obfuscation delta
    if norm_report and norm_report.get('obfuscation_delta', 0) >= 0.05:
        delta = norm_report['obfuscation_delta']
        det, reason, conf = _apply_cap('YELLOW', f"Text normalization delta ({delta:.1%}) suggests obfuscation", 0.35)
        return det, reason, conf, channel_details

    # REVIEW: any channel has non-zero score
    any_signal = any(ch.score > 0.05 for ch in channels)
    if any_signal:
        weak_parts = [ch.explanation for ch in channels if ch.score > 0.05]
        review_reason = f"Weak signals below threshold: {' + '.join(weak_parts[:2])}"
        if disagreement:
            review_reason += ' [channels disagree]'
        return 'REVIEW', review_reason, 0.10, channel_details

    # GREEN
    return 'GREEN', 'No significant signals', 0.0, channel_details


# ==============================================================================
# PIPELINE ORCHESTRATION
# ==============================================================================
"""Full analysis pipeline orchestration."""


def analyze_prompt(text, task_id='', occupation='', attempter='', stage='',
                   run_l3=True, api_key=None, dna_provider='anthropic',
                   dna_model=None, dna_samples=3,
                   ground_truth=None, language=None, domain=None,
                   mode='auto', cal_table=None):
    """Run full v0.61 pipeline on a single prompt. Returns result dict."""
    # Normalization pre-pass
    normalized_text, norm_report = normalize_text(text)
    word_count_raw = len(text.split())
    word_count = len(normalized_text.split())

    # Fairness / language support gate
    lang_gate = check_language_support(normalized_text, word_count)

    text_for_analysis = normalized_text

    # Run all analyzers
    preamble_score, preamble_severity, preamble_hits, preamble_spans = run_preamble(text_for_analysis)
    fingerprint_score, fingerprint_hits, fingerprint_rate = run_fingerprint(text_for_analysis)
    prompt_sig = run_prompt_signature_enhanced(text_for_analysis)
    voice_dis = run_voice_dissonance_enhanced(text_for_analysis)
    instr_density = run_instruction_density_enhanced(
        text_for_analysis,
        constraint_active=(prompt_sig.get('pack_constraint_score', 0) > 0.08),
        schema_active=(voice_dis.get('pack_schema_score', 0) > 0.05),
    )

    self_sim = None
    if run_l3:
        self_sim = run_self_similarity(text_for_analysis)

    cont_result = None
    if run_l3 and api_key:
        cont_result = run_continuation_api(
            text_for_analysis, api_key=api_key, provider=dna_provider,
            model=dna_model, n_samples=dna_samples,
        )
    elif run_l3:
        cont_result = run_continuation_local_multi(text_for_analysis)

    semantic = run_semantic_resonance(text_for_analysis)
    tocsin = run_token_cohesiveness(text_for_analysis)
    ppl = run_perplexity(text_for_analysis)
    surprisal_traj = score_surprisal_windows(ppl.get('_token_losses', []))

    # Topic-scrubbed stylometry
    masked_text, mask_count = mask_topical_content(text_for_analysis)
    stylo_features = extract_stylometric_features(text_for_analysis, masked_text)

    # Windowed scoring
    window_result = score_windows(text_for_analysis)

    # Span-level explainability (diagnostic)
    _base_spans = collect_spans(text_for_analysis)

    # Evidence fusion
    det, reason, confidence, channel_details = determine(
        preamble_score, preamble_severity, prompt_sig, voice_dis, instr_density, word_count,
        self_sim=self_sim, cont_result=cont_result,
        lang_gate=lang_gate, norm_report=norm_report,
        mode=mode, fingerprint_score=fingerprint_score,
        semantic=semantic, ppl=ppl,
        window_result=window_result,
    )

    # Conformal calibration
    if word_count < 100:
        length_bin = 'short'
    elif word_count < 300:
        length_bin = 'medium'
    elif word_count < 800:
        length_bin = 'long'
    else:
        length_bin = 'very_long'

    cal_result = apply_calibration(confidence, cal_table, domain=domain, length_bin=length_bin)

    # Audit trail
    audit_trail = {
        'pipeline_version': 'v0.65',
        'mode_resolved': channel_details.get('mode', mode),
        'channels': channel_details.get('channels', {}),
        'fairness_gate': {
            'support_level': lang_gate.get('support_level'),
            'fw_coverage': lang_gate.get('function_word_coverage'),
        },
        'normalization': {
            'obfuscation_delta': norm_report.get('obfuscation_delta', 0),
            'invisible_chars': norm_report.get('invisible_chars', 0),
            'homoglyphs': norm_report.get('homoglyphs', 0),
            'ftfy_applied': norm_report.get('ftfy_applied', False),
        },
        'calibration': cal_result,
        'semantic_available': HAS_SEMANTIC,
        'perplexity_available': HAS_PERPLEXITY,
    }

    result = {
        'task_id': task_id,
        'occupation': occupation,
        'attempter': attempter,
        'stage': stage,
        'word_count': word_count,
        'word_count_raw': word_count_raw,
        'determination': det,
        'reason': reason,
        'confidence': confidence,
        'calibrated_confidence': cal_result['calibrated_confidence'],
        'confidence_quantile': cal_result['confidence_quantile'],
        'calibration_stratum': cal_result['stratum_used'],
        'mode': channel_details.get('mode', mode),
        'channel_details': channel_details,
        'audit_trail': audit_trail,
        # Normalization
        'norm_obfuscation_delta': norm_report.get('obfuscation_delta', 0.0),
        'norm_invisible_chars': norm_report.get('invisible_chars', 0),
        'norm_homoglyphs': norm_report.get('homoglyphs', 0),
        # Fairness gate
        'lang_support_level': lang_gate.get('support_level', 'SUPPORTED'),
        'lang_fw_coverage': lang_gate.get('function_word_coverage', 0.0),
        'lang_non_latin_ratio': lang_gate.get('non_latin_ratio', 0.0),
        # Preamble
        'preamble_score': preamble_score,
        'preamble_severity': preamble_severity,
        'preamble_hits': len(preamble_hits),
        'preamble_details': preamble_hits,
        # Fingerprint (diagnostic-only)
        'fingerprint_score': fingerprint_score,
        'fingerprint_hits': fingerprint_hits,
        # Prompt signature
        'prompt_signature_composite': prompt_sig['composite'],
        'prompt_signature_cfd': prompt_sig['cfd'],
        'prompt_signature_distinct_frames': prompt_sig['distinct_frames'],
        'prompt_signature_mfsr': prompt_sig['mfsr'],
        'prompt_signature_framing': prompt_sig['framing_completeness'],
        'prompt_signature_conditional_density': prompt_sig['conditional_density'],
        'prompt_signature_meta_design': prompt_sig['meta_design_hits'],
        'prompt_signature_contractions': prompt_sig['contractions'],
        'prompt_signature_must_rate': prompt_sig['must_rate'],
        'prompt_signature_numbered_criteria': prompt_sig['numbered_criteria'],
        # Instruction density
        'instruction_density_idi': instr_density['idi'],
        'instruction_density_imperatives': instr_density['imperatives'],
        'instruction_density_conditionals': instr_density['conditionals'],
        'instruction_density_binary_specs': instr_density['binary_specs'],
        'instruction_density_missing_refs': instr_density['missing_refs'],
        'instruction_density_flag_count': instr_density['flag_count'],
        # Voice dissonance
        'voice_dissonance_voice_score': voice_dis['voice_score'],
        'voice_dissonance_spec_score': voice_dis['spec_score'],
        'voice_dissonance_vsd': voice_dis['vsd'],
        'voice_dissonance_voice_gated': voice_dis['voice_gated'],
        'voice_dissonance_casual_markers': voice_dis['casual_markers'],
        'voice_dissonance_misspellings': voice_dis['misspellings'],
        'voice_dissonance_camel_cols': voice_dis['camel_cols'],
        'voice_dissonance_calcs': voice_dis['calcs'],
        'voice_dissonance_hedges': voice_dis['hedges'],
        # SSI
        'ssi_triggered': (
            voice_dis['spec_score'] >= (5.0 if voice_dis['contractions'] == 0 else 7.0)
            and voice_dis['voice_score'] < 0.5
            and voice_dis['hedges'] == 0
            and word_count >= 150
        ),
        # Metadata
        'ground_truth': ground_truth,
        'language': language,
        'domain': domain,
        # Windowed scoring
        'window_max_score': window_result.get('max_window_score', 0.0),
        'window_mean_score': window_result.get('mean_window_score', 0.0),
        'window_variance': window_result.get('window_variance', 0.0),
        'window_hot_span': window_result.get('hot_span_length', 0),
        'window_n_windows': window_result.get('n_windows', 0),
        'window_mixed_signal': window_result.get('mixed_signal', False),
        'window_fw_trajectory_cv': window_result.get('fw_trajectory_cv', 0.0),
        'window_comp_trajectory_mean': window_result.get('comp_trajectory_mean', 0.0),
        'window_comp_trajectory_cv': window_result.get('comp_trajectory_cv', 0.0),
        'window_changepoint': window_result.get('changepoint'),
        # Pack diagnostics
        'pack_constraint_score': prompt_sig.get('pack_constraint_score', 0.0),
        'pack_exec_spec_score': prompt_sig.get('pack_exec_spec_score', 0.0),
        'pack_schema_score': voice_dis.get('pack_schema_score', 0.0),
        'pack_active_families': prompt_sig.get('pack_active_families', 0),
        'pack_prompt_boost': prompt_sig.get('pack_boost', 0.0),
        'pack_idi_boost': instr_density.get('pack_idi_boost', 0.0),
        # Stylometric features
        'stylo_fw_ratio': stylo_features.get('function_word_ratio', 0.0),
        'stylo_sent_dispersion': stylo_features.get('sent_length_dispersion', 0.0),
        'stylo_ttr': stylo_features.get('type_token_ratio', 0.0),
        'stylo_avg_word_len': stylo_features.get('avg_word_length', 0.0),
        'stylo_short_word_ratio': stylo_features.get('short_word_ratio', 0.0),
        'stylo_mask_count': mask_count,
    }

    # Semantic resonance
    result.update({
        'semantic_resonance_ai_score': semantic.get('semantic_ai_score', 0.0),
        'semantic_resonance_human_score': semantic.get('semantic_human_score', 0.0),
        'semantic_resonance_ai_mean': semantic.get('semantic_ai_mean', 0.0),
        'semantic_resonance_human_mean': semantic.get('semantic_human_mean', 0.0),
        'semantic_resonance_delta': semantic.get('semantic_delta', 0.0),
        'semantic_resonance_determination': semantic.get('determination'),
        'semantic_resonance_confidence': semantic.get('confidence', 0.0),
    })

    # Perplexity
    result.update({
        'perplexity_value': ppl.get('perplexity', 0.0),
        'perplexity_determination': ppl.get('determination'),
        'perplexity_confidence': ppl.get('confidence', 0.0),
        'perplexity_surprisal_variance': ppl.get('surprisal_variance', 0.0),
        'perplexity_first_half_variance': ppl.get('first_half_variance', 0.0),
        'perplexity_second_half_variance': ppl.get('second_half_variance', 0.0),
        'perplexity_volatility_decay': ppl.get('volatility_decay', 1.0),
        'perplexity_n_tokens': ppl.get('n_tokens', 0),
        'perplexity_comp_ratio': ppl.get('comp_ratio', 0.0),
        'perplexity_zlib_normalized_ppl': ppl.get('zlib_normalized_ppl', 0.0),
        'perplexity_comp_ppl_ratio': ppl.get('comp_ppl_ratio', 0.0),
    })

    # Surprisal trajectory (FEAT 10)
    result.update({
        'surprisal_trajectory_cv': surprisal_traj.get('surprisal_trajectory_cv', 0.0),
        'surprisal_var_of_var': surprisal_traj.get('surprisal_var_of_var', 0.0),
        'surprisal_stationarity': surprisal_traj.get('surprisal_stationarity', 0.0),
        'surprisal_n_windows': surprisal_traj.get('n_surprisal_windows', 0),
    })

    # Token cohesiveness (TOCSIN, FEAT 8)
    result.update({
        'tocsin_cohesiveness': tocsin.get('cohesiveness', 0.0),
        'tocsin_cohesiveness_std': tocsin.get('cohesiveness_std', 0.0),
        'tocsin_determination': tocsin.get('determination'),
        'tocsin_confidence': tocsin.get('confidence', 0.0),
    })

    # Self-similarity (NSSI)
    if self_sim:
        result.update({
            'self_similarity_nssi_score': self_sim.get('nssi_score', 0.0),
            'self_similarity_nssi_signals': self_sim.get('nssi_signals', 0),
            'self_similarity_determination': self_sim.get('determination'),
            'self_similarity_confidence': self_sim.get('confidence', 0.0),
            'self_similarity_formulaic_density': self_sim.get('formulaic_density', 0.0),
            'self_similarity_power_adj_density': self_sim.get('power_adj_density', 0.0),
            'self_similarity_demonstrative_density': self_sim.get('demonstrative_density', 0.0),
            'self_similarity_transition_density': self_sim.get('transition_density', 0.0),
            'self_similarity_scare_quote_density': self_sim.get('scare_quote_density', 0.0),
            'self_similarity_emdash_density': self_sim.get('emdash_density', 0.0),
            'self_similarity_this_the_start_rate': self_sim.get('this_the_start_rate', 0.0),
            'self_similarity_section_depth': self_sim.get('section_depth', 0),
            'self_similarity_sent_length_cv': self_sim.get('sent_length_cv', 0.0),
            'self_similarity_comp_ratio': self_sim.get('comp_ratio', 0.0),
            'self_similarity_hapax_ratio': self_sim.get('hapax_ratio', 0.0),
            'self_similarity_hapax_count': self_sim.get('hapax_count', 0),
            'self_similarity_unique_words': self_sim.get('unique_words', 0),
            'self_similarity_shuffled_comp_ratio': self_sim.get('shuffled_comp_ratio', 0.0),
            'self_similarity_structural_compression_delta': self_sim.get('structural_compression_delta', 0.0),
        })
    else:
        result.update({
            'self_similarity_nssi_score': 0.0, 'self_similarity_nssi_signals': 0,
            'self_similarity_determination': None, 'self_similarity_confidence': 0.0,
            'self_similarity_formulaic_density': 0.0, 'self_similarity_power_adj_density': 0.0,
            'self_similarity_demonstrative_density': 0.0, 'self_similarity_transition_density': 0.0,
            'self_similarity_scare_quote_density': 0.0, 'self_similarity_emdash_density': 0.0,
            'self_similarity_this_the_start_rate': 0.0, 'self_similarity_section_depth': 0,
            'self_similarity_sent_length_cv': 0.0, 'self_similarity_comp_ratio': 0.0,
            'self_similarity_hapax_ratio': 0.0, 'self_similarity_hapax_count': 0,
            'self_similarity_unique_words': 0,
            'self_similarity_shuffled_comp_ratio': 0.0,
            'self_similarity_structural_compression_delta': 0.0,
        })

    # Continuation (DNA-GPT)
    if cont_result:
        proxy = cont_result.get('proxy_features', {})
        result.update({
            'continuation_bscore': cont_result.get('bscore', 0.0),
            'continuation_bscore_max': cont_result.get('bscore_max', 0.0),
            'continuation_determination': cont_result.get('determination'),
            'continuation_confidence': cont_result.get('confidence', 0.0),
            'continuation_n_samples': cont_result.get('n_samples', 0),
            'continuation_mode': 'local' if proxy else 'api',
            'continuation_ncd': proxy.get('ncd', 0.0),
            'continuation_internal_overlap': proxy.get('internal_overlap', 0.0),
            'continuation_cond_surprisal': proxy.get('cond_surprisal', 0.0),
            'continuation_repeat4': proxy.get('repeat4', 0.0),
            'continuation_ttr': proxy.get('ttr', 0.0),
            'continuation_composite': proxy.get('composite', 0.0),
            'continuation_composite_stability': proxy.get('composite_stability', 0.0),
            'continuation_composite_variance': proxy.get('composite_variance', 0.0),
            'continuation_improvement_rate': proxy.get('improvement_rate', 0.0),
            'continuation_ncd_matrix_mean': proxy.get('ncd_matrix_mean', 0.0),
            'continuation_ncd_matrix_variance': proxy.get('ncd_matrix_variance', 0.0),
        })
    else:
        result.update({
            'continuation_bscore': 0.0, 'continuation_bscore_max': 0.0,
            'continuation_determination': None, 'continuation_confidence': 0.0,
            'continuation_n_samples': 0, 'continuation_mode': None,
            'continuation_ncd': 0.0, 'continuation_internal_overlap': 0.0,
            'continuation_cond_surprisal': 0.0, 'continuation_repeat4': 0.0,
            'continuation_ttr': 0.0, 'continuation_composite': 0.0,
            'continuation_composite_stability': 0.0, 'continuation_composite_variance': 0.0,
            'continuation_improvement_rate': 0.0,
            'continuation_ncd_matrix_mean': 0.0, 'continuation_ncd_matrix_variance': 0.0,
        })

    # Span-level explainability — merge all span sources
    all_spans = list(_base_spans)
    all_spans.extend(preamble_spans)
    all_spans.extend(prompt_sig.get('pack_spans', []))
    all_spans.extend(voice_dis.get('pack_spans', []))
    all_spans.extend(instr_density.get('pack_spans', []))
    # Hot window sentence ranges
    for w in window_result.get('windows', []):
        if w['score'] >= 0.30:
            all_spans.append({
                'start_sentence': w['start'],
                'end_sentence': w['end'],
                'score': w['score'],
                'type': 'hot_window',
            })
    all_spans.sort(key=lambda s: s.get('start', s.get('start_sentence', 0)))
    result['_spans'] = all_spans
    result['detection_spans'] = all_spans

    return result


# ==============================================================================
# REPORTING: ATTEMPTER PROFILING
# ==============================================================================


def profile_attempters(results, min_submissions=2):
    """Aggregate detection results by attempter.

    Returns list of attempter profiles sorted by flag rate (descending).
    """
    by_attempter = {}
    for r in results:
        att = r.get('attempter', '').strip()
        if att:
            by_attempter.setdefault(att, []).append(r)

    profiles = []
    for att, submissions in by_attempter.items():
        if len(submissions) < min_submissions:
            continue

        det_counts = {}
        for r in submissions:
            d = r.get('determination', 'GREEN')
            det_counts[d] = det_counts.get(d, 0) + 1

        n_total = len(submissions)
        n_flagged = det_counts.get('RED', 0) + det_counts.get('AMBER', 0) + det_counts.get('MIXED', 0)
        flag_rate = n_flagged / n_total

        # Primary detection channel for flagged submissions
        flagged_channels = {}
        for r in submissions:
            if r.get('determination') in ('RED', 'AMBER', 'MIXED'):
                cd = r.get('channel_details', {}).get('channels', {})
                for ch_name, ch_info in cd.items():
                    if ch_info.get('severity') in ('RED', 'AMBER'):
                        flagged_channels[ch_name] = flagged_channels.get(ch_name, 0) + 1

        primary_channel = max(flagged_channels, key=flagged_channels.get) if flagged_channels else None

        flagged_confs = [r['confidence'] for r in submissions
                         if r.get('determination') in ('RED', 'AMBER', 'MIXED')]
        mean_conf = sum(flagged_confs) / len(flagged_confs) if flagged_confs else 0.0

        profiles.append({
            'attempter': att,
            'total_submissions': n_total,
            'flagged': n_flagged,
            'flag_rate': round(flag_rate, 3),
            'red': det_counts.get('RED', 0),
            'amber': det_counts.get('AMBER', 0),
            'yellow': det_counts.get('YELLOW', 0),
            'green': det_counts.get('GREEN', 0),
            'mixed': det_counts.get('MIXED', 0),
            'mean_flagged_confidence': round(mean_conf, 3),
            'primary_detection_channel': primary_channel,
            'occupations': list(set(r.get('occupation', '') for r in submissions if r.get('occupation'))),
        })

    profiles.sort(key=lambda p: (-p['flag_rate'], -p['flagged']))
    return profiles


def print_attempter_report(profiles):
    """Print attempter profiling summary to stdout."""
    if not profiles:
        print("\n  No attempter data available for profiling.")
        return

    flagged_profiles = [p for p in profiles if p['flagged'] > 0]
    total_submissions = sum(p['total_submissions'] for p in profiles)
    total_flagged = sum(p['flagged'] for p in profiles)

    print(f"\n{'='*90}")
    print(f"  ATTEMPTER PROFILING: {len(profiles)} contributors, "
          f"{total_submissions} submissions")
    print(f"{'='*90}")

    if flagged_profiles and total_flagged > 0:
        top_n = max(1, int(len(profiles) * 0.10))
        top_flagged = sum(p['flagged'] for p in flagged_profiles[:top_n])
        concentration = top_flagged / total_flagged * 100
        print(f"\n  Concentration: Top {top_n} contributor(s) account for "
              f"{concentration:.0f}% of all flagged submissions")

        print(f"\n  {'Attempter':<25} {'Subs':>5} {'Flag':>5} {'Rate':>7} "
              f"{'R':>3} {'A':>3} {'Y':>3} {'G':>3} {'Primary Channel':<20}")
        print(f"  {'-'*85}")

        for p in flagged_profiles:
            ch = p['primary_detection_channel'] or '-'
            print(f"  {p['attempter'][:24]:<25} {p['total_submissions']:>5} "
                  f"{p['flagged']:>5} {p['flag_rate']:>6.0%} "
                  f"{p['red']:>3} {p['amber']:>3} {p['yellow']:>3} {p['green']:>3} "
                  f"{ch:<20}")

    clean = [p for p in profiles if p['flagged'] == 0]
    if clean:
        print(f"\n  Clean contributors ({len(clean)}): "
              f"{', '.join(p['attempter'][:20] for p in clean[:10])}"
              f"{'...' if len(clean) > 10 else ''}")


# ==============================================================================
# REPORTING: FINANCIAL IMPACT
# ==============================================================================


def financial_impact(results, cost_per_prompt=400.0):
    """Calculate financial impact of detection.

    Returns dict with impact metrics.
    """
    n_total = len(results)
    det_counts = {}
    for r in results:
        d = r.get('determination', 'GREEN')
        det_counts[d] = det_counts.get(d, 0) + 1

    n_flagged = det_counts.get('RED', 0) + det_counts.get('AMBER', 0) + det_counts.get('MIXED', 0)
    flag_rate = n_flagged / max(n_total, 1)
    total_spend = n_total * cost_per_prompt
    waste_at_flag = n_flagged * cost_per_prompt
    clean_count = n_total - n_flagged
    clean_yield = clean_count / max(n_total, 1)
    annual_waste = waste_at_flag * 4
    annual_savings = annual_waste * 0.60

    return {
        'total_submissions': n_total,
        'total_spend': total_spend,
        'flagged_count': n_flagged,
        'flag_rate': round(flag_rate, 3),
        'waste_estimate': waste_at_flag,
        'clean_yield': round(clean_yield, 3),
        'clean_count': clean_count,
        'projected_annual_waste': annual_waste,
        'projected_annual_savings_60pct': annual_savings,
    }


def print_financial_report(impact, cost_per_prompt=400.0):
    """Print financial impact summary to stdout."""
    print(f"\n{'='*90}")
    print(f"  FINANCIAL IMPACT ESTIMATE (${cost_per_prompt:.0f}/prompt)")
    print(f"{'='*90}")
    print(f"    Total submissions:       {impact['total_submissions']:>8}")
    print(f"    Total spend:             ${impact['total_spend']:>10,.0f}")
    print(f"    Flagged (RED+AMBER):     {impact['flagged_count']:>8}  "
          f"({impact['flag_rate']:.1%})")
    print(f"    Estimated waste:         ${impact['waste_estimate']:>10,.0f}")
    print(f"    Clean yield:             {impact['clean_count']:>8}  "
          f"({impact['clean_yield']:.1%})")
    print(f"")
    print(f"    Projected annual waste:  ${impact['projected_annual_waste']:>10,.0f}  "
          f"(4 quarterly batches)")
    print(f"    Annual savings (60%):    ${impact['projected_annual_savings_60pct']:>10,.0f}  "
          f"(conservative catch rate)")


# ==============================================================================
# REPORTING: HTML REPORT GENERATOR
# ==============================================================================

_HTML_CSS = """
body { font-family: 'Segoe UI', system-ui, sans-serif; max-width: 900px;
       margin: 40px auto; padding: 0 20px; background: #fafafa; color: #1a1a1a; }
.header { border-bottom: 3px solid #1a1a1a; padding-bottom: 16px; margin-bottom: 24px; }
.det { font-size: 28px; font-weight: 700; }
.det-RED { color: #d32f2f; } .det-AMBER { color: #f57c00; }
.det-YELLOW { color: #fbc02d; } .det-GREEN { color: #388e3c; }
.det-MIXED { color: #1976d2; }
.meta { color: #666; font-size: 14px; margin-top: 8px; }
.text-container { background: white; border: 1px solid #e0e0e0; border-radius: 8px;
                  padding: 24px; line-height: 1.8; font-size: 15px; white-space: pre-wrap;
                  word-wrap: break-word; }
.signal { padding: 2px 0; border-bottom: 3px solid; cursor: help; }
.signal-CRITICAL { border-color: #ff1744; background: #ffebee; }
.signal-HIGH { border-color: #ff5722; background: #fbe9e7; }
.signal-MEDIUM { border-color: #ff9800; background: #fff3e0; }
.signal-LOW { border-color: #42a5f5; background: #e3f2fd; }
.signal-hot_window { border-color: #ef5350; background: #ffcdd2; }
.legend { margin-top: 24px; padding: 16px; background: #f5f5f5; border-radius: 8px;
          font-size: 13px; }
.legend span { display: inline-block; margin-right: 16px; }
.channels { margin-top: 24px; }
.ch-row { display: flex; align-items: center; padding: 8px 0;
          border-bottom: 1px solid #eee; font-size: 14px; }
.ch-name { width: 160px; font-weight: 600; }
.ch-sev { width: 80px; font-weight: 600; }
"""


def _apply_highlights(text, spans):
    """Apply highlight markup to text at span positions.

    Handles overlapping spans by using the highest-severity span at each position.
    """
    import html as _html
    if not spans:
        return _html.escape(text)

    severity_rank = {'CRITICAL': 4, 'HIGH': 3, 'MEDIUM': 2, 'LOW': 1}
    char_map = [None] * len(text)
    for span in spans:
        start = span.get('start', -1)
        end = span.get('end', start)
        if start < 0:
            continue
        sev = span.get('severity', span.get('type', 'LOW'))
        tooltip = span.get('pack', span.get('pattern', ''))
        rank = severity_rank.get(sev, 0)
        for i in range(max(0, start), min(end, len(text))):
            if char_map[i] is None or rank > char_map[i][2]:
                char_map[i] = (sev, tooltip, rank)

    out = []
    i = 0
    while i < len(text):
        if char_map[i] is None:
            j = i
            while j < len(text) and char_map[j] is None:
                j += 1
            out.append(_html.escape(text[i:j]))
            i = j
        else:
            sev, tooltip, rank = char_map[i]
            j = i
            while j < len(text) and char_map[j] is not None and char_map[j][:2] == (sev, tooltip):
                j += 1
            css_class = f"signal-{sev}"
            out.append(
                f'<span class="signal {css_class}" title="{_html.escape(tooltip)}">'
                f'{_html.escape(text[i:j])}</span>'
            )
            i = j

    return ''.join(out)


def generate_html_report(text, result, output_path=None):
    """Generate an HTML report with highlighted detection spans.

    Args:
        text: Original input text.
        result: Pipeline result dict (must include 'detection_spans' or '_spans').
        output_path: Where to write the HTML file. If None, returns HTML string.
    """
    import html as _html

    spans = result.get('detection_spans', result.get('_spans', []))
    det = result.get('determination', 'GREEN')
    reason = result.get('reason', '')
    confidence = result.get('confidence', 0)
    task_id = result.get('task_id', '')
    word_count = result.get('word_count', 0)

    char_spans = sorted(
        [s for s in spans if 'start' in s and 'end' in s],
        key=lambda s: s['start'],
    )

    highlighted = _apply_highlights(text, char_spans)

    cd = result.get('channel_details', {}).get('channels', {})
    channel_rows = []
    for ch_name in ['prompt_structure', 'stylometry', 'continuation', 'windowing']:
        info = cd.get(ch_name, {})
        sev = info.get('severity', 'GREEN')
        expl = info.get('explanation', '')[:80]
        channel_rows.append(
            f'<div class="ch-row">'
            f'<div class="ch-name">{ch_name}</div>'
            f'<div class="ch-sev det-{sev}">{sev}</div>'
            f'<div style="flex:1">{_html.escape(expl)}</div>'
            f'</div>'
        )

    n_char_spans = len(char_spans)
    n_hot = sum(1 for s in spans if s.get('type') == 'hot_window')
    span_summary = f"{n_char_spans} character spans"
    if n_hot:
        span_summary += f", {n_hot} hot windows"

    report = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>BEET Detection Report — {_html.escape(task_id)}</title>
<style>{_HTML_CSS}</style></head><body>
<div class="header">
    <div class="det det-{det}">{det}</div>
    <div class="meta">
        Task: {_html.escape(task_id)} | Words: {word_count} |
        Confidence: {confidence:.1%} |
        Mode: {result.get('mode', '?')} |
        Spans: {span_summary}
    </div>
    <div class="meta" style="margin-top:4px">{_html.escape(reason[:200])}</div>
</div>
<div class="text-container">{highlighted}</div>
<div class="legend">
    <strong>Legend:</strong>
    <span class="signal signal-CRITICAL">CRITICAL</span>
    <span class="signal signal-HIGH">HIGH</span>
    <span class="signal signal-MEDIUM">MEDIUM</span>
    <span class="signal signal-LOW">LOW (fingerprint/lexicon)</span>
</div>
<div class="channels">
    <h3>Channel Scores</h3>
    {''.join(channel_rows)}
</div>
</body></html>"""

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(report)
        return output_path

    return report


# ==============================================================================
# COMMAND-LINE INTERFACE
# ==============================================================================
"""Command-line interface for the LLM Detection Pipeline."""


def _is_frozen():
    """Check if running as a PyInstaller bundle."""
    return getattr(sys, 'frozen', False)


def print_result(r, verbose=False):
    """Pretty-print a single result."""
    icons = {'RED': '\U0001f534', 'AMBER': '\U0001f7e0', 'YELLOW': '\U0001f7e1',
             'GREEN': '\U0001f7e2', 'MIXED': '\U0001f535', 'REVIEW': '\u26aa'}
    icon = icons.get(r['determination'], '?')

    print(f"\n  {icon} [{r['determination']}] {r['task_id'][:20]}  |  {r['occupation'][:45]}")
    print(f"     Attempter: {r['attempter'] or '(unknown)'} | Stage: {r['stage']} | Words: {r['word_count']} | Mode: {r.get('mode', '?')}")
    print(f"     Reason: {r['reason']}")

    cal_conf = r.get('calibrated_confidence')
    cq_val = r.get('confidence_quantile')
    if cal_conf is not None and cal_conf != r.get('confidence'):
        cal_str = f"     Calibrated: conf={cal_conf:.3f}"
        if cq_val is not None:
            cal_str += f"  q={cq_val:.3f}"
        cal_str += f"  [{r.get('calibration_stratum', '?')}]"
        print(cal_str)

    if verbose or r['determination'] in ('RED', 'AMBER'):
        delta = r.get('norm_obfuscation_delta', 0)
        lang = r.get('lang_support_level', 'SUPPORTED')
        if delta > 0 or lang != 'SUPPORTED':
            print(f"     NORM obfuscation: {delta:.1%}  invisible={r.get('norm_invisible_chars', 0)} homoglyphs={r.get('norm_homoglyphs', 0)}")
            print(f"     GATE support:     {lang} (fw_coverage={r.get('lang_fw_coverage', 0):.2f}, non_latin={r.get('lang_non_latin_ratio', 0):.2f})")
        print(f"     Preamble:         {r['preamble_score']:.2f} ({r['preamble_severity']}, {r['preamble_hits']} hits)")
        if r['preamble_details']:
            for name, sev in r['preamble_details']:
                print(f"         -> [{sev}] {name}")
        print(f"     Fingerprints:     {r['fingerprint_score']:.2f} ({r['fingerprint_hits']} hits)")
        print(f"     Prompt Sig:       {r['prompt_signature_composite']:.2f}")
        print(f"         CFD={r['prompt_signature_cfd']:.3f} frames={r['prompt_signature_distinct_frames']} MFSR={r['prompt_signature_mfsr']:.3f}")
        print(f"         meta={r['prompt_signature_meta_design']} FC={r['prompt_signature_framing']}/3 must={r['prompt_signature_must_rate']:.3f}/sent")
        print(f"         contractions={r['prompt_signature_contractions']} numbered_criteria={r['prompt_signature_numbered_criteria']}")
        print(f"     IDI:              {r['instruction_density_idi']:.1f}  (imp={r['instruction_density_imperatives']} cond={r['instruction_density_conditionals']} Y/N={r['instruction_density_binary_specs']} MISS={r['instruction_density_missing_refs']} flag={r['instruction_density_flag_count']})")
        print(f"     VSD:              {r['voice_dissonance_vsd']:.1f}  (voice={r['voice_dissonance_voice_score']:.1f} x spec={r['voice_dissonance_spec_score']:.1f})")
        print(f"         gated={'YES' if r['voice_dissonance_voice_gated'] else 'no'} casual={r['voice_dissonance_casual_markers']} typos={r['voice_dissonance_misspellings']}")
        print(f"         cols={r['voice_dissonance_camel_cols']} calcs={r['voice_dissonance_calcs']} hedges={r['voice_dissonance_hedges']}")
        if r.get('ssi_triggered'):
            print(f"     SSI:  TRIGGERED  (spec={r['voice_dissonance_spec_score']:.1f}, voice=0, hedges=0, {r['word_count']}w)")
        nssi_score = r.get('self_similarity_nssi_score', 0.0)
        nssi_signals = r.get('self_similarity_nssi_signals', 0)
        nssi_det = r.get('self_similarity_determination')
        if nssi_score > 0 or nssi_det:
            det_str = nssi_det or 'n/a'
            print(f"     NSSI:             {nssi_score:.3f}  ({nssi_signals} signals, det={det_str})")
            print(f"         formulaic={r.get('self_similarity_formulaic_density', 0):.3f} power_adj={r.get('self_similarity_power_adj_density', 0):.3f}"
                  f" demo={r.get('self_similarity_demonstrative_density', 0):.3f} trans={r.get('self_similarity_transition_density', 0):.3f}")
            print(f"         sent_cv={r.get('self_similarity_sent_length_cv', 0):.3f} comp_ratio={r.get('self_similarity_comp_ratio', 0):.3f}"
                  f" hapax={r.get('self_similarity_hapax_ratio', 0):.3f} (unique={r.get('self_similarity_unique_words', 0)})")
        bscore = r.get('continuation_bscore', 0.0)
        dna_det = r.get('continuation_determination')
        if bscore > 0 or dna_det:
            det_str = dna_det or 'n/a'
            print(f"     DNA-GPT:          BScore={bscore:.4f}  (max={r.get('continuation_bscore_max', 0):.4f}, "
                  f"samples={r.get('continuation_n_samples', 0)}, det={det_str})")

        cd = r.get('channel_details', {})
        if cd.get('channels'):
            print(f"     -- Channels --")
            for ch_name, ch_info in cd['channels'].items():
                if ch_info['severity'] != 'GREEN':
                    eligible = 'Y' if ch_info.get('mode_eligible') else 'o'
                    print(f"     {eligible} {ch_name:18s} {ch_info['severity']:6s} score={ch_info['score']:.2f}  {ch_info['explanation'][:60]}")


def main():
    parser = argparse.ArgumentParser(description='LLM Detection Pipeline v0.61')
    parser.add_argument('input', nargs='?', help='Input file (.xlsx, .csv, or .pdf)')
    parser.add_argument('--gui', action='store_true', help='Launch desktop GUI mode')
    parser.add_argument('--text', help='Analyze a single text string')
    parser.add_argument('--sheet', help='Sheet name for xlsx files')
    parser.add_argument('--prompt-col', default='prompt', help='Column name containing prompts')
    parser.add_argument('--verbose', '-v', action='store_true', help='Show all layer details')
    parser.add_argument('--output', '-o', help='Output CSV path')
    parser.add_argument('--attempter', help='Filter by attempter name (substring match)')
    parser.add_argument('--no-similarity', action='store_true',
                        help='Skip cross-submission similarity analysis')
    parser.add_argument('--similarity-threshold', type=float, default=0.40,
                        help='Jaccard threshold for text similarity (default: 0.40)')
    parser.add_argument('--semantic-similarity', action='store_true',
                        help='Enable semantic embedding similarity (requires sentence-transformers)')
    parser.add_argument('--similarity-store', metavar='PATH',
                        help='JSONL path for cross-batch MinHash similarity store')
    parser.add_argument('--instructions', metavar='PATH',
                        help='Text file with shared instructions to factor out of similarity')
    parser.add_argument('--collect', metavar='PATH',
                        help='Append scored results to JSONL file for baseline accumulation')
    parser.add_argument('--analyze-baselines', metavar='JSONL',
                        help='Read accumulated JSONL and print per-occupation percentile tables')
    parser.add_argument('--baselines-csv', metavar='PATH',
                        help='Write baseline percentile tables to CSV (use with --analyze-baselines)')
    parser.add_argument('--no-layer3', action='store_true',
                        help='Skip Layer 3 entirely (NSSI + DNA-GPT)')
    parser.add_argument('--api-key', metavar='KEY',
                        help='API key for DNA-GPT continuation analysis. Falls back to '
                             'ANTHROPIC_API_KEY or OPENAI_API_KEY env var.')
    parser.add_argument('--provider', default='anthropic', choices=['anthropic', 'openai'],
                        help='LLM provider for DNA-GPT (default: anthropic)')
    parser.add_argument('--dna-model', metavar='MODEL',
                        help='Model name for DNA-GPT (default: auto per provider)')
    parser.add_argument('--dna-samples', type=int, default=3,
                        help='Number of regeneration samples for DNA-GPT (default: 3)')
    parser.add_argument('--mode', default='auto', choices=['task_prompt', 'generic_aigt', 'auto'],
                        help='Detection mode: task_prompt (prompt-structure primary), '
                             'generic_aigt (all channels), auto (heuristic). Default: auto')
    parser.add_argument('--calibrate', metavar='JSONL',
                        help='Build calibration table from labeled baseline JSONL and save to --cal-table')
    parser.add_argument('--cal-table', metavar='JSON',
                        help='Path to calibration table JSON (load for scoring, or save target for --calibrate)')
    parser.add_argument('--cost-per-prompt', type=float, default=400.0,
                        help='Cost per prompt for financial impact estimate (default: $400)')
    parser.add_argument('--html-report', metavar='DIR',
                        help='Generate HTML reports for flagged submissions in DIR')
    parser.add_argument('--memory', metavar='DIR', default=None,
                        help='Path to BEET memory store directory (enables cross-batch memory)')
    parser.add_argument('--confirm', nargs=3, metavar=('TASK_ID', 'LABEL', 'REVIEWER'),
                        help='Record ground truth confirmation: --confirm task_001 ai reviewer_A')
    parser.add_argument('--attempter-history', metavar='NAME',
                        help='Show historical profile for an attempter from memory store')
    parser.add_argument('--memory-summary', action='store_true',
                        help='Print memory store summary and exit')
    parser.add_argument('--rebuild-calibration', action='store_true',
                        help='Rebuild calibration from confirmed labels in memory store')
    args = parser.parse_args()

    if args.gui:
        launch_gui()
        return

    if not args.api_key:
        env_key = 'ANTHROPIC_API_KEY' if args.provider == 'anthropic' else 'OPENAI_API_KEY'
        args.api_key = os.environ.get(env_key)

    # Memory store initialization
    store = None
    if getattr(args, 'memory', None):
        store = MemoryStore(args.memory)

    # Standalone memory commands
    if getattr(args, 'memory_summary', False) and store:
        store.print_summary()
        return

    if getattr(args, 'confirm', None) and store:
        task_id, label, reviewer = args.confirm
        store.record_confirmation(task_id, label, verified_by=reviewer)
        return

    if getattr(args, 'attempter_history', None) and store:
        history = store.get_attempter_history(args.attempter_history)
        _print_attempter_history(history)
        return

    if getattr(args, 'rebuild_calibration', False) and store:
        cal = store.rebuild_calibration()
        return

    if args.analyze_baselines:
        if not os.path.exists(args.analyze_baselines):
            print(f"ERROR: File not found: {args.analyze_baselines}")
            return
        analyze_baselines(args.analyze_baselines, output_csv=args.baselines_csv)
        return

    if args.calibrate:
        if not os.path.exists(args.calibrate):
            print(f"ERROR: File not found: {args.calibrate}")
            return
        cal = calibrate_from_baselines(args.calibrate)
        if cal is None:
            print("ERROR: Insufficient labeled human data for calibration (need >=20)")
            return
        cal_path = args.cal_table or args.calibrate.replace('.jsonl', '_calibration.json')
        save_calibration(cal, cal_path)
        print(f"  Global quantiles: {cal['global']}")
        print(f"  Strata: {len(cal.get('strata', {}))} domain x length_bin tables")
        return

    cal_table = None
    if args.cal_table and os.path.exists(args.cal_table):
        cal_table = load_calibration(args.cal_table)
        print(f"Loaded calibration table: {cal_table['n_calibration']} records, "
              f"{len(cal_table.get('strata', {}))} strata")

    run_l3 = not args.no_layer3

    if args.text:
        result = analyze_prompt(
            args.text, run_l3=run_l3,
            api_key=args.api_key, dna_provider=args.provider,
            dna_model=args.dna_model, dna_samples=args.dna_samples,
            mode=args.mode, cal_table=cal_table,
        )
        print_result(result, verbose=True)
        return

    if not args.input:
        if _is_frozen():
            launch_gui()
            return
        parser.print_help()
        return

    ext = os.path.splitext(args.input)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        tasks = load_xlsx(args.input, sheet=args.sheet, prompt_col=args.prompt_col)
    elif ext == '.csv':
        tasks = load_csv(args.input, prompt_col=args.prompt_col)
    elif ext == '.pdf':
        tasks = load_pdf(args.input)
    else:
        print(f"ERROR: Unsupported file type: {ext}")
        return

    if not tasks:
        print("ERROR: No tasks found.")
        return

    if args.attempter:
        tasks = [t for t in tasks if args.attempter.lower() in t.get('attempter', '').lower()]
        print(f"Filtered to {len(tasks)} tasks matching attempter '{args.attempter}'")

    layer3_label = " + L3" if run_l3 else ""
    dna_label = " + DNA-GPT" if args.api_key else ""
    print(f"Processing {len(tasks)} tasks through pipeline v0.61{layer3_label}{dna_label}...")

    results = []
    text_map = {}
    for i, task in enumerate(tasks):
        r = analyze_prompt(
            task['prompt'],
            task_id=task.get('task_id', ''),
            occupation=task.get('occupation', ''),
            attempter=task.get('attempter', ''),
            stage=task.get('stage', ''),
            run_l3=run_l3,
            api_key=args.api_key,
            dna_provider=args.provider,
            dna_model=args.dna_model,
            dna_samples=args.dna_samples,
            mode=args.mode,
            cal_table=cal_table,
        )
        results.append(r)
        tid = task.get('task_id', f'_row{i}')
        text_map[tid] = task['prompt']
        if (i + 1) % 50 == 0:
            print(f"  Processed {i+1}/{len(tasks)}...")

    det_counts = Counter(r['determination'] for r in results)
    print(f"\n{'='*90}")
    print(f"  PIPELINE v0.61 RESULTS (n={len(results)})")
    print(f"{'='*90}")
    for det in ['RED', 'AMBER', 'YELLOW', 'GREEN']:
        ct = det_counts.get(det, 0)
        pct = ct / len(results) * 100
        icons = {'RED': '\U0001f534', 'AMBER': '\U0001f7e0', 'YELLOW': '\U0001f7e1', 'GREEN': '\U0001f7e2'}
        print(f"  {icons[det]} {det:>8}: {ct:>4} ({pct:.1f}%)")

    flagged = [r for r in results if r['determination'] in ('RED', 'AMBER')]
    if flagged:
        print(f"\n{'='*90}")
        print(f"  FLAGGED SUBMISSIONS: {len(flagged)}")
        print(f"{'='*90}")
        for r in sorted(flagged, key=lambda x: x['confidence'], reverse=True):
            print_result(r, verbose=args.verbose)

    yellow = [r for r in results if r['determination'] == 'YELLOW']
    if yellow:
        print(f"\n  YELLOW ({len(yellow)} minor signals):")
        for r in sorted(yellow, key=lambda x: x['confidence'], reverse=True)[:10]:
            print(f"    \U0001f7e1 {r['task_id'][:12]:12} {r['occupation'][:40]:40} | {r['reason'][:50]}")

    # Instruction template factoring (FEAT 15)
    instruction_shingles = None
    if getattr(args, 'instructions', None):
        with open(args.instructions, 'r') as _f:
            instruction_shingles = _word_shingles(_f.read())

    if not args.no_similarity and len(results) >= 2:
        sim_pairs = analyze_similarity(
            results, text_map,
            jaccard_threshold=args.similarity_threshold,
            semantic=getattr(args, 'semantic_similarity', False),
            instruction_shingles=instruction_shingles,
            similarity_store_path=getattr(args, 'similarity_store', None),
        )
        n_upgrades = apply_similarity_feedback(results, sim_pairs)
        print_similarity_report(sim_pairs)
        if n_upgrades:
            print(f"\n  {n_upgrades} determination(s) upgraded via similarity feedback")
    else:
        sim_pairs = []

    # Attempter profiling
    if len(results) >= 5:
        profiles = profile_attempters(results)
        print_attempter_report(profiles)

    # Financial impact
    if len(results) >= 10:
        impact = financial_impact(results, cost_per_prompt=args.cost_per_prompt)
        print_financial_report(impact, cost_per_prompt=args.cost_per_prompt)

    # HTML reports for flagged submissions
    if getattr(args, 'html_report', None) and flagged:
        os.makedirs(args.html_report, exist_ok=True)
        for r in flagged:
            tid = r.get('task_id', 'unknown')[:20]
            path = os.path.join(args.html_report, f"{tid}_{r['determination']}.html")
            generate_html_report(
                text_map.get(r.get('task_id', ''), ''), r, path)
        print(f"\n  HTML reports written to {args.html_report}/ ({len(flagged)} files)")

    default_name = os.path.basename(args.input).rsplit('.', 1)[0] + '_pipeline_v065.csv'
    input_dir = os.path.dirname(os.path.abspath(args.input))
    output_path = args.output or os.path.join(input_dir, default_name)

    flat = []
    for r in results:
        row = {k: v for k, v in r.items() if k != 'preamble_details'}
        row['preamble_details'] = str(r.get('preamble_details', []))
        flat.append(row)

    if sim_pairs:
        sim_lookup = defaultdict(list)
        for p in sim_pairs:
            sim_lookup[p['id_a']].append(f"{p['id_b']}(J={p['jaccard']:.2f})")
            sim_lookup[p['id_b']].append(f"{p['id_a']}(J={p['jaccard']:.2f})")
        for row in flat:
            tid = row.get('task_id', '')
            row['similarity_flags'] = '; '.join(sim_lookup.get(tid, []))

    pd.DataFrame(flat).to_csv(output_path, index=False)
    print(f"\n  Results saved to: {output_path}")

    if args.collect:
        collect_baselines(results, args.collect)

    # Memory store: cross-batch similarity + batch recording
    if store:
        cross_flags = store.cross_batch_similarity(results, text_map)
        if cross_flags:
            print(f"\n  CROSS-BATCH MEMORY: {len(cross_flags)} matches to previous submissions")
            for cf in cross_flags[:5]:
                print(f"    {cf['current_id'][:15]} <-> {cf['historical_id'][:15]} "
                      f"(MH={cf['minhash_similarity']:.2f})")
        store.record_batch(results, text_map)


def main_gui():
    """Entry point that always launches the GUI (for gui-scripts / executable)."""
    launch_gui()


if __name__ == '__main__':
    main()


# ==============================================================================
# DESKTOP GUI
# ==============================================================================
"""Desktop GUI for the LLM Detection Pipeline."""

import os
import threading
from collections import Counter




class DetectorGUI:
    """Simple desktop GUI for single-text and file analysis."""

    def __init__(self, root):
        self.root = root
        self.root.title("LLM Detector Pipeline v0.61")
        self.root.geometry("1040x760")

        self.file_var = tk.StringVar()
        self.prompt_col_var = tk.StringVar(value='prompt')
        self.sheet_var = tk.StringVar()
        self.attempter_var = tk.StringVar()
        self.provider_var = tk.StringVar(value='anthropic')
        self.api_key_var = tk.StringVar()
        self.status_var = tk.StringVar(value='Ready')

        self._build_layout()

    def _build_layout(self):
        frame = ttk.Frame(self.root, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        file_row = ttk.Frame(frame)
        file_row.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(file_row, text='Input file (CSV/XLSX):').pack(side=tk.LEFT)
        ttk.Entry(file_row, textvariable=self.file_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)
        ttk.Button(file_row, text='Browse', command=self._browse_file).pack(side=tk.LEFT)

        opts = ttk.Frame(frame)
        opts.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(opts, text='Prompt column').grid(row=0, column=0, sticky='w')
        ttk.Entry(opts, textvariable=self.prompt_col_var, width=18).grid(row=0, column=1, sticky='w', padx=6)
        ttk.Label(opts, text='Sheet (xlsx)').grid(row=0, column=2, sticky='w')
        ttk.Entry(opts, textvariable=self.sheet_var, width=16).grid(row=0, column=3, sticky='w', padx=6)
        ttk.Label(opts, text='Attempter filter').grid(row=0, column=4, sticky='w')
        ttk.Entry(opts, textvariable=self.attempter_var, width=18).grid(row=0, column=5, sticky='w', padx=6)

        l3 = ttk.LabelFrame(frame, text='Continuation Analysis (DNA-GPT)')
        l3.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(l3, text='Provider').grid(row=0, column=0, sticky='w', padx=6, pady=6)
        ttk.Combobox(l3, textvariable=self.provider_var, values=['anthropic', 'openai'], width=12, state='readonly').grid(row=0, column=1, sticky='w', pady=6)
        ttk.Label(l3, text='API Key (optional)').grid(row=0, column=2, sticky='w', padx=(16, 6), pady=6)
        ttk.Entry(l3, textvariable=self.api_key_var, show='*').grid(row=0, column=3, sticky='ew', padx=(0, 6), pady=6)
        l3.columnconfigure(3, weight=1)

        ttk.Label(frame, text='Single text input (optional):').pack(anchor='w')
        self.text_input = tk.Text(frame, height=10, wrap=tk.WORD)
        self.text_input.pack(fill=tk.BOTH, pady=(4, 8))

        actions = ttk.Frame(frame)
        actions.pack(fill=tk.X, pady=(0, 8))
        ttk.Button(actions, text='Analyze Text', command=lambda: self._run_async(self._analyze_text)).pack(side=tk.LEFT)
        ttk.Button(actions, text='Analyze File', command=lambda: self._run_async(self._analyze_file)).pack(side=tk.LEFT, padx=8)
        ttk.Button(actions, text='Clear Output', command=self._clear_output).pack(side=tk.LEFT)

        ttk.Label(frame, text='Results:').pack(anchor='w')
        self.output = tk.Text(frame, height=20, wrap=tk.WORD)
        self.output.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, textvariable=self.status_var).pack(anchor='w', pady=(8, 0))

    def _browse_file(self):
        path = filedialog.askopenfilename(filetypes=[('Data files', '*.csv *.xlsx *.xlsm'), ('All files', '*.*')])
        if path:
            self.file_var.set(path)

    def _clear_output(self):
        self.output.delete('1.0', tk.END)
        self.status_var.set('Ready')

    def _run_async(self, fn):
        self.status_var.set('Running...')

        def runner():
            try:
                fn()
                self.root.after(0, lambda: self.status_var.set('Done'))
            except Exception as exc:
                self.root.after(0, lambda: self.status_var.set('Error'))
                self.root.after(0, lambda: messagebox.showerror('Analysis Error', str(exc)))

        threading.Thread(target=runner, daemon=True).start()

    def _append(self, text):
        self.root.after(0, lambda: (self.output.insert(tk.END, text), self.output.see(tk.END)))

    def _analyze_text(self):
        text = self.text_input.get('1.0', tk.END).strip()
        if not text:
            self.root.after(0, lambda: messagebox.showinfo('Input required', 'Enter text to analyze.'))
            return
        result = analyze_prompt(
            text,
            run_l3=True,
            api_key=self.api_key_var.get().strip() or None,
            dna_provider=self.provider_var.get(),
        )
        self._append(self._format_result(result) + '\n')

    def _analyze_file(self):
        path = self.file_var.get().strip()
        if not path:
            self.root.after(0, lambda: messagebox.showinfo('Input required', 'Choose a CSV/XLSX file to analyze.'))
            return
        ext = os.path.splitext(path)[1].lower()
        if ext in ('.xlsx', '.xlsm'):
            tasks = load_xlsx(path, sheet=self.sheet_var.get().strip() or None, prompt_col=self.prompt_col_var.get().strip() or 'prompt')
        elif ext == '.csv':
            tasks = load_csv(path, prompt_col=self.prompt_col_var.get().strip() or 'prompt')
        else:
            self.root.after(0, lambda: messagebox.showerror('Unsupported file', f'Unsupported extension: {ext}'))
            return
        if self.attempter_var.get().strip():
            needle = self.attempter_var.get().strip().lower()
            tasks = [t for t in tasks if needle in t.get('attempter', '').lower()]
        if not tasks:
            self.root.after(0, lambda: messagebox.showinfo('No tasks', 'No qualifying prompts found.'))
            return

        api_key = self.api_key_var.get().strip() or None
        counts = Counter()
        for i, task in enumerate(tasks, 1):
            r = analyze_prompt(
                task['prompt'],
                task_id=task.get('task_id', ''),
                occupation=task.get('occupation', ''),
                attempter=task.get('attempter', ''),
                stage=task.get('stage', ''),
                run_l3=True,
                api_key=api_key,
                dna_provider=self.provider_var.get(),
            )
            counts[r['determination']] += 1
            self._append(f"[{i}/{len(tasks)}] {self._format_result(r)}\n")

        summary = (
            f"\nSummary: RED={counts.get('RED', 0)} | AMBER={counts.get('AMBER', 0)} "
            f"| YELLOW={counts.get('YELLOW', 0)} | GREEN={counts.get('GREEN', 0)}\n"
        )
        self._append(summary)

    @staticmethod
    def _format_result(result):
        return (
            f"{result.get('determination')} | conf={result.get('confidence', 0):.2f} | "
            f"words={result.get('word_count', 0)} | reason={result.get('reason', '')}"
        )


def launch_gui():
    """Launch Tkinter GUI mode."""
    if not HAS_TK:
        print('ERROR: tkinter is not available in this Python environment.')
        return
    root = tk.Tk()
    DetectorGUI(root)
    root.mainloop()


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == '__main__':
    main()
